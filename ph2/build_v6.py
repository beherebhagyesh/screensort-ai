"""Build chips_inventory_v6.xlsx — Ph1 (7 sheets) + Ph2 (5 new sheets).

Run from project root:
    /c/Puthon313/python ph2/build_v6.py
"""
import sys, json, re
from pathlib import Path
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
raw       = json.loads((ROOT / "ph1/packet_data.json").read_text(encoding="utf-8"))
merged_j  = json.loads((ROOT / "ph2/merged_data.json").read_text(encoding="utf-8"))
off_j     = json.loads((ROOT / "ph2/off_data.json").read_text(encoding="utf-8"))
serper_j  = json.loads((ROOT / "ph2/serper_data.json").read_text(encoding="utf-8"))
indiamart_j = json.loads((ROOT / "ph2/indiamart_data.json").read_text(encoding="utf-8"))
fssai_j   = json.loads((ROOT / "ph2/fssai_data.json").read_text(encoding="utf-8"))

enriched_skus = {e["packet_num"]: e for e in merged_j["enriched_skus"]}

# ── BRAND NORMALISATION ────────────────────────────────────────────────────────
def normalise_brand(b_en, b_lo):
    b = (b_en or b_lo or "").strip()
    bl = b.lower()
    if "haldiram" in bl: return "Balaji"
    if "balaji wafers" in bl: return "Balaji Wafers"
    if "balaji" in bl: return "Balaji"
    if "surya" in bl: return "Surya"
    lo = (b_lo or "").strip()
    if "सूर्य" in lo: return "Surya"
    if "बालाजी" in lo: return "Balaji"
    return b.title() or "Unknown"

def parse_ingredients(raw_str):
    if not raw_str: return []
    s = re.sub(r'\([^)]{0,60}\)', '', raw_str.strip())
    parts = re.split(r'[,;]', s)
    cleaned = []
    for p in parts:
        p = p.strip().strip('*').strip()
        if len(p) > 1 and not re.match(r'^[\d\.\s%]+$', p):
            cleaned.append(p.title())
    return [x for x in cleaned if x]

def ingredients_as_list(raw_str):
    items = parse_ingredients(raw_str)
    return "\n".join(f"• {i}" for i in items) if items else ""

# ── DATA CLEANING ─────────────────────────────────────────────────────────────
data = []
for d in raw:
    b_en = d.get("brand_english", "") or ""
    b_lo = d.get("brand_local", "") or ""
    brand = normalise_brand(b_en, b_lo)
    cleaned = dict(d)
    cleaned["brand"] = brand
    try: cleaned["mrp_inr"] = float(d.get("mrp_inr") or 5)
    except: cleaned["mrp_inr"] = 5.0
    try: cleaned["net_weight_g"] = float(d.get("net_weight_g"))
    except: cleaned["net_weight_g"] = None
    data.append(cleaned)

N = len(data)

# ── TRADE PRICING ─────────────────────────────────────────────────────────────
TRADE_DEFAULT = {
    "Balaji":       {"mfr": 4.25, "dist": 4.50},
    "Balaji Wafers":{"mfr": 4.25, "dist": 4.50},
    "Surya":        {"mfr": 3.50, "dist": 4.00},
}
def get_trade(brand, product_name):
    pn = (product_name or "").lower()
    if "mung" in pn or "moong" in pn or "dal" in pn:
        return {"mfr": 4.50, "dist": 4.75}
    return TRADE_DEFAULT.get(brand, {"mfr": None, "dist": None})

# ── STYLES ────────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)
THIN  = Border(**{s: Side(style="thin",   color="D0D0D0") for s in "left right top bottom".split()})
MED   = Border(**{s: Side(style="medium", color="2E75B6") for s in "left right top bottom".split()})
WR    = Alignment(wrap_text=True, vertical="top")
CTR   = Alignment(horizontal="center", vertical="top", wrap_text=True)

H1F=fill("1F4E79"); H1T=fnt(True,"FFFFFF",9)
H2F=fill("2E75B6"); H2T=fnt(True,"FFFFFF",9)
ALTF=fill("EBF3FB")
DATAF=fnt(size=9); CALCF=fnt(color="000000",size=9)
INPF=fnt(color="0000FF",size=9)
WARNF=fill("FFF2CC"); GREENF=fill("E2EFDA"); REDF=fill("FFE0E0")
PH2_H1F=fill("1A3C5E"); PH2_H2F=fill("28699A")  # darker blue for Ph2 sheets

def hdr1(ws, r, c, v, span=1, bg=None):
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    x = ws.cell(r, c, v)
    x.fill = bg or H1F; x.font=H1T; x.alignment=CTR; x.border=MED

def hdr2(ws, r, c, v):
    x = ws.cell(r, c, v); x.fill=H2F; x.font=H2T; x.alignment=CTR; x.border=THIN

def wr(ws, r, c, v, f=None, a=None, fl=None):
    x = ws.cell(r, c, v); x.border=THIN
    x.font = f or DATAF; x.alignment = a or WR
    if fl: x.fill = fl
    return x

def rfill(ri): return ALTF if ri % 2 == 0 else fill("FFFFFF")
def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTER INVENTORY
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Master Inventory"
ws1.freeze_panes = "D3"
ws1.row_dimensions[1].height = 16
ws1.row_dimensions[2].height = 36

secs = [(1,3,"IDENTIFICATION"),(4,6,"PRICING & WEIGHT"),(7,10,"MANUFACTURER"),
        (11,12,"COMPLIANCE"),(13,14,"PRODUCT"),(15,17,"PACK COLORS"),
        (18,21,"PACK DESIGN"),(22,23,"SCORES"),(24,25,"COMPETITIVE INTEL")]
for cs, ce, title in secs:
    hdr1(ws1, 1, cs, title, span=ce-cs+1)

c2 = ["#","Brand","Product","MRP (Rs)","Net Wt (g)","Price/g (Rs)",
      "Manufacturer","City","State","Phone","FSSAI License","Best Before (mo)",
      "Type","Flavor / Variant","Primary Color","Secondary Color","Accent Colors",
      "Pack Format","Front Imagery","Design Style","Font Style",
      "Appeal /10","Shelf Vis /10","Claims on Pack","Competitive Notes"]
for ci, h in enumerate(c2, 1): hdr2(ws1, 2, ci, h)
set_widths(ws1, [4,16,32,9,9,10,26,14,14,14,20,10,18,20,16,16,18,14,24,18,14,12,12,32,36])

for ri, d in enumerate(data, 3):
    rf = rfill(ri)
    brand = d.get("brand","?"); mrp = d.get("mrp_inr"); wt = d.get("net_weight_g")
    prod_en = d.get("product_name_english","") or ""
    prod_lo = d.get("product_name_local","") or ""
    prod = prod_en + (f"\n[{prod_lo}]" if prod_lo and prod_lo != prod_en else "")
    ws1.row_dimensions[ri].height = 32

    vals = [
        d.get("packet_num"), brand, prod,
        mrp, wt,
        f"=D{ri}/E{ri}" if (mrp and wt) else None,
        d.get("manufacturer_name"), d.get("manufacturer_city"), d.get("manufacturer_state"), d.get("manufacturer_phone"),
        d.get("fssai_license"), d.get("best_before_months"),
        d.get("product_type"), d.get("variant_flavor"),
        d.get("pack_primary_color"), d.get("pack_secondary_color"), d.get("pack_accent_colors"),
        d.get("pack_format"), d.get("front_imagery"), d.get("design_style"), d.get("font_style"),
        d.get("pack_appeal_score"), d.get("shelf_visibility_score"),
        d.get("claims_on_pack"), d.get("competitive_notes"),
    ]
    for ci, v in enumerate(vals, 1):
        a = CTR if ci in {1,4,5,6,12,22,23} else WR
        f = CALCF if ci == 6 else INPF if ci in {4,5} else DATAF
        fl = WARNF if (ci == 5 and v is None) else rf
        c = wr(ws1, ri, ci, v, f=f, a=a, fl=fl)
        if ci == 6 and v: c.number_format = "0.000"
        if ci == 4 and v: c.number_format = '"Rs"#,##0.00'

ws1.auto_filter.ref = f"A2:{get_column_letter(len(c2))}2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — INGREDIENTS
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Ingredients")
ws2.freeze_panes = "C3"
ws2.row_dimensions[1].height = 16
ws2.row_dimensions[2].height = 36

hdr1(ws2, 1, 1, "PRODUCT", span=2)
hdr1(ws2, 1, 3, "INGREDIENTS — ITEMIZED LIST (one per line)", span=2)
hdr1(ws2, 1, 5, "ALLERGENS & DIETARY")
hdr1(ws2, 1, 6, "INGREDIENT COUNT")

cols2 = ["#","Product","Ingredients (itemized)","Raw text (original)","Allergens","# Ingredients"]
for ci, h in enumerate(cols2, 1): hdr2(ws2, 2, ci, h)
set_widths(ws2, [4, 30, 48, 52, 24, 12])

all_ingredients = []

for ri, d in enumerate(data, 3):
    rf = rfill(ri)
    raw_ing = d.get("ingredients_full","") or ""
    parsed = parse_ingredients(raw_ing)
    all_ingredients.extend(parsed)
    itemized = ingredients_as_list(raw_ing)
    ws2.row_dimensions[ri].height = max(30, len(parsed) * 12)

    wr(ws2, ri, 1, d.get("packet_num"), a=CTR, fl=rf)
    wr(ws2, ri, 2, d.get("product_name_english") or d.get("product_name_local"), fl=rf)
    x = wr(ws2, ri, 3, itemized, fl=rf)
    x.alignment = Alignment(wrap_text=True, vertical="top")
    wr(ws2, ri, 4, raw_ing, f=fnt(size=8, italic=True, color="666666"), fl=rf)
    wr(ws2, ri, 5, d.get("allergens"), fl=rf)
    wr(ws2, ri, 6, len(parsed) if parsed else None, a=CTR, fl=rf)

ws2.auto_filter.ref = "A2:F2"

sep_row = N + 5
ws2.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=6)
ws2.cell(sep_row, 1, "INGREDIENT FREQUENCY ANALYSIS — How often each ingredient appears across all 19 SKUs")
ws2.cell(sep_row, 1).font = fnt(True, "1F4E79", 11)
ws2.cell(sep_row, 1).fill = fill("D6E4F0")

freq_row = sep_row + 2
hdr2(ws2, freq_row, 1, "Ingredient")
hdr2(ws2, freq_row, 2, "SKU Count")
hdr2(ws2, freq_row, 3, "% of SKUs")
hdr2(ws2, freq_row, 4, "Category")
set_widths(ws2, [4, 30, 48, 52, 24, 12])

counter = Counter(i.lower() for i in all_ingredients)
CATEGORIES = {
    "edible oil": "Fat/Oil", "palm": "Fat/Oil", "palmolein": "Fat/Oil",
    "salt": "Seasoning", "sugar": "Sweetener",
    "rice": "Grain/Starch", "wheat": "Grain/Starch", "corn": "Grain/Starch",
    "maize": "Grain/Starch", "potato": "Vegetable", "tapioca": "Grain/Starch",
    "dal": "Pulse/Legume", "gram": "Pulse/Legume", "soy": "Pulse/Legume", "pea": "Pulse/Legume",
    "chilli": "Spice", "pepper": "Spice", "turmeric": "Spice", "cumin": "Spice",
    "coriander": "Spice", "amchur": "Spice", "mango": "Spice",
    "acid": "Acidity Reg.", "citric": "Acidity Reg.",
    "flavour": "Flavouring", "flavor": "Flavouring", "natural": "Flavouring",
    "colour": "Colour", "color": "Colour",
    "iodised": "Salt type", "iodized": "Salt type",
}
def categorize(ing):
    il = ing.lower()
    for k, cat in CATEGORIES.items():
        if k in il: return cat
    return "Other"

r = freq_row + 1
for ing, count in counter.most_common(40):
    pct = count / N * 100
    rf2 = GREENF if count >= 10 else (WARNF if count >= 5 else fill("FFFFFF"))
    ws2.cell(r, 1, ing.title()).fill = rf2; ws2.cell(r,1).border=THIN; ws2.cell(r,1).font=DATAF; ws2.cell(r,1).alignment=WR
    ws2.cell(r, 2, count).fill = rf2; ws2.cell(r,2).border=THIN; ws2.cell(r,2).font=CALCF; ws2.cell(r,2).alignment=CTR
    ws2.cell(r, 3, pct/100).fill = rf2; ws2.cell(r,3).border=THIN; ws2.cell(r,3).font=CALCF
    ws2.cell(r, 3).number_format = "0%"
    ws2.cell(r, 4, categorize(ing)).fill = rf2; ws2.cell(r,4).border=THIN; ws2.cell(r,4).font=DATAF
    r += 1

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — NUTRITION ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Nutrition Analysis")
ws3.freeze_panes = "B3"

hdr1(ws3, 1, 1, "NUTRITION ANALYSIS — All values per 100g", span=7)
nh = ["#","Product","Brand","Calories\n(kcal/100g)","Sodium\n(mg/100g)","Fat (g)\n/100g","Carbs (g)\n/100g"]
for ci, h in enumerate(nh, 1): hdr2(ws3, 2, ci, h)
set_widths(ws3, [4, 38, 16, 14, 14, 12, 12])

for ri, d in enumerate(data, 3):
    rf = rfill(ri)
    cal = d.get("calories_per_100g"); sod = d.get("sodium_mg"); fat = d.get("fat_g"); carb = d.get("carbs_g")
    wr(ws3, ri, 1, d.get("packet_num"), a=CTR, fl=rf)
    wr(ws3, ri, 2, d.get("product_name_english"), fl=rf)
    wr(ws3, ri, 3, d.get("brand"), fl=rf)
    c = wr(ws3, ri, 4, cal, a=CTR, fl=REDF if (cal and cal > 570) else (GREENF if (cal and cal < 480) else rf))
    c = wr(ws3, ri, 5, sod, a=CTR, fl=REDF if (sod and sod > 900) else (GREENF if (sod and sod < 400) else rf))
    wr(ws3, ri, 6, fat, a=CTR, fl=rf)
    wr(ws3, ri, 7, carb, a=CTR, fl=rf)
    ws3.row_dimensions[ri].height = 28

ws3.auto_filter.ref = "A2:G2"

# Add ranking tables
sep = N + 5
labels = ["TOP 5 HIGHEST CALORIES", "TOP 5 LOWEST CALORIES", "TOP 5 HIGHEST SODIUM", "TOP 5 LOWEST SODIUM"]
keys   = ["calories_per_100g", "calories_per_100g", "sodium_mg", "sodium_mg"]
revs   = [True, False, True, False]
fills2 = [REDF, GREENF, REDF, GREENF]

col_off = 1
for label, key, rev, rfill2 in zip(labels, keys, revs, fills2):
    sorted_d = sorted([x for x in data if x.get(key) is not None], key=lambda x: x[key], reverse=rev)[:5]
    ws3.merge_cells(start_row=sep, start_column=col_off, end_row=sep, end_column=col_off+2)
    ws3.cell(sep, col_off, label).font = fnt(True, "1F4E79", 10)
    ws3.cell(sep, col_off).fill = fill("D6E4F0")
    hdr2(ws3, sep+1, col_off, "Product"); hdr2(ws3, sep+1, col_off+1, "Brand"); hdr2(ws3, sep+1, col_off+2, key.replace("_"," ").title())
    for ri2, d2 in enumerate(sorted_d, sep+2):
        for ci2, v2 in enumerate([d2.get("product_name_english"), d2.get("brand"), d2.get(key)], col_off):
            ws3.cell(ri2, ci2, v2).font = DATAF; ws3.cell(ri2,ci2).fill=rfill2; ws3.cell(ri2,ci2).border=THIN; ws3.cell(ri2,ci2).alignment=WR
    col_off += 4

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — DESIGN ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Design Analysis")
ws4.freeze_panes = "C3"

hdr1(ws4, 1, 1, "PACK DESIGN ANALYSIS", span=8)
dh = ["#","Product","Brand","Primary Color","Secondary Color","Pack Format","Appeal /10","Shelf Vis /10"]
for ci, h in enumerate(dh, 1): hdr2(ws4, 2, ci, h)
set_widths(ws4, [4,32,16,18,18,18,12,12])

for ri, d in enumerate(data, 3):
    rf = rfill(ri)
    vals = [d.get("packet_num"), d.get("product_name_english"), d.get("brand"),
            d.get("pack_primary_color"), d.get("pack_secondary_color"), d.get("pack_format"),
            d.get("pack_appeal_score"), d.get("shelf_visibility_score")]
    for ci, v in enumerate(vals, 1):
        a = CTR if ci in {1,7,8} else WR
        wr(ws4, ri, ci, v, a=a, fl=rf)
    ws4.row_dimensions[ri].height = 28

ws4.auto_filter.ref = "A2:H2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — TRADE MARGINS
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Trade Margins")
ws5.freeze_panes = "C3"

hdr1(ws5, 1, 1, "TRADE MARGIN ANALYSIS — Rs 5 MRP", span=9)
mh = ["#","Product","Brand","MRP (Rs)","Mfr Price (Rs)","Dist Price (Rs)",
      "Retailer Margin (Rs)","Retailer Margin %","Note"]
for ci, h in enumerate(mh, 1): hdr2(ws5, 2, ci, h)
set_widths(ws5, [4,32,16,10,14,14,16,16,36])

for ri, d in enumerate(data, 3):
    rf = rfill(ri)
    brand = d.get("brand","")
    trade = get_trade(brand, d.get("product_name_english",""))
    mrp = d.get("mrp_inr",5)
    mfr = trade.get("mfr"); dist = trade.get("dist")

    pn = (d.get("product_name_english","") or "").lower()
    is_mungdal = "mung" in pn or "moong" in pn or "dal" in pn
    row_fill = REDF if is_mungdal else rf

    wr(ws5, ri, 1, d.get("packet_num"), a=CTR, fl=row_fill)
    wr(ws5, ri, 2, d.get("product_name_english"), fl=row_fill)
    wr(ws5, ri, 3, brand, fl=row_fill)
    wr(ws5, ri, 4, mrp, f=INPF, a=CTR, fl=row_fill)
    wr(ws5, ri, 5, mfr, f=INPF, a=CTR, fl=row_fill)
    wr(ws5, ri, 6, dist, f=INPF, a=CTR, fl=row_fill)
    if dist is not None:
        ret_margin = mrp - dist
        ret_pct = ret_margin / mrp
        wr(ws5, ri, 7, ret_margin, f=CALCF, a=CTR, fl=row_fill)
        c = wr(ws5, ri, 8, ret_pct, f=CALCF, a=CTR, fl=row_fill)
        c.number_format = "0.0%"
    else:
        wr(ws5, ri, 7, None, fl=row_fill)
        wr(ws5, ri, 8, None, fl=row_fill)
    note = "⚠ Lowest margin SKU — 5% retailer margin. Least pushed at retail." if is_mungdal else ""
    wr(ws5, ri, 9, note, f=fnt(bold=is_mungdal, color="CC0000" if is_mungdal else "000000"), fl=row_fill)
    ws5.row_dimensions[ri].height = 28

for ci in [4,5,6,7]:
    for ri in range(3, N+3):
        ws5.cell(ri, ci).number_format = '"Rs"0.00'

ws5.auto_filter.ref = "A2:I2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — MANUFACTURER REGISTER
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Manufacturer Register")
ws6.freeze_panes = "B3"

hdr1(ws6, 1, 1, "MANUFACTURER REGISTER — Unique entities with FSSAI licenses", span=7)
mfh = ["FSSAI License","Manufacturer Name","City","State","Phone","SKUs (count)","Products"]
for ci, h in enumerate(mfh, 1): hdr2(ws6, 2, ci, h)
set_widths(ws6, [22, 36, 16, 14, 18, 10, 60])

mfr_index = {}
for d in data:
    key = (d.get("fssai_license","") or "UNKNOWN", d.get("manufacturer_name","") or "Unknown")
    if key not in mfr_index:
        mfr_index[key] = {"city": d.get("manufacturer_city"), "state": d.get("manufacturer_state"),
                          "phone": d.get("manufacturer_phone"), "products": []}
    pname = d.get("product_name_english") or d.get("product_name_local") or "?"
    mfr_index[key]["products"].append(pname)

for ri, ((lic, name), v) in enumerate(mfr_index.items(), 3):
    rf = rfill(ri)
    wr(ws6, ri, 1, lic, fl=rf)
    wr(ws6, ri, 2, name, fl=rf)
    wr(ws6, ri, 3, v["city"], fl=rf)
    wr(ws6, ri, 4, v["state"], fl=rf)
    wr(ws6, ri, 5, v["phone"], fl=rf)
    wr(ws6, ri, 6, len(v["products"]), a=CTR, fl=rf)
    wr(ws6, ri, 7, "\n".join(f"• {p}" for p in v["products"]), fl=rf)
    ws6.row_dimensions[ri].height = max(30, len(v["products"]) * 14)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — PM SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("PM Summary")
ws7.column_dimensions["A"].width = 22
ws7.column_dimensions["B"].width = 80

hdr1(ws7, 1, 1, "PM COMPETITIVE RESEARCH — SUMMARY DASHBOARD", span=2)

stats_title_row = 2
ws7.cell(3, 1, "INVENTORY OVERVIEW").font = fnt(True, "1F4E79", 10)
ws7.cell(3, 1).fill = fill("D6E4F0")
ws7.merge_cells("A3:B3")

stats = [
    ("Total SKUs inventoried", N),
    ("Brands covered", len(set(d["brand"] for d in data))),
    ("All MRP", "Rs 5"),
    ("Weight range", f"{min(d['net_weight_g'] for d in data if d.get('net_weight_g')):.0f}g — {max(d['net_weight_g'] for d in data if d.get('net_weight_g')):.0f}g"),
    ("Unique FSSAI licenses", len(set(d.get('fssai_license') for d in data if d.get('fssai_license')))),
    ("Unique manufacturers", len(mfr_index)),
    ("Calories range", f"{min(d['calories_per_100g'] for d in data if d.get('calories_per_100g')):.0f} — {max(d['calories_per_100g'] for d in data if d.get('calories_per_100g')):.0f} kcal/100g"),
    ("Sodium range", f"{min(d['sodium_mg'] for d in data if d.get('sodium_mg')):.0f} — {max(d['sodium_mg'] for d in data if d.get('sodium_mg')):.0f} mg/100g"),
]
for r2, (k, v) in enumerate(stats, 4):
    rf = rfill(r2)
    ws7.cell(r2, 1, k).font=DATAF; ws7.cell(r2,1).fill=rf; ws7.cell(r2,1).border=THIN
    ws7.cell(r2, 2, str(v)).font=DATAF; ws7.cell(r2,2).fill=rf; ws7.cell(r2,2).border=THIN

r = 4 + len(stats) + 2
ws7[f"A{r}"] = "KEY COMPETITIVE INSIGHTS"; ws7[f"A{r}"].font = fnt(True, size=11, color="1F4E79")
ws7.merge_cells(f"A{r}:B{r}")
insights = [
    ("Margin war", "Surya gives retailer 2x the margin (20% vs 10%) — retailers push Surya harder. To compete, target ≥18% retailer margin."),
    ("Mung Dal weakness", "Mung Dal has the lowest retailer margin (5%) — least pushed product. Opportunity to capture with better margin to retailer."),
    ("Edible Oil dominates", "Edible Vegetable Oil / Palmolein is the universal base ingredient across virtually all SKUs — not a differentiator."),
    ("Calories cluster", "Most products 500-570 kcal/100g — tight band. Energy density is not a differentiator in this category."),
    ("Sodium variation", "Sodium varies widely (300-1300mg/100g) — high sodium products are a health risk; low-sodium is an unclaimed positioning."),
    ("Color gap", "Dominant colors: Yellow, Red, Orange. Blue/White/Premium metallic packs = 0 at this price point — shelf standout opportunity."),
    ("Claims gap", "Zero products claim: 'high protein', 'baked', 'no MSG', 'clean label', 'high fiber' — open positioning at Rs5."),
    ("Marathi presence", "Both Surya & Balaji use Devanagari/Marathi on pack — local script essential for this market."),
    ("Weight/Value play", "Balaji 16g packs = Rs 0.31/g vs 24g packs at Rs 0.21/g — 48% worse value. Launch 25g at Rs5 to own value positioning."),
    ("Surya positioning", "Surya leans into Maharashtra-specific themes (Ladki Bahin scheme) — hyper-local cultural marketing at Rs5 is unique."),
]
r += 1
hdr2(ws7, r, 1, "Theme"); hdr2(ws7, r, 2, "Insight")
r += 1
for theme, insight in insights:
    rf = rfill(r)
    ws7.cell(r, 1, theme).font = fnt(True, size=9); ws7.cell(r,1).fill=rf; ws7.cell(r,1).border=THIN; ws7.cell(r,1).alignment=WR
    ws7.cell(r, 2, insight).font = fnt(size=9); ws7.cell(r,2).fill=rf; ws7.cell(r,2).border=THIN; ws7.cell(r,2).alignment=WR
    ws7.row_dimensions[r].height = 30
    r += 1

set_widths(ws7, [22, 80])

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — ONLINE PRESENCE (PH2)
# ═══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Online Presence")
ws8.freeze_panes = "C3"

hdr1(ws8, 1, 1, "PH2 — ONLINE PRESENCE INTELLIGENCE", span=9, bg=PH2_H1F)
oh = ["#","Product","Brand","Presence\nScore /4","Google\nShopping","IndiaMART\nB2B","Open Food\nFacts","FSSAI\nVerified","Top Online\nSource / Price"]
for ci, h in enumerate(oh, 1):
    x = ws8.cell(2, ci, h); x.fill=PH2_H2F; x.font=H2T; x.alignment=CTR; x.border=THIN
set_widths(ws8, [4, 38, 16, 10, 12, 12, 10, 10, 40])

for ri, d in enumerate(data, 3):
    pnum = d.get("packet_num")
    e = enriched_skus.get(pnum, {})
    score = e.get("online_presence_score", 0)
    rf = GREENF if score >= 3 else (WARNF if score >= 2 else (REDF if score == 1 else fill("FFD0D0")))

    wr(ws8, ri, 1, pnum, a=CTR, fl=rf)
    wr(ws8, ri, 2, d.get("product_name_english"), fl=rf)
    wr(ws8, ri, 3, d.get("brand"), fl=rf)
    wr(ws8, ri, 4, f"{score}/4", a=CTR, fl=rf)
    wr(ws8, ri, 5, "YES" if e.get("serper_online_presence") else "NO", a=CTR,
       fl=GREENF if e.get("serper_online_presence") else REDF)
    wr(ws8, ri, 6, "YES" if e.get("indiamart_present") else "NO", a=CTR,
       fl=GREENF if e.get("indiamart_present") else REDF)
    wr(ws8, ri, 7, "YES" if e.get("off_matched") else "NO", a=CTR,
       fl=GREENF if e.get("off_matched") else REDF)
    wr(ws8, ri, 8, "YES" if e.get("fssai_verified") else "MANUAL", a=CTR,
       fl=WARNF if not e.get("fssai_verified") else GREENF)
    price = e.get("serper_price")
    src   = e.get("serper_source", "")
    src_short = (src or "").replace("https://","").split("/")[0] if src else ""
    top_info = f"{src_short}: {price}" if (src_short and price) else (src_short or "—")
    wr(ws8, ri, 9, top_info, fl=rf)
    ws8.row_dimensions[ri].height = 28

ws8.auto_filter.ref = "A2:I2"

# Summary note below
note_r = N + 5
ws8.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=9)
sum_s = merged_j.get("summary", {})
note_txt = (
    f"Avg Online Presence Score: {sum_s.get('avg_online_presence_score','?')}/4  |  "
    f"Google Shopping hits: {serper_j.get('summary',{}).get('skus_with_online_presence','?')}/19  |  "
    f"OFF matches: {sum_s.get('off_matched','?')}/19  |  "
    f"IndiaMART: Balaji + Surya both confirmed  |  "
    f"FSSAI: {sum_s.get('fssai_verified','?')} verified via API (rest need manual check at foscos.fssai.gov.in)"
)
ws8.cell(note_r, 1, note_txt).font = fnt(italic=True, color="444444", size=8)
ws8.cell(note_r, 1).fill = fill("EBF3FB")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 9 — B2B INTELLIGENCE (IndiaMART)
# ═══════════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("B2B Intelligence")
ws9.freeze_panes = "B3"

hdr1(ws9, 1, 1, "PH2 — INDIAMART B2B INTELLIGENCE", span=5, bg=PH2_H1F)
bh = ["Company / Supplier","Location","Price Mention","Products Listed","Source URL"]
for ci, h in enumerate(bh, 1):
    x = ws9.cell(2, ci, h); x.fill=PH2_H2F; x.font=H2T; x.alignment=CTR; x.border=THIN
set_widths(ws9, [36, 18, 20, 50, 40])

suppliers = indiamart_j.get("suppliers", [])
ri = 3
for s in suppliers:
    rf = rfill(ri)
    company = s.get("company","") or s.get("raw_text","")[:60]
    location = s.get("location","")
    price = s.get("price","")
    raw_text = s.get("raw_text","")
    # Extract product names from raw_text (first 2 lines)
    products = raw_text[:200].replace("\n"," | ")
    src_url = s.get("source_url","")
    wr(ws9, ri, 1, company, fl=rf)
    wr(ws9, ri, 2, location, fl=rf)
    wr(ws9, ri, 3, price, a=CTR, fl=rf)
    wr(ws9, ri, 4, products, fl=rf)
    wr(ws9, ri, 5, src_url, f=fnt(size=8, color="0000FF"), fl=rf)
    ws9.row_dimensions[ri].height = 28
    ri += 1

# Surya finding note
ri += 1
ws9.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=5)
surya_note = (
    "SURYA GRUH UDHYOG — CONFIRMED ON INDIAMART: Found via 'surya gruh udhyog' search. "
    f"Appears in B2B listings. {indiamart_j.get('notes','')}"
)
c = ws9.cell(ri, 1, surya_note)
c.font = fnt(bold=True, color="1F4E79", size=9)
c.fill = fill("D6E4F0")
ws9.row_dimensions[ri].height = 40

# IndiaMART catalogue products
ri += 2
ws9.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=5)
ws9.cell(ri, 1, "CATALOGUE PRODUCTS FOUND ON INDIAMART").font = fnt(True, "1F4E79", 10)
ws9.cell(ri, 1).fill = fill("D6E4F0")
ri += 1
hdr2(ws9, ri, 1, "Product Name Found"); hdr2(ws9, ri, 2, "URL")
ri += 1
for p in indiamart_j.get("catalogue_products", []):
    rf = rfill(ri)
    wr(ws9, ri, 1, p.get("name",""), fl=rf)
    wr(ws9, ri, 2, p.get("url",""), f=fnt(size=8, color="0000FF"), fl=rf)
    ri += 1

# Price mentions
ri += 1
ws9.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=5)
ws9.cell(ri, 1, "PRICE MENTIONS EXTRACTED FROM INDIAMART PAGES").font = fnt(True, "1F4E79", 10)
ws9.cell(ri, 1).fill = fill("D6E4F0")
ri += 1
for pm in indiamart_j.get("price_mentions", [])[:20]:
    rf = rfill(ri)
    wr(ws9, ri, 1, pm, fl=rf)
    ri += 1

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 10 — PRODUCT DATABASE (Open Food Facts)
# ═══════════════════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("Product DB (OFF)")
ws10.freeze_panes = "C3"

hdr1(ws10, 1, 1, "PH2 — OPEN FOOD FACTS DATABASE MATCHES", span=7, bg=PH2_H1F)
ph = ["#","Product","Brand","Match\nType","OFF Product Name","Barcode\n(EAN)","Verified\nIngredients (sample)"]
for ci, h in enumerate(ph, 1):
    x = ws10.cell(2, ci, h); x.fill=PH2_H2F; x.font=H2T; x.alignment=CTR; x.border=THIN
set_widths(ws10, [4, 36, 16, 14, 36, 16, 60])

off_by_pnum = {}
for s in off_j.get("sku_matches", []):
    off_by_pnum[s.get("ph1_packet_num")] = s

for ri, d in enumerate(data, 3):
    pnum = d.get("packet_num")
    off = off_by_pnum.get(pnum, {})
    matched = off.get("match_found", False)
    rf = GREENF if matched else fill("FFFFFF")

    wr(ws10, ri, 1, pnum, a=CTR, fl=rf)
    wr(ws10, ri, 2, d.get("product_name_english"), fl=rf)
    wr(ws10, ri, 3, d.get("brand"), fl=rf)
    wr(ws10, ri, 4, off.get("match_type","—") if matched else "No match", a=CTR, fl=rf)
    wr(ws10, ri, 5, off.get("off_product_name","—") if matched else "—", fl=rf)
    wr(ws10, ri, 6, off.get("off_barcode","") if matched else "", a=CTR, fl=rf)
    ing_sample = (off.get("off_ingredients","") or "")[:120]
    wr(ws10, ri, 7, ing_sample, f=fnt(size=8), fl=rf)
    ws10.row_dimensions[ri].height = 28

ws10.auto_filter.ref = "A2:G2"

# OFF catalogue discoveries
sep_r = N + 5
ws10.merge_cells(start_row=sep_r, start_column=1, end_row=sep_r, end_column=7)
ws10.cell(sep_r, 1, f"OFF CATALOGUE DISCOVERIES — {len(off_j.get('catalogue_discoveries',[]))} Balaji products found on Open Food Facts (not in our 19 SKUs)").font = fnt(True, "1F4E79", 10)
ws10.cell(sep_r, 1).fill = fill("D6E4F0")
sep_r += 1
for ci2, h2 in enumerate(["OFF Product Name","Barcode","Quantity","Brand","OFF URL"], 1):
    x = ws10.cell(sep_r, ci2, h2); x.fill=PH2_H2F; x.font=H2T; x.alignment=CTR; x.border=THIN
sep_r += 1
for disc in off_j.get("catalogue_discoveries", [])[:30]:
    rf = rfill(sep_r)
    wr(ws10, sep_r, 1, disc.get("product_name",""), fl=rf)
    wr(ws10, sep_r, 2, disc.get("code",""), fl=rf)
    wr(ws10, sep_r, 3, disc.get("quantity",""), fl=rf)
    wr(ws10, sep_r, 4, disc.get("brand",""), fl=rf)
    wr(ws10, sep_r, 5, disc.get("url",""), f=fnt(size=8, color="0000FF"), fl=rf)
    sep_r += 1

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 11 — GOOGLE SHOPPING (Serper)
# ═══════════════════════════════════════════════════════════════════════════════
ws11 = wb.create_sheet("Google Shopping")
ws11.freeze_panes = "C3"

hdr1(ws11, 1, 1, "PH2 — GOOGLE SHOPPING INTELLIGENCE (via Serper.dev)", span=7, bg=PH2_H1F)
gh = ["#","Product","Brand","Results\nCount","Top Online\nPrice","Top Source","Rating /\nReviews"]
for ci, h in enumerate(gh, 1):
    x = ws11.cell(2, ci, h); x.fill=PH2_H2F; x.font=H2T; x.alignment=CTR; x.border=THIN
set_widths(ws11, [4, 38, 16, 10, 16, 32, 20])

serper_by_pnum = {s.get("ph1_packet_num"): s for s in serper_j.get("sku_results", [])}

for ri, d in enumerate(data, 3):
    pnum = d.get("packet_num")
    sr = serper_by_pnum.get(pnum, {})
    present = sr.get("online_presence", False)
    rf = GREENF if present else fill("FFFFFF")
    results = sr.get("results", [])
    top = results[0] if results else {}

    wr(ws11, ri, 1, pnum, a=CTR, fl=rf)
    wr(ws11, ri, 2, d.get("product_name_english"), fl=rf)
    wr(ws11, ri, 3, d.get("brand"), fl=rf)
    wr(ws11, ri, 4, sr.get("results_count", 0), a=CTR, fl=rf)
    price_str = top.get("price","") or "—"
    wr(ws11, ri, 5, price_str, a=CTR, fl=rf)
    src = (top.get("source","") or top.get("link","") or "—")
    src_short = src.replace("https://","").split("/")[0] if src.startswith("http") else src
    wr(ws11, ri, 6, src_short[:40], fl=rf)
    rating = top.get("rating",""); reviews = top.get("ratingCount","") or top.get("reviews","")
    rating_str = f"{rating} ({reviews} reviews)" if rating else "—"
    wr(ws11, ri, 7, rating_str, a=CTR, fl=rf)
    ws11.row_dimensions[ri].height = 28

ws11.auto_filter.ref = "A2:G2"

# Summary
s_sum = serper_j.get("summary", {})
note_r2 = N + 5
ws11.merge_cells(start_row=note_r2, start_column=1, end_row=note_r2, end_column=7)
ws11.cell(note_r2, 1,
    f"Google Shopping Summary: {s_sum.get('skus_with_online_presence','?')}/19 SKUs found online  |  "
    f"Total results: {s_sum.get('total_results_found','?')}  |  "
    f"Surya online presence: {'YES — confirmed' if s_sum.get('surya_online_presence') else 'NO'}  |  "
    f"Data via Serper.dev Google Shopping API (gl=in)"
).font = fnt(italic=True, color="444444", size=8)
ws11.cell(note_r2, 1).fill = fill("EBF3FB")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 12 — SCRAPE AUDIT
# ═══════════════════════════════════════════════════════════════════════════════
ws12 = wb.create_sheet("Scrape Audit")
ws12.column_dimensions["A"].width = 22
ws12.column_dimensions["B"].width = 20
ws12.column_dimensions["C"].width = 18
ws12.column_dimensions["D"].width = 16
ws12.column_dimensions["E"].width = 55

hdr1(ws12, 1, 1, "PH2 — WEB SCRAPING AUDIT LOG", span=5, bg=PH2_H1F)
ah = ["Source","Method","Status","Records\nFound","Notes / Findings"]
for ci, h in enumerate(ah, 1):
    x = ws12.cell(2, ci, h); x.fill=PH2_H2F; x.font=H2T; x.alignment=CTR; x.border=THIN

audit_rows = [
    (
        "Open Food Facts",
        "REST API (free)",
        "✓ SUCCESS",
        f"{len(off_j.get('sku_matches',[]))} SKU checks\n{len(off_j.get('catalogue_discoveries',[]))} new products",
        "Barcode match: 1 confirmed (Mung Dal). 50 Balaji catalogue products discovered. "
        "Name-search false positives (returned Sting Energy) — only barcode matches are reliable."
    ),
    (
        "FSSAI / FoSCoS",
        "REST API",
        "⚠ PARTIAL",
        f"{len(fssai_j.get('licenses_found',[]))} licenses",
        "API returns HTTP 200 but no structured JSON without CAPTCHA. "
        "5 unique licenses extracted from packet data. Manual lookup needed at foscos.fssai.gov.in."
    ),
    (
        "IndiaMART (v1)",
        "requests + BeautifulSoup",
        "✗ FAILED",
        "0",
        "Pages loaded HTTP 200 but content is JS-rendered SPA. BeautifulSoup extracts no product data. "
        "Replaced by Playwright version."
    ),
    (
        "IndiaMART (v2)",
        "Playwright (headless Chromium)",
        "✓ SUCCESS",
        f"{len(indiamart_j.get('suppliers',[]))} suppliers\n{len(indiamart_j.get('catalogue_products',[]))} products",
        f"Surya Gruh Udhyog CONFIRMED on IndiaMART. "
        f"{indiamart_j.get('pages_success',0)}/{indiamart_j.get('pages_attempted',0)} pages scraped. "
        f"17 price mentions extracted. Surya company page returns 404 (no B2B profile) but appears in search results."
    ),
    (
        "Google Shopping",
        "Serper.dev API (gl=in)",
        "✓ SUCCESS",
        f"{s_sum.get('total_results_found','?')} results\nacross 19 SKUs",
        f"{s_sum.get('skus_with_online_presence','?')}/19 SKUs found online. "
        "Surya confirmed with online presence. "
        "Most Balaji Rs5 packs not on major ecomm (Amazon/BigBasket) — "
        "found on niche B2C stores. API key: Serper.dev (22 queries consumed)."
    ),
]

for ri2, (src, method, status, records, notes) in enumerate(audit_rows, 3):
    rf = GREENF if "SUCCESS" in status else (WARNF if "PARTIAL" in status else REDF)
    wr(ws12, ri2, 1, src, f=fnt(bold=True, size=9), fl=rf)
    wr(ws12, ri2, 2, method, fl=rf)
    wr(ws12, ri2, 3, status, a=CTR, fl=rf)
    wr(ws12, ri2, 4, records, a=CTR, fl=rf)
    wr(ws12, ri2, 5, notes, f=fnt(size=8), fl=rf)
    ws12.row_dimensions[ri2].height = 50

# FSSAI license table
r_f = 3 + len(audit_rows) + 2
ws12.merge_cells(start_row=r_f, start_column=1, end_row=r_f, end_column=5)
ws12.cell(r_f, 1, "FSSAI LICENSES — MANUAL LOOKUP REQUIRED AT foscos.fssai.gov.in").font = fnt(True, "1F4E79", 10)
ws12.cell(r_f, 1).fill = fill("D6E4F0")
r_f += 1
for ci2, h2 in enumerate(["FSSAI License", "Likely Entity", "SKUs Count", "Manual Lookup URL","Status"], 1):
    x = ws12.cell(r_f, ci2, h2); x.fill=PH2_H2F; x.font=H2T; x.alignment=CTR; x.border=THIN
r_f += 1
lic_map = {}
for d in data:
    lic = d.get("fssai_license","")
    brand = d.get("brand","")
    if lic:
        if lic not in lic_map: lic_map[lic] = {"brand": brand, "count": 0}
        lic_map[lic]["count"] += 1

for lic, v in lic_map.items():
    rf = rfill(r_f)
    wr(ws12, r_f, 1, lic, fl=rf)
    wr(ws12, r_f, 2, v["brand"], fl=rf)
    wr(ws12, r_f, 3, v["count"], a=CTR, fl=rf)
    wr(ws12, r_f, 4, f"https://foscos.fssai.gov.in/", f=fnt(size=8, color="0000FF"), fl=rf)
    wr(ws12, r_f, 5, "Needs manual check", fl=WARNF)
    r_f += 1

# ── SAVE ──────────────────────────────────────────────────────────────────────
out = ROOT / "chips_inventory_v6.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: {wb.sheetnames}")
print(f"SKUs: {N} | New discovered: {len(merged_j.get('new_skus_discovered',[]))}")
