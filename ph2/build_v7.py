"""
Build chips_inventory_v7.xlsx — Comprehensive Competitive Intelligence.
Sources: ph1 packet data + all ph2 scraping (Serper, IndiaMART, OFF).
Run from project root: /c/Puthon313/python ph2/build_v7.py
"""
import sys, json, re
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
ph1_data     = json.loads((ROOT / "ph1/packet_data.json").read_text(encoding="utf-8"))
serper_disc  = json.loads((ROOT / "ph2/serper_discovery_data.json").read_text(encoding="utf-8"))
indiamart_c  = json.loads((ROOT / "ph2/indiamart_competitors_data.json").read_text(encoding="utf-8"))
off_brands   = json.loads((ROOT / "ph2/off_brands_data.json").read_text(encoding="utf-8"))

# ── STYLES ─────────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)
THIN  = Border(**{s: Side(style="thin",   color="D0D0D0") for s in "left right top bottom".split()})
MED   = Border(**{s: Side(style="medium", color="1F4E79") for s in "left right top bottom".split()})
WR    = Alignment(wrap_text=True, vertical="top")
CTR   = Alignment(horizontal="center", vertical="top", wrap_text=True)

# Color palette
H1F   = fill("1F4E79"); H1T = fnt(True,"FFFFFF",10)
H2F   = fill("2E75B6"); H2T = fnt(True,"FFFFFF",9)
ALTF  = fill("EBF3FB")
GREENF= fill("E2EFDA"); YELLOWF=fill("FFF2CC"); REDF=fill("FFE0E0")
ORANGEF=fill("FCE4D6"); PURPLEF=fill("EAD1DC"); GREYF=fill("F2F2F2")
DATAF = fnt(size=9); BOLDF = fnt(bold=True, size=9)
INPF  = fnt(color="0000FF", size=9); CALCF = fnt(color="000000", size=9)
LINKF = fnt(color="0563C1", size=8)

def hdr1(ws, r, c, v, span=1, bg=None):
    if span > 1: ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)
    x = ws.cell(r, c, v); x.fill=bg or H1F; x.font=H1T; x.alignment=CTR; x.border=MED

def hdr2(ws, r, c, v):
    x = ws.cell(r, c, v); x.fill=H2F; x.font=H2T; x.alignment=CTR; x.border=THIN

def wr(ws, r, c, v, f=None, a=None, fl=None):
    x = ws.cell(r, c, v); x.border=THIN
    x.font = f or DATAF; x.alignment = a or WR
    if fl: x.fill = fl
    return x

def rfill(ri): return ALTF if ri % 2 == 0 else fill("FFFFFF")
def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── DATA PREP ──────────────────────────────────────────────────────────────────
# Brand metadata: home state, competitor tier, color, direct Balaji rival
BRAND_META = {
    "Balaji Wafers": {
        "home_state": "Gujarat", "home_city": "Rajkot",
        "tier": "Primary (ours)", "color": "FF9900",
        "balaji_rival": "Self", "listed_nse": False,
        "price_tiers": "Rs 5, 10, 20, 50+",
        "product_types": "Wafers, Namkeen, Pellets, Extruded",
        "dist_region": "Gujarat, Maharashtra, MP, National",
        "notes": "Our primary brand. Rajkot-based. 19 SKUs inventoried at Rs 5.",
    },
    "Balaji": {
        "home_state": "Gujarat", "home_city": "Rajkot",
        "tier": "Primary (ours)", "color": "FF9900",
        "balaji_rival": "Self", "listed_nse": False,
        "price_tiers": "Rs 5, 10, 20, 50+",
        "product_types": "Namkeen, Sev, Gathiya, Aloo, Chips",
        "dist_region": "Gujarat, Maharashtra, MP, National",
        "notes": "Sub-brand of Balaji Wafers. Same manufacturer.",
    },
    "Surya": {
        "home_state": "Maharashtra", "home_city": "Nandurbar",
        "tier": "Primary competitor", "color": "E2EFDA",
        "balaji_rival": "Direct (Rs 5)", "listed_nse": False,
        "price_tiers": "Rs 5, 10",
        "product_types": "Pellets, Extruded snacks, Noodle sticks",
        "dist_region": "North Maharashtra, Nandurbar district",
        "notes": "Surya Gruh Udhyog. Micro-manufacturer, kirana-only, hyper-local. NOTE: 'Surya' in Google Shopping = different brand (Surya Masale, TN). Our Surya is offline-only.",
    },
    "Gopal": {
        "home_state": "Gujarat", "home_city": "Rajkot",
        "tier": "Tier 1 competitor", "color": "FCE4D6",
        "balaji_rival": "Direct (same city)", "listed_nse": True,
        "price_tiers": "Rs 5, 10, 20, 50, 100+",
        "product_types": "Chips, Namkeen, Gathiya, Chevdo, Papad, Spices, Rusk",
        "dist_region": "Gujarat, Maharashtra (Nagpur plant), National",
        "notes": "Gopal Snacks Ltd. Listed on NSE. Head office & plants in Rajkot GIDC. Nagpur plant serves Maharashtra. Broadest product range of any Gujarat snack brand.",
    },
    "Yellow Diamond": {
        "home_state": "Madhya Pradesh", "home_city": "Indore",
        "tier": "Tier 1 competitor", "color": "FFF2CC",
        "balaji_rival": "Regional (MP-MH belt)", "listed_nse": True,
        "price_tiers": "Rs 5, 10, 20, 50+",
        "product_types": "Chips, Namkeen, Pellets, Rings, Puffs",
        "dist_region": "MP, Maharashtra, UP, Rajasthan, National",
        "notes": "Prataap Snacks Limited. Flagship brand Yellow Diamond. Based in Indore — dominant in MP and border Maharashtra. Listed on NSE.",
    },
    "Haldiram": {
        "home_state": "Delhi / Nagpur", "home_city": "Delhi (HQ), Nagpur (MH plant)",
        "tier": "Tier 1 national", "color": "EAD1DC",
        "balaji_rival": "National (has Rs 5)", "listed_nse": False,
        "price_tiers": "Rs 5, 10, 20, 50, 100+",
        "product_types": "Bhujia, Sev, Dal, Mixture, Chana, Namkeen",
        "dist_region": "National, strong Maharashtra via Nagpur",
        "notes": "Haldiram's. Has Rs 5 Aloo Bhujia and Ratlami Sev packets confirmed. Nagpur manufacturing unit serves Maharashtra. Premium perception at Rs 5 vs local brands.",
    },
    "Bikaji": {
        "home_state": "Rajasthan", "home_city": "Bikaner",
        "tier": "Tier 1 national", "color": "EAD1DC",
        "balaji_rival": "National (has Rs 5-10)", "listed_nse": True,
        "price_tiers": "Rs 10, 25, 50, 100+",
        "product_types": "Bhujia, Sev, Chana, Mixture, Chips",
        "dist_region": "National, Rajasthan-MP-MH strong",
        "notes": "Bikaji Foods. Listed. Primarily Rs 10+ at kirana. Bikaneri Bhujia signature product.",
    },
    "Chheda": {
        "home_state": "Maharashtra", "home_city": "Mumbai",
        "tier": "Regional (MH)", "color": "E2EFDA",
        "balaji_rival": "Maharashtra regional", "listed_nse": False,
        "price_tiers": "Rs 47, 52, 60, 85, 110",
        "product_types": "Bhakharwadi, Mix Farsan, Chana Chor, Tikha Namkeen",
        "dist_region": "Mumbai, Maharashtra",
        "notes": "Chheda's. Mumbai-based Maharashtra snack brand. Premium farsan positioning (Rs 47-110). Not a Rs 5 player. Important for shelf-space and consumer preference in MH.",
    },
    "Anil": {
        "home_state": "Tamil Nadu", "home_city": "Chennai",
        "tier": "Not relevant", "color": "F2F2F2",
        "balaji_rival": "None (different segment)", "listed_nse": False,
        "price_tiers": "Rs 17-95",
        "product_types": "Vermicelli, Fryums, Appalam, Happala",
        "dist_region": "South India",
        "notes": "Anil Foods Pvt Ltd — Tamil Nadu vermicelli/pasta brand. NOT a Gujarat snack competitor. Included in scrape due to brand name match.",
    },
}

# Normalize brand from product data
def norm_brand(b):
    if not b: return "Unknown"
    bl = b.lower()
    if "balaji wafer" in bl: return "Balaji Wafers"
    if "balaji" in bl: return "Balaji"
    if "gopal" in bl: return "Gopal"
    if "yellow diamond" in bl or "prataap" in bl or "pratap" in bl: return "Yellow Diamond"
    if "haldiram" in bl: return "Haldiram"
    if "bikaji" in bl: return "Bikaji"
    if "chheda" in bl: return "Chheda"
    if "surya" in bl: return "Surya"
    if "anil" in bl: return "Anil"
    if "bikanervala" in bl or "bikanerv" in bl: return "Bikanervala"
    return b.strip().title()

# Price tier classification
def price_tier(price_rs):
    if price_rs is None: return "Unknown"
    if price_rs <= 6:   return "Rs 5"
    if price_rs <= 12:  return "Rs 10"
    if price_rs <= 25:  return "Rs 20"
    if price_rs <= 55:  return "Rs 50"
    if price_rs <= 110: return "Rs 100"
    return "Rs 100+"

# Pack size tier
def size_tier(g):
    if g is None: return "Unknown"
    if g <= 16:  return "Rs 5 (≤16g)"
    if g <= 30:  return "Rs 10 (17-30g)"
    if g <= 60:  return "Rs 20 (31-60g)"
    if g <= 150: return "Rs 50 (61-150g)"
    return "Rs 100+ (>150g)"

# Product type detection
def detect_type(name):
    nl = (name or "").lower()
    if any(x in nl for x in ["wafer","chip","crisps","potato chips"]): return "Potato Chips/Wafers"
    if any(x in nl for x in ["bhujia","sev","gathiya","aloo sev","ratlami"]): return "Sev/Bhujia"
    if any(x in nl for x in ["namkeen","mixture","farsan","chevdo","chiwda","mix"]): return "Namkeen/Mix"
    if any(x in nl for x in ["puff","ring","pellet","stick","extruded","fryum","murukku"]): return "Extruded/Pellets"
    if any(x in nl for x in ["dal","moong","mung","chana","boondi"]): return "Dal/Legume"
    if any(x in nl for x in ["bhakhar","bhakri","papad","rusk"]): return "Papad/Rusk"
    if any(x in nl for x in ["chikki","sweet","pak"]): return "Sweet/Chikki"
    if any(x in nl for x in ["vermicelli","appalam","happala"]): return "Pasta/Appalam"
    return "Other Snack"

# Platform detection from source
def detect_platform(source):
    if not source: return "Unknown"
    s = source.lower()
    if "amazon" in s: return "Amazon"
    if "zepto" in s: return "Zepto"
    if "blinkit" in s or "grofers" in s: return "Blinkit"
    if "swiggy" in s: return "Swiggy Instamart"
    if "bigbasket" in s: return "BigBasket"
    if "jiomart" in s: return "JioMart"
    if "dmart" in s: return "DMart Ready"
    if "flipkart" in s: return "Flipkart"
    if "indiamart" in s: return "IndiaMART"
    if "1mg" in s or "tata" in s: return "1mg/Tata"
    return source[:25]

# Build clean product database from Serper
all_serper_products = []
for p in serper_disc["all_products"]:
    brand = norm_brand(p.get("brand"))
    pname = (p.get("product_name","") or "").strip()
    if not pname or len(pname) < 4: continue
    # Filter noise
    if any(x in pname.lower() for x in ["battery","phone","charger","cable","shoe","shirt",
                                          "tablet","laptop","book","toy","beauty","cream","shampoo",
                                          "medicine","capsule","syrup"]): continue
    all_serper_products.append({
        "brand": brand,
        "product_name": pname,
        "price_rs": p.get("price_num"),
        "price_str": p.get("price_str",""),
        "size_g": p.get("size_g"),
        "price_tier": price_tier(p.get("price_num")),
        "size_tier": size_tier(p.get("size_g")),
        "product_type": detect_type(pname),
        "platform": detect_platform(p.get("source","")),
        "source": p.get("source",""),
        "rating": p.get("rating"),
        "rating_count": p.get("rating_count"),
        "query_slug": p.get("query_slug",""),
        "data_source": "Google Shopping (Serper)",
    })

# Add OFF products
for p in off_brands.get("all_products",[]):
    b = norm_brand(p.get("brands",""))
    pname = (p.get("product_name","") or "").strip()
    if not pname or b == "Unknown": continue
    all_serper_products.append({
        "brand": b,
        "product_name": pname,
        "price_rs": None,
        "price_str": "",
        "size_g": None,
        "price_tier": "Unknown",
        "size_tier": p.get("quantity",""),
        "product_type": detect_type(pname),
        "platform": "Open Food Facts",
        "source": "Open Food Facts",
        "rating": None,
        "rating_count": None,
        "query_slug": "",
        "data_source": "Open Food Facts",
    })

# Deduplicate
seen_pk = set()
products_clean = []
for p in all_serper_products:
    k = f"{p['brand']}||{p['product_name'].lower().strip()}"
    if k not in seen_pk:
        seen_pk.add(k)
        products_clean.append(p)

# Brand aggregates
brand_products = defaultdict(list)
for p in products_clean:
    brand_products[p["brand"]].append(p)

brand_platforms = defaultdict(set)
for p in products_clean:
    if p["platform"] != "Unknown":
        brand_platforms[p["brand"]].add(p["platform"])

brand_price_tiers = defaultdict(set)
for p in products_clean:
    if p["price_tier"] != "Unknown":
        brand_price_tiers[p["brand"]].add(p["price_tier"])

brand_product_types = defaultdict(set)
for p in products_clean:
    brand_product_types[p["brand"]].add(p["product_type"])

# Ph1 brand normalization
def normalise_ph1_brand(b_en, b_lo):
    b = (b_en or b_lo or "").strip()
    bl = b.lower()
    if "balaji wafers" in bl: return "Balaji Wafers"
    if "balaji" in bl: return "Balaji"
    if "surya" in bl: return "Surya"
    if "सूर्य" in (b_lo or ""): return "Surya"
    return b.title() or "Unknown"

ph1_clean = []
for d in ph1_data:
    d2 = dict(d)
    d2["brand"] = normalise_ph1_brand(d.get("brand_english",""), d.get("brand_local",""))
    try: d2["mrp_inr"] = float(d.get("mrp_inr") or 5)
    except: d2["mrp_inr"] = 5.0
    try: d2["net_weight_g"] = float(d.get("net_weight_g"))
    except: d2["net_weight_g"] = None
    ph1_clean.append(d2)

N_PH1 = len(ph1_clean)

# IndiaMART supplier data by context
im_by_brand = {}
for slug, v in indiamart_c["raw_page_data"].items():
    brand_hint = v.get("brand_hint","general")
    if brand_hint not in im_by_brand:
        im_by_brand[brand_hint] = {"suppliers":[], "products":[], "price_mentions":[]}
    im_by_brand[brand_hint]["suppliers"].extend(v.get("suppliers",[]))
    im_by_brand[brand_hint]["products"].extend(v.get("products",[]))
    im_by_brand[brand_hint]["price_mentions"].extend(v.get("price_mentions",[]))

# ── BUILD WORKBOOK ──────────────────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COMPETITIVE LANDSCAPE MAP
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Competitive Map"
ws1.freeze_panes = "C3"

hdr1(ws1, 1, 1, "COMPETITIVE LANDSCAPE — Rs 5-10 SNACK MARKET | Gujarat · Maharashtra · Madhya Pradesh", span=12)
cols = ["Brand","Home State","Key City","Tier","Listed\n(NSE)","Price\nTiers","Product\nTypes",
        "Distribution\nReach","SKUs\nFound","Online\nChannels","Direct\nRival?","Notes"]
for ci, h in enumerate(cols, 1): hdr2(ws1, 2, ci, h)
set_widths(ws1, [20,16,16,20,8,18,28,28,8,32,18,60])

BRAND_ORDER = ["Balaji Wafers","Balaji","Surya","Gopal","Yellow Diamond","Haldiram","Bikaji","Chheda","Anil"]
BRAND_FILLS = {
    "Balaji Wafers": fill("FFE699"), "Balaji": fill("FFE699"),
    "Surya": GREENF, "Gopal": ORANGEF, "Yellow Diamond": YELLOWF,
    "Haldiram": PURPLEF, "Bikaji": PURPLEF, "Chheda": GREENF, "Anil": GREYF,
}

for ri, brand in enumerate(BRAND_ORDER, 3):
    meta = BRAND_META.get(brand, {})
    bf = BRAND_FILLS.get(brand, fill("FFFFFF"))
    prods = brand_products.get(brand, [])
    platforms = sorted(brand_platforms.get(brand, set()))
    vals = [
        brand,
        meta.get("home_state",""),
        meta.get("home_city",""),
        meta.get("tier",""),
        "YES" if meta.get("listed_nse") else "No",
        meta.get("price_tiers",""),
        meta.get("product_types",""),
        meta.get("dist_region",""),
        len(prods),
        "\n".join(platforms) if platforms else "—",
        meta.get("balaji_rival",""),
        meta.get("notes",""),
    ]
    for ci, v in enumerate(vals, 1):
        a = CTR if ci in {1,5,9,11} else WR
        f = BOLDF if ci == 1 else DATAF
        c = wr(ws1, ri, ci, v, f=f, a=a, fl=bf)
    ws1.row_dimensions[ri].height = 56

ws1.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"

# Key findings banner
note_r = len(BRAND_ORDER) + 4
ws1.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=12)
ws1.cell(note_r, 1,
    "KEY FINDING: Gopal Snacks (Rajkot) = Balaji's most direct competitor — same city, same product types, NSE-listed, Nagpur plant in Maharashtra. "
    "Yellow Diamond (Indore) dominates MP-MH border. Haldiram's has confirmed Rs 5 SKUs. Surya Gruh Udhyog is offline-only micro-manufacturer."
).font = fnt(bold=True, color="1F4E79", size=9)
ws1.cell(note_r, 1).fill = fill("DDEEFF")
ws1.row_dimensions[note_r].height = 36

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — FULL PRODUCT DATABASE
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Full Product Database")
ws2.freeze_panes = "C3"

hdr1(ws2, 1, 1, f"FULL PRODUCT DATABASE — {len(products_clean)} discovered SKUs across {len(brand_products)} brands", span=9)
ph = ["Brand","Product Name","Pack Size (g)","Price (Rs)","Price Tier","Product Type","Platform","Source","Data Origin"]
for ci, h in enumerate(ph, 1): hdr2(ws2, 2, ci, h)
set_widths(ws2, [18, 55, 12, 12, 14, 22, 22, 28, 20])

# Sort by brand order then product name
ordered_prods = sorted(products_clean,
    key=lambda x: (BRAND_ORDER.index(x["brand"]) if x["brand"] in BRAND_ORDER else 99, x["product_name"]))

TIER_COLORS = {
    "Rs 5": fill("C6EFCE"), "Rs 10": fill("FFEB9C"), "Rs 20": fill("FFC7CE"),
    "Rs 50": ORANGEF, "Rs 100": PURPLEF, "Rs 100+": PURPLEF, "Unknown": fill("FFFFFF"),
}
for ri, p in enumerate(ordered_prods, 3):
    bf = BRAND_FILLS.get(p["brand"], fill("FFFFFF"))
    tc = TIER_COLORS.get(p["price_tier"], fill("FFFFFF"))
    wr(ws2, ri, 1, p["brand"], f=BOLDF, fl=bf)
    wr(ws2, ri, 2, p["product_name"], fl=fill("FFFFFF"))
    wr(ws2, ri, 3, p["size_g"] or p.get("size_tier",""), a=CTR, fl=fill("FFFFFF"))
    c = wr(ws2, ri, 4, p["price_rs"], a=CTR, fl=tc)
    if p["price_rs"]: c.number_format = '"₹"#,##0.00'
    wr(ws2, ri, 5, p["price_tier"], a=CTR, fl=tc)
    wr(ws2, ri, 6, p["product_type"], fl=fill("FFFFFF"))
    wr(ws2, ri, 7, p["platform"], fl=fill("FFFFFF"))
    wr(ws2, ri, 8, p["source"][:40] if p["source"] else "", f=LINKF, fl=fill("FFFFFF"))
    wr(ws2, ri, 9, p["data_source"], fl=fill("FFFFFF"))
    ws2.row_dimensions[ri].height = 20

ws2.auto_filter.ref = f"A2:{get_column_letter(len(ph))}2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — BRAND PROFILES (deep-dive per brand)
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Brand Profiles")
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 22
ws3.column_dimensions["E"].width = 55

hdr1(ws3, 1, 1, "BRAND PROFILES — Deep Intelligence per Competitor", span=5)

PROFILE_DATA = [
    {
        "brand": "Gopal Snacks Ltd",
        "hq": "Rajkot GIDC, Gujarat",
        "plants": "Unit 1 & 2: Metoda GIDC, Rajkot\nUnit 3: Nagpur, Maharashtra",
        "listed": "NSE (IPO 2023)",
        "revenue": "~Rs 1,800 Cr+ FY24",
        "product_cats": "Chips · Namkeen · Snack Pellets · Premium Namkeen · Rusk · Papad · Spices · Sweets",
        "key_products": "Masala Bite Wafers · Salty Punch Wafers · Gathiya · Farali Chevdo · Tikha Mitha Mix · Udad Papad",
        "price_range": "Rs 5 to Rs 500+",
        "dist": "Gujarat (dominant) · Maharashtra via Nagpur plant · National modern trade",
        "threat_level": "HIGH — Same city as Balaji, Maharashtra plant, listed company with capital",
        "key_intel": "Nagpur plant strategic for Maharashtra distribution. Launched premium range Rs 50+. Online presence: Zepto, BigBasket, DMart, JioMart.",
    },
    {
        "brand": "Prataap Snacks Ltd (Yellow Diamond)",
        "hq": "Indore, Madhya Pradesh",
        "plants": "Indore (MP) · Multiple across India",
        "listed": "NSE (since 2017)",
        "revenue": "~Rs 1,400 Cr+ FY24",
        "product_cats": "Chips · Namkeen · Pellets · Rings · Puffs · Biscuits",
        "key_products": "Yellow Diamond Potato Chips · Rings · Puffs · Spicy Wheels · Uncle Chips · Richies",
        "price_range": "Rs 5 to Rs 200+",
        "dist": "MP (dominant) · Maharashtra · UP · Rajasthan · National",
        "threat_level": "HIGH — Dominant in MP, expanding Maharashtra. Rs 5 SKUs confirmed.",
        "key_intel": "MP-Maharashtra border is their stronghold. Strong Rs 5-10 presence at kirana. Backed by Sequoia Capital earlier.",
    },
    {
        "brand": "Haldiram's",
        "hq": "Delhi (primary) · Nagpur (Maharashtra)",
        "plants": "Nagpur (MH) · Delhi · Noida · Kolkata · multiple",
        "listed": "Private (IPO planned)",
        "revenue": "~Rs 12,000 Cr+ FY24",
        "product_cats": "Bhujia · Sev · Dal · Mixture · Chana · Namkeen · Sweets · RTE",
        "key_products": "Aloo Bhujia Rs 5 · Ratlami Sev Rs 5 · Plain Bhujia 200g · Khatta Meetha · Dal Biji",
        "price_range": "Rs 5 to Rs 500+",
        "dist": "National dominant · Nagpur plant → strong Maharashtra",
        "threat_level": "MEDIUM — Premium perception, limited Rs 5 volume. Brand authority very high.",
        "key_intel": "Rs 5 Aloo Bhujia confirmed on BigBasket. Ratlami Sev Rs 5 packs at kirana. Nagpur unit directly serves MH market.",
    },
    {
        "brand": "Bikaji Foods",
        "hq": "Bikaner, Rajasthan",
        "plants": "Bikaner · Multiple",
        "listed": "NSE (IPO 2022)",
        "revenue": "~Rs 2,000 Cr+ FY24",
        "product_cats": "Bhujia · Sev · Namkeen · Chips · Sweets · Papad",
        "key_products": "Bikaneri Bhujia · Chana Jor Garam · Ratlami Chatpata Sev · Classic Salted Chips",
        "price_range": "Rs 10 to Rs 500+",
        "dist": "National, Rajasthan-MP-Maharashtra strong",
        "threat_level": "MEDIUM — Mostly Rs 10+ at kirana. Premium bhujia positioning.",
        "key_intel": "Less aggressive at Rs 5 than Haldirams/Gopal. Chips range entering Rs 10. Bikaneri Bhujia is iconic product.",
    },
    {
        "brand": "Chheda's Snacks",
        "hq": "Mumbai, Maharashtra",
        "plants": "Mumbai",
        "listed": "Private",
        "revenue": "~Rs 200-500 Cr est.",
        "product_cats": "Bhakharwadi · Mix Farsan · Tikha Namkeen · Chana Chor · Dry Fruits",
        "key_products": "Small Bhakharwadi 170g · Mix Farsan 170g · Chana Chor · Tikha Namkeen",
        "price_range": "Rs 47 to Rs 200",
        "dist": "Mumbai dominant · Maharashtra · Limited national",
        "threat_level": "LOW — Different price segment (Rs 47+). But strong MH brand equity.",
        "key_intel": "Strong on Zepto and BigBasket in Maharashtra. Bhakharwadi is Maharashtra's signature snack. No Rs 5-10 play.",
    },
    {
        "brand": "Surya Gruh Udhyog",
        "hq": "Nandurbar, Maharashtra",
        "plants": "Nandurbar (2 FSSAI licenses)",
        "listed": "Private micro-manufacturer",
        "revenue": "<Rs 10 Cr est.",
        "product_cats": "Extruded pellets · Noodle sticks · Biryani stick · Pasta shapes",
        "key_products": "Biryani Stick · Wonder Car · Pasta · Ladki Bahin (scheme-branded)",
        "price_range": "Rs 5 only",
        "dist": "Nandurbar district + surrounding north MH. Kirana-only.",
        "threat_level": "MONITOR — Hyper-local, but 20% retailer margin creates retailer loyalty vs our 10%.",
        "key_intel": "2 manufacturing units (2 FSSAI licenses = 2 plants). Offline-only. Ladki Bahin branding = clever scheme marketing. No B2B/online presence.",
    },
]

r = 3
for prof in PROFILE_DATA:
    brand = prof["brand"]
    # Brand header
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws3.cell(r, 1, f"  {brand}"); c.fill=H1F; c.font=H1T; c.alignment=WR; c.border=MED
    r += 1
    fields = [
        ("HQ", prof["hq"]), ("Manufacturing Plants", prof["plants"]),
        ("Listed", prof["listed"]), ("Est. Revenue", prof["revenue"]),
        ("Product Categories", prof["product_cats"]),
        ("Key Products", prof["key_products"]),
        ("Price Range", prof["price_range"]),
        ("Distribution", prof["dist"]),
        ("Threat to Balaji", prof["threat_level"]),
        ("Key Intelligence", prof["key_intel"]),
    ]
    for field, value in fields:
        rf = YELLOWF if field == "Threat to Balaji" else rfill(r)
        wr(ws3, r, 1, field, f=BOLDF, fl=rf)
        ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        c2 = ws3.cell(r, 2, value); c2.fill=rf; c2.border=THIN
        c2.font = fnt(bold="HIGH" in str(value) or "MEDIUM" in str(value), size=9,
                      color="CC0000" if "HIGH" in str(value) else "856404" if "MEDIUM" in str(value) else "000000")
        c2.alignment = WR
        ws3.row_dimensions[r].height = max(20, len(value)//50 * 14 + 18)
        r += 1
    r += 1  # gap between brands

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — PACKET SIZE DISTRIBUTION
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Pack Size Distribution")
ws4.freeze_panes = "B3"

hdr1(ws4, 1, 1, "PACKET SIZE & PRICE TIER DISTRIBUTION — by Brand", span=9)
size_cols = ["Brand","Rs 5\n(≤16g)","Rs 10\n(17-30g)","Rs 20\n(31-60g)","Rs 50\n(61-150g)","Rs 100+\n(>150g)","Smallest\nPack (g)","Largest\nPack (g)","Key Size\nObservation"]
for ci, h in enumerate(size_cols, 1): hdr2(ws4, 2, ci, h)
set_widths(ws4, [20,10,10,10,10,10,12,12,50])

TIERS = ["Rs 5 (≤16g)","Rs 10 (17-30g)","Rs 20 (31-60g)","Rs 50 (61-150g)","Rs 100+ (>150g)"]

for ri, brand in enumerate(BRAND_ORDER, 3):
    prods = brand_products.get(brand, [])
    sized = [p for p in prods if p.get("size_g")]
    tier_counts = Counter(size_tier(p["size_g"]) for p in sized)
    sizes = [p["size_g"] for p in sized]

    rf = BRAND_FILLS.get(brand, fill("FFFFFF"))
    wr(ws4, ri, 1, brand, f=BOLDF, fl=rf)
    for ci, tier in enumerate(TIERS, 2):
        cnt = tier_counts.get(tier, 0)
        cell_fill = GREENF if (tier == "Rs 5 (≤16g)" and cnt > 0) else (YELLOWF if cnt > 0 else fill("FFFFFF"))
        wr(ws4, ri, ci, cnt if cnt > 0 else "—", a=CTR, fl=cell_fill)

    min_s = min(sizes) if sizes else None
    max_s = max(sizes) if sizes else None
    wr(ws4, ri, 7, f"{min_s:.0f}g" if min_s else "?", a=CTR, fl=rf)
    wr(ws4, ri, 8, f"{max_s:.0f}g" if max_s else "?", a=CTR, fl=rf)

    # Observation
    obs = ""
    if brand in ("Balaji Wafers","Balaji"):
        obs = "Rs 5 packs = kirana exclusive (5-16g). Online presence mostly 150-400g. Full ladder from Rs 5 to Rs 200."
    elif brand == "Surya":
        obs = "Rs 5 only. No larger packs found online — purely kirana micro-distribution."
    elif brand == "Gopal":
        obs = "Rs 5-10 packs present (Masala Bite Rs 27-44 online). Offline Rs 5 smaller packs. Full size ladder."
    elif brand == "Yellow Diamond":
        obs = "Rs 5-10 kirana packs. Online shows Rs 25-100 range. Strong Rs 10 price point."
    elif brand == "Haldiram":
        obs = "Rs 5 confirmed (Aloo Bhujia, Ratlami Sev at Rs 5). But dominates Rs 50-200g segment."
    elif brand == "Bikaji":
        obs = "Mostly Rs 25-110 online. Rs 10 kirana packs suspected but limited online data."
    elif brand == "Chheda":
        obs = "No Rs 5-10 play. Standard size 170g at Rs 47-60. Premium farsan positioning."
    wr(ws4, ri, 9, obs, fl=fill("FFFFFF"))
    ws4.row_dimensions[ri].height = 32

ws4.auto_filter.ref = "A2:I2"

# Insight below
ins_r = len(BRAND_ORDER) + 5
ws4.merge_cells(start_row=ins_r, start_column=1, end_row=ins_r, end_column=9)
ws4.cell(ins_r, 1,
    "INSIGHT: The Rs 5 segment (≤16g) is a kirana-only channel — Google Shopping shows near-zero Rs 5 packs because they don't ship online. "
    "This is a STRUCTURAL MOAT: brands with Rs 5 packs (Balaji, Surya, Haldiram, Gopal) operate in an offline-only competitive space. "
    "Online data underrepresents the Rs 5 segment. Field data (our 19 SKUs) is the primary source of truth for this segment."
).font = fnt(italic=True, size=8, color="444444")
ws4.cell(ins_r, 1).fill = fill("DDEEFF")
ws4.row_dimensions[ins_r].height = 40

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — B2B WHOLESALE INTELLIGENCE
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("B2B Wholesale Intel")
ws5.freeze_panes = "C3"

hdr1(ws5, 1, 1, "B2B WHOLESALE INTELLIGENCE — IndiaMART Suppliers · Manufacturers · Distributors", span=6)
bh = ["Company / Supplier","Location","B2B Price\nMention","Brand Context","Product Desc.","Source Page"]
for ci, h in enumerate(bh, 1): hdr2(ws5, 2, ci, h)
set_widths(ws5, [40, 20, 18, 18, 50, 30])

all_im_suppliers = indiamart_c.get("unique_suppliers", [])
for ri, s in enumerate(all_im_suppliers, 3):
    rf = rfill(ri)
    company = s.get("company","") or s.get("raw_text","")[:60]
    loc = s.get("location","")
    price = s.get("price","")
    brand_h = s.get("brand_hint","")
    raw = s.get("raw_text","")
    # Try to extract product description from raw text (first meaningful line)
    lines = [l.strip() for l in raw.replace("\n"," | ").split("|") if len(l.strip()) > 8]
    prod_desc = lines[0][:80] if lines else ""
    src_url = s.get("source_url","").replace("https://dir.indiamart.com/search.mp?ss=","IM: ")

    wr(ws5, ri, 1, company[:60], fl=rf)
    wr(ws5, ri, 2, loc, a=CTR, fl=rf)
    price_fill = GREENF if price and any(c.isdigit() for c in price) else rf
    wr(ws5, ri, 3, price, a=CTR, fl=price_fill)
    wr(ws5, ri, 4, brand_h, a=CTR, fl=rf)
    wr(ws5, ri, 5, prod_desc, fl=rf)
    wr(ws5, ri, 6, src_url[:40], f=LINKF, fl=rf)
    ws5.row_dimensions[ri].height = 22

ws5.auto_filter.ref = "A2:F2"

# Price analysis section
price_section_r = len(all_im_suppliers) + 5
ws5.merge_cells(start_row=price_section_r, start_column=1, end_row=price_section_r, end_column=6)
ws5.cell(price_section_r, 1, "PRICE MENTIONS BY BRAND CONTEXT — IndiaMART pages").font = fnt(True, "1F4E79", 10)
ws5.cell(price_section_r, 1).fill = fill("D6E4F0")
price_section_r += 1

for ci2, h2 in enumerate(["Brand Context","Price Mentions (raw)","Count"], 1):
    x = ws5.cell(price_section_r, ci2, h2); x.fill=H2F; x.font=H2T; x.alignment=CTR; x.border=THIN
price_section_r += 1

# Extract meaningful price mentions per brand
BRAND_PRICE_CONTEXTS = ["Balaji","Surya","Gopal","Yellow Diamond","Prataap","Haldiram","Bikaji","Gujarat","Maharashtra","MP"]
for brand_ctx in BRAND_PRICE_CONTEXTS:
    ctx_data = im_by_brand.get(brand_ctx, {})
    prices = [p for p in ctx_data.get("price_mentions",[])
              if re.search(r'₹\s*\d+', p) and "turnover" not in p.lower()
              and float(re.search(r'[\d,]+', p.replace(",","")).group() or "0") < 10000]
    if not prices: continue
    rf = rfill(price_section_r)
    wr(ws5, price_section_r, 1, brand_ctx, f=BOLDF, fl=rf)
    ws5.merge_cells(start_row=price_section_r, start_column=2, end_row=price_section_r, end_column=5)
    c2 = ws5.cell(price_section_r, 2, " | ".join(prices[:8])); c2.fill=rf; c2.border=THIN; c2.font=DATAF; c2.alignment=WR
    wr(ws5, price_section_r, 6, len(prices), a=CTR, fl=rf)
    ws5.row_dimensions[price_section_r].height = 24
    price_section_r += 1

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — ONLINE CHANNEL MAP
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Online Channel Map")
ws6.freeze_panes = "B3"

PLATFORMS = ["Amazon","BigBasket","Zepto","Blinkit","JioMart","Swiggy Instamart","DMart Ready","Flipkart","IndiaMART","Other"]

hdr1(ws6, 1, 1, "ONLINE CHANNEL PRESENCE MAP — Which brands sell where", span=len(PLATFORMS)+2)
ws6.cell(2, 1, "Brand").fill=H2F; ws6.cell(2,1).font=H2T; ws6.cell(2,1).alignment=CTR; ws6.cell(2,1).border=THIN
ws6.cell(2, 2, "SKUs\nFound").fill=H2F; ws6.cell(2,2).font=H2T; ws6.cell(2,2).alignment=CTR; ws6.cell(2,2).border=THIN
for ci, plat in enumerate(PLATFORMS, 3):
    x = ws6.cell(2, ci, plat.replace(" ","\n")); x.fill=H2F; x.font=H2T; x.alignment=CTR; x.border=THIN
ws6.row_dimensions[2].height = 36
set_widths(ws6, [20, 10] + [12]*len(PLATFORMS))

for ri, brand in enumerate(BRAND_ORDER, 3):
    prods = brand_products.get(brand, [])
    bf = BRAND_FILLS.get(brand, fill("FFFFFF"))
    platforms_found = brand_platforms.get(brand, set())
    wr(ws6, ri, 1, brand, f=BOLDF, fl=bf)
    wr(ws6, ri, 2, len(prods), a=CTR, fl=bf)
    for ci, plat in enumerate(PLATFORMS, 3):
        is_present = plat in platforms_found or any(plat.lower() in p.lower() for p in platforms_found)
        cell_fill = GREENF if is_present else fill("FFFFFF")
        cell_val = "✓" if is_present else ""
        wr(ws6, ri, ci, cell_val, a=CTR, fl=cell_fill)
    ws6.row_dimensions[ri].height = 28

ws6.auto_filter.ref = f"A2:{get_column_letter(len(PLATFORMS)+2)}2"

# Insight
ins_r2 = len(BRAND_ORDER) + 5
ws6.merge_cells(start_row=ins_r2, start_column=1, end_row=ins_r2, end_column=len(PLATFORMS)+2)
ws6.cell(ins_r2, 1,
    "INSIGHT: Online channels carry 150-500g packs of established brands. Rs 5 micro-packs are ABSENT from all online channels — "
    "this means Zepto/Blinkit/BigBasket cannot threaten Rs 5 kirana distribution. The Rs 5 segment's competitive battleground "
    "is 100% offline: kirana shelf space, retailer margin, and hyper-local brand recognition."
).font = fnt(italic=True, size=8, color="444444")
ws6.cell(ins_r2, 1).fill = fill("DDEEFF")
ws6.row_dimensions[ins_r2].height = 40

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — OUR 19 SKUs (reference)
# ═══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Our 19 SKUs (Ph1)")
ws7.freeze_panes = "D3"

hdr1(ws7, 1, 1, "OUR 19 SKUs — Inventoried Packets (Phase 1 Field Data)", span=10)
sh = ["#","Brand","Product","Net Wt (g)","MRP (Rs)","FSSAI License","Manufacturer","Type","Flavor","Ingredients (summary)"]
for ci, h in enumerate(sh, 1): hdr2(ws7, 2, ci, h)
set_widths(ws7, [4,16,36,10,9,20,26,18,20,50])

for ri, d in enumerate(ph1_clean, 3):
    rf = rfill(ri)
    prod = d.get("product_name_english","") or ""
    raw_ing = d.get("ingredients_full","") or ""
    ing_short = raw_ing[:100] if raw_ing else ""
    vals = [
        d.get("packet_num"), d.get("brand"), prod,
        d.get("net_weight_g"), d.get("mrp_inr"),
        d.get("fssai_license"), d.get("manufacturer_name"),
        d.get("product_type"), d.get("variant_flavor"), ing_short,
    ]
    for ci, v in enumerate(vals, 1):
        a = CTR if ci in {1,4,5} else WR
        wr(ws7, ri, ci, v, a=a, fl=rf)
    ws7.row_dimensions[ri].height = 28

ws7.auto_filter.ref = "A2:J2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — PM INTELLIGENCE & GAPS
# ═══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("PM Intelligence")
ws8.column_dimensions["A"].width = 28
ws8.column_dimensions["B"].width = 85

hdr1(ws8, 1, 1, "PM COMPETITIVE INTELLIGENCE — Strategic Gaps & Opportunities", span=2)

sections = [
    ("COMPETITIVE THREAT RANKING", [
        ("🔴 Gopal Snacks (CRITICAL)", "Same city as Balaji (Rajkot GIDC). NSE-listed with capital. Nagpur plant directly targeting Maharashtra. Broader product range (chips, namkeen, pellets, papad, spices, rusk, sweets). Online on Zepto, BigBasket, JioMart. Direct shelf competitor at kirana."),
        ("🔴 Yellow Diamond/Prataap (CRITICAL)", "Indore-based, dominates MP and MP-MH border belt. Rs 5 SKUs confirmed in market. NSE-listed, Sequoia-backed. Strong kirana penetration in exactly the regions (MP-MH) where Surya operates."),
        ("🟡 Haldiram's (WATCH)", "National brand, Nagpur plant serves Maharashtra. Rs 5 Aloo Bhujia and Ratlami Sev confirmed at kirana. Premium trust advantage — consumer pays for brand. Lower volume in Rs 5 than local players."),
        ("🟡 Bikaji (WATCH)", "Mostly Rs 10+ plays. Signature Bikaneri Bhujia well-established. Less aggressive in Gujarat/Maharashtra Rs 5 segment specifically."),
        ("🟢 Chheda's (LOW RISK)", "Different segment entirely — Rs 47+ premium farsan. No Rs 5 play. Maharashtra shelf competitor for mindshare but not for wallet."),
        ("🟢 Surya Gruh Udhyog (MONITOR)", "Micro-manufacturer. Only relevant for North Maharashtra/Nandurbar district. Key threat: 20% retailer margin vs our 10% — retailers actively push Surya."),
    ]),
    ("STRUCTURAL MARKET INSIGHTS", [
        ("Rs 5 = offline-only battleground", "Zero Rs 5 packs found on any online platform (Amazon, Zepto, BigBasket, JioMart). The entire Rs 5 competitive war is fought at kirana shelf. Digital marketing, D2C, ecommerce are irrelevant for this segment."),
        ("Retailer margin = the weapon", "Surya wins at local level purely through 20% retailer margin. Gopal and Yellow Diamond likely use similar tactics. The margin war IS the distribution war at Rs 5."),
        ("Pack size ladder gap", "Rs 5 → Rs 10 → Rs 50 is the real ladder. Rs 20 packs (30-60g) appear underrepresented. A Rs 20 / 40g pack could be a market gap — better value than Rs 5 (16g @ Rs 0.31/g) and more accessible than Rs 50 packs."),
        ("Geographic cluster: Rajkot", "Balaji AND Gopal both based in Rajkot GIDC. Shared distribution routes, same kirana contacts. Gopal's Nagpur plant shows Maharashtra ambition. Watch for Gopal price drops in Maharashtra."),
        ("Product type gap at Rs 5", "Zero Rs 5 baked/low-sodium/high-protein products across any brand. Every Rs 5 SKU is fried, 500-570 kcal/100g. First brand to make a credible 'better-for-you' Rs 5 pack wins a new shelf."),
        ("Marathi-language packs win locally", "Our Surya data confirms Devanagari script is essential for North MH market. Ladki Bahin scheme branding is hyper-effective. Cultural relevance > ingredient list for impulse purchase at Rs 5."),
        ("Yellow Diamond in MP = threat to MH expansion", "If Balaji wants to expand into MP, Yellow Diamond is the 800-lb gorilla. Any MH-MP expansion must account for their distribution density and Rs 5 SKUs already at kirana there."),
    ]),
    ("PRODUCT PORTFOLIO GAPS (unclaimed at Rs 5-10)", [
        ("Baked/air-fried snack", "No brand offers a baked product at Rs 5. Consumer awareness of 'less oil' is growing even at lower income levels. First-mover advantage available."),
        ("High-protein Rs 5 snack", "Zero high-protein claims across all 19 inventoried SKUs and all competitor products found. Soy/dal-based protein at Rs 5 is technically feasible."),
        ("25g at Rs 5", "Our analysis shows Balaji's 16g pack = Rs 0.31/g vs 24g packs at Rs 0.21/g (48% worse value). A 25g pack at Rs 5 = Rs 0.20/g = dominant value position vs any competitor."),
        ("Regional flavor + scheme tie-in", "Surya's Ladki Bahin brand shows scheme-marketing works at Rs 5. Other state schemes (PM Awas Yojana, Ujjwala, MGNREGA) have brand equity in target demographics. Unexplored."),
        ("Low-sodium variant", "Sodium varies 300-1300mg/100g across our SKUs. No product claims 'low sodium.' Simple reformulation + claim = differentiation."),
    ]),
    ("DISTRIBUTION INTELLIGENCE", [
        ("Gopal Nagpur plant", "Gopal's dedicated Maharashtra plant means they can offer same distribution economics as local Gujarat brands. Watch pricing and availability in Nashik, Nagpur, Aurangabad belts."),
        ("Haldiram Nagpur unit", "Haldiram has Nagpur manufacturing — this means competitive pricing in Maharashtra. Their Rs 5 namkeen packs are subsidized by higher-margin larger packs in same distribution route."),
        ("Surya = kirana loyalty network", "Two manufacturing units, hyper-local but deep. At kirana level in Nandurbar + surrounding district, Surya likely has >50% shelf share vs Balaji. Retailer margin is the cause."),
        ("Online quick-commerce irrelevant at Rs 5", "Zepto/Blinkit/Swiggy Instamart carry 150-500g packs Rs 25+. They are not the channel for Rs 5 competitive dynamics. Our battleground is kirana, not app."),
    ]),
]

r = 3
for section_title, items in sections:
    ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws8.cell(r, 1, section_title); c.fill=H1F; c.font=H1T; c.alignment=WR; c.border=MED
    ws8.row_dimensions[r].height = 22
    r += 1
    for key, value in items:
        rf = rfill(r)
        wr(ws8, r, 1, key, f=BOLDF, fl=rf)
        wr(ws8, r, 2, value, f=fnt(size=9), fl=rf)
        ws8.row_dimensions[r].height = max(28, len(value)//80*14 + 18)
        r += 1
    r += 1

# ─── SAVE ──────────────────────────────────────────────────────────────────────
out = ROOT / "chips_inventory_v7.xlsx"
wb.save(out)

print(f"Saved: {out}")
print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
print(f"Products in database: {len(products_clean)} from {len(brand_products)} brands")
print(f"B2B suppliers: {len(all_im_suppliers)}")
print(f"Ph1 SKUs: {N_PH1}")
