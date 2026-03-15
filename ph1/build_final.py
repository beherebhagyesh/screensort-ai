import sys, json, re
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

raw = json.loads(Path("packet_data.json").read_text(encoding="utf-8"))

# ── BRAND NORMALISATION ───────────────────────────────────────────────────────
def normalise_brand(b_en, b_lo):
    b = (b_en or b_lo or "").strip()
    bl = b.lower()
    if "haldiram" in bl:
        return "Balaji"
    if "balaji wafers" in bl or bl == "balaji wafers":
        return "Balaji Wafers"
    if "balaji" in bl:
        return "Balaji"
    if "surya" in bl:
        return "Surya"
    lo = (b_lo or "").strip()
    if "सूर्य" in lo: return "Surya"
    if "बालाजी" in lo: return "Balaji"
    return b.title() or "Unknown"

# ── INGREDIENT PARSING ────────────────────────────────────────────────────────
def parse_ingredients(raw_str):
    if not raw_str: return []
    s = raw_str.strip()
    # Remove percentage annotations like (45%), (one or more of the following:)
    s = re.sub(r'\([^)]{0,60}\)', '', s)
    # Split on comma or semicolon
    parts = re.split(r'[,;]', s)
    cleaned = []
    for p in parts:
        p = p.strip().strip('*').strip()
        if len(p) > 1 and not re.match(r'^[\d\.\s%]+$', p):
            cleaned.append(p.title())
    return [x for x in cleaned if x]

def ingredients_as_list(raw_str):
    items = parse_ingredients(raw_str)
    return "\n".join(f"• {i}" for i in items) if items else ""

# ── DATA CLEANING ─────────────────────────────────────────────────────────────
data = []
for d in raw:
    b_en = d.get("brand_english", "") or ""
    b_lo = d.get("brand_local", "") or ""
    brand = normalise_brand(b_en, b_lo)
    cleaned = dict(d)
    cleaned["brand"] = brand
    try: cleaned["mrp_inr"] = float(d.get("mrp_inr") or 5)
    except: cleaned["mrp_inr"] = 5.0
    try: cleaned["net_weight_g"] = float(d.get("net_weight_g"))
    except: cleaned["net_weight_g"] = None
    data.append(cleaned)

N = len(data)

# ── TRADE PRICING ─────────────────────────────────────────────────────────────
# Default per brand
TRADE_DEFAULT = {
    "Balaji":       {"mfr": 4.25, "dist": 4.50},
    "Balaji Wafers":{"mfr": 4.25, "dist": 4.50},
    "Surya":        {"mfr": 3.50, "dist": 4.00},
}
# SKU-level overrides (by product name keyword)
def get_trade(brand, product_name):
    pn = (product_name or "").lower()
    if "mung" in pn or "moong" in pn or "dal" in pn:
        return {"mfr": 4.50, "dist": 4.75}
    return TRADE_DEFAULT.get(brand, {"mfr": None, "dist": None})

# ── STYLES ────────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)
THIN  = Border(**{s: Side(style="thin",   color="D0D0D0") for s in "left right top bottom".split()})
MED   = Border(**{s: Side(style="medium", color="2E75B6") for s in "left right top bottom".split()})
WR    = Alignment(wrap_text=True, vertical="top")
CTR   = Alignment(horizontal="center", vertical="top", wrap_text=True)

H1F=fill("1F4E79"); H1T=fnt(True,"FFFFFF",9)
H2F=fill("2E75B6"); H2T=fnt(True,"FFFFFF",9)
ALTF=fill("EBF3FB")
DATAF=fnt(size=9); CALCF=fnt(color="000000",size=9)
INPF=fnt(color="0000FF",size=9)
WARNF=fill("FFF2CC"); GREENF=fill("E2EFDA"); REDF=fill("FFE0E0")

def hdr1(ws, r, c, v, span=1):
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    x = ws.cell(r, c, v); x.fill=H1F; x.font=H1T; x.alignment=CTR; x.border=MED

def hdr2(ws, r, c, v):
    x = ws.cell(r, c, v); x.fill=H2F; x.font=H2T; x.alignment=CTR; x.border=THIN

def wr(ws, r, c, v, f=None, a=None, fl=None):
    x = ws.cell(r, c, v); x.border=THIN
    x.font = f or DATAF; x.alignment = a or WR
    if fl: x.fill = fl
    return x

def rfill(ri): return ALTF if ri % 2 == 0 else fill("FFFFFF")
def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTER INVENTORY
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Master Inventory"
ws1.freeze_panes = "D3"
ws1.row_dimensions[1].height = 16
ws1.row_dimensions[2].height = 36

secs = [(1,3,"IDENTIFICATION"),(4,6,"PRICING & WEIGHT"),(7,10,"MANUFACTURER"),
        (11,12,"COMPLIANCE"),(13,14,"PRODUCT"),(15,17,"PACK COLORS"),
        (18,21,"PACK DESIGN"),(22,23,"SCORES"),(24,25,"COMPETITIVE INTEL")]
for cs, ce, title in secs:
    hdr1(ws1, 1, cs, title, span=ce-cs+1)

c2 = ["#","Brand","Product","MRP (Rs)","Net Wt (g)","Price/g (Rs)",
      "Manufacturer","City","State","Phone","FSSAI License","Best Before (mo)",
      "Type","Flavor / Variant","Primary Color","Secondary Color","Accent Colors",
      "Pack Format","Front Imagery","Design Style","Font Style",
      "Appeal /10","Shelf Vis /10","Claims on Pack","Competitive Notes"]
for ci, h in enumerate(c2, 1): hdr2(ws1, 2, ci, h)
set_widths(ws1, [4,16,32,9,9,10,26,14,14,14,20,10,18,20,16,16,18,14,24,18,14,12,12,32,36])

for ri, d in enumerate(data, 3):
    rf = rfill(ri)
    brand = d.get("brand","?"); mrp = d.get("mrp_inr"); wt = d.get("net_weight_g")
    prod_en = d.get("product_name_english","") or ""
    prod_lo = d.get("product_name_local","") or ""
    prod = prod_en + (f"\n[{prod_lo}]" if prod_lo and prod_lo != prod_en else "")
    ws1.row_dimensions[ri].height = 32

    vals = [
        d.get("packet_num"), brand, prod,
        mrp, wt,
        f"=D{ri}/E{ri}" if (mrp and wt) else None,
        d.get("manufacturer_name"), d.get("manufacturer_city"), d.get("manufacturer_state"), d.get("manufacturer_phone"),
        d.get("fssai_license"), d.get("best_before_months"),
        d.get("product_type"), d.get("variant_flavor"),
        d.get("pack_primary_color"), d.get("pack_secondary_color"), d.get("pack_accent_colors"),
        d.get("pack_format"), d.get("front_imagery"), d.get("design_style"), d.get("font_style"),
        d.get("pack_appeal_score"), d.get("shelf_visibility_score"),
        d.get("claims_on_pack"), d.get("competitive_notes"),
    ]
    for ci, v in enumerate(vals, 1):
        a = CTR if ci in {1,4,5,6,12,22,23} else WR
        f = CALCF if ci == 6 else INPF if ci in {4,5} else DATAF
        fl = WARNF if (ci == 5 and v is None) else rf
        c = wr(ws1, ri, ci, v, f=f, a=a, fl=fl)
        if ci == 6 and v: c.number_format = "0.000"
        if ci == 4 and v: c.number_format = '"Rs"#,##0.00'

ws1.auto_filter.ref = f"A2:{get_column_letter(len(c2))}2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — INGREDIENTS (itemized) + INGREDIENT ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Ingredients")
ws2.freeze_panes = "C3"
ws2.row_dimensions[1].height = 16
ws2.row_dimensions[2].height = 36

hdr1(ws2, 1, 1, "PRODUCT", span=2)
hdr1(ws2, 1, 3, "INGREDIENTS — ITEMIZED LIST (one per line)", span=2)
hdr1(ws2, 1, 5, "ALLERGENS & DIETARY")
hdr1(ws2, 1, 6, "INGREDIENT COUNT")

cols2 = ["#","Product","Ingredients (itemized)","Raw text (original)","Allergens","# Ingredients"]
for ci, h in enumerate(cols2, 1): hdr2(ws2, 2, ci, h)
set_widths(ws2, [4, 30, 48, 52, 24, 12])

all_ingredients = []  # for frequency analysis

for ri, d in enumerate(data, 3):
    rf = rfill(ri)
    raw_ing = d.get("ingredients_full","") or ""
    parsed = parse_ingredients(raw_ing)
    all_ingredients.extend(parsed)
    itemized = ingredients_as_list(raw_ing)
    ws2.row_dimensions[ri].height = max(30, len(parsed) * 12)

    wr(ws2, ri, 1, d.get("packet_num"), a=CTR, fl=rf)
    wr(ws2, ri, 2, d.get("product_name_english") or d.get("product_name_local"), fl=rf)
    x = wr(ws2, ri, 3, itemized, fl=rf)
    x.alignment = Alignment(wrap_text=True, vertical="top")
    wr(ws2, ri, 4, raw_ing, f=fnt(size=8, italic=True, color="666666"), fl=rf)
    wr(ws2, ri, 5, d.get("allergens"), fl=rf)
    wr(ws2, ri, 6, len(parsed) if parsed else None, a=CTR, fl=rf)

ws2.auto_filter.ref = "A2:F2"

# ── INGREDIENT FREQUENCY ANALYSIS (below the data) ───────────────────────────
sep_row = N + 5
ws2.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=6)
ws2.cell(sep_row, 1, "INGREDIENT FREQUENCY ANALYSIS — How often each ingredient appears across all 19 SKUs")
ws2.cell(sep_row, 1).font = fnt(True, "1F4E79", 11)
ws2.cell(sep_row, 1).fill = fill("D6E4F0")
ws2.cell(sep_row, 1).alignment = CTR

hdr2(ws2, sep_row+1, 1, "Rank")
hdr2(ws2, sep_row+1, 2, "Ingredient")
hdr2(ws2, sep_row+1, 3, "# Products")
hdr2(ws2, sep_row+1, 4, "% of SKUs")
hdr2(ws2, sep_row+1, 5, "Category")
hdr2(ws2, sep_row+1, 6, "PM Note")

freq = Counter(all_ingredients).most_common(40)

def ingredient_category(name):
    n = name.lower()
    if any(x in n for x in ["oil","palmolein","cottonseed","groundnut"]): return "Fat/Oil"
    if any(x in n for x in ["salt","sodium","iodized"]): return "Salt/Mineral"
    if any(x in n for x in ["chilli","chili","pepper","masala","spice","cumin","coriander","turmeric","ajwain","fennel","mango","amchur","tamarind"]): return "Spices/Seasoning"
    if any(x in n for x in ["flour","maida","wheat","corn","rice","potato","gram","chickpea","lentil","dal","mung","moong"]): return "Starch/Base"
    if any(x in n for x in ["sugar","glucose","maltodextrin"]): return "Sugar"
    if any(x in n for x in ["ins","acidity","antical","anti-cal","flavour enhancer","flavour"]): return "Additive/Flavour"
    if any(x in n for x in ["protein","soya","milk","cheese","whey"]): return "Protein"
    if any(x in n for x in ["onion","garlic","ginger","tomato","dehydrated veg"]): return "Vegetable"
    return "Other"

for i, (ing, cnt) in enumerate(freq, 1):
    r = sep_row + 1 + i
    pct = cnt / N
    rf2 = GREENF if pct >= 0.5 else WARNF if pct >= 0.25 else rfill(r)
    note = "Universal base" if pct >= 0.8 else "Very common" if pct >= 0.5 else "Common" if pct >= 0.25 else ""
    wr(ws2, r, 1, i, a=CTR, fl=rf2)
    wr(ws2, r, 2, ing, fl=rf2)
    wr(ws2, r, 3, cnt, a=CTR, fl=rf2)
    wr(ws2, r, 4, pct, a=CTR, fl=rf2).number_format = "0%"
    wr(ws2, r, 5, ingredient_category(ing), a=CTR, fl=rf2)
    wr(ws2, r, 6, note, a=CTR, fl=rf2)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — NUTRITION ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Nutrition Analysis")
ws3.freeze_panes = "A2"

# Build nutrition data
nut_data = []
for d in data:
    def n(key):
        try: return float(d.get(key))
        except: return None
    nut_data.append({
        "num": d.get("packet_num"),
        "brand": d.get("brand"),
        "product": d.get("product_name_english") or d.get("product_name_local",""),
        "wt": d.get("net_weight_g"),
        "cal": n("calories_per_100g"),
        "fat": n("total_fat_g"),
        "sat": n("saturated_fat_g"),
        "trans": n("trans_fat_g"),
        "carbs": n("total_carbs_g"),
        "fiber": n("dietary_fiber_g"),
        "sugar": n("sugar_g"),
        "protein": n("protein_g"),
        "sodium": n("sodium_mg"),
    })

# Section helper
def section_header(ws, r, title, ncols):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.cell(r, 1, title).font = fnt(True, "1F4E79", 11)
    ws.cell(r, 1).fill = fill("D6E4F0")
    ws.cell(r, 1).alignment = CTR
    ws.row_dimensions[r].height = 20
    return r + 1

def rank_table(ws, start_row, title, key, label, ncols=5, reverse=True, fmt="0.0"):
    r = section_header(ws, start_row, title, ncols)
    hdrs = ["Rank", "Product", "Brand", label, "vs Average"]
    for ci, h in enumerate(hdrs, 1): hdr2(ws, r, ci, h)
    r += 1
    valid = [(d, d[key]) for d in nut_data if d[key] is not None]
    ranked = sorted(valid, key=lambda x: x[1], reverse=reverse)
    avg = sum(v for _, v in ranked) / len(ranked) if ranked else 0
    for rank, (d, val) in enumerate(ranked, 1):
        rf2 = REDF if rank == 1 and reverse else GREENF if rank == 1 and not reverse else rfill(r)
        wr(ws, r, 1, rank, a=CTR, fl=rf2)
        wr(ws, r, 2, d["product"], fl=rf2)
        wr(ws, r, 3, d["brand"], a=CTR, fl=rf2)
        c = wr(ws, r, 4, val, a=CTR, fl=rf2); c.number_format = fmt
        diff = val - avg
        sign = "+" if diff > 0 else ""
        wr(ws, r, 5, f"{sign}{diff:.1f}", a=CTR, fl=rf2)
        r += 1
    return r + 1

set_widths(ws3, [6, 32, 16, 14, 14])
ws3.row_dimensions[1].height = 20

cur = 1
cur = rank_table(ws3, cur, "🔥 HIGHEST ENERGY (kcal per 100g) — Most calorie-dense products", "cal", "kcal/100g")
cur = rank_table(ws3, cur, "🧂 HIGHEST SODIUM (mg per 100g) — Saltiest products", "sodium", "Sodium (mg)")
cur = rank_table(ws3, cur, "💪 HIGHEST PROTEIN (g per 100g)", "protein", "Protein (g)")
cur = rank_table(ws3, cur, "🧈 HIGHEST TOTAL FAT (g per 100g)", "fat", "Fat (g)")
cur = rank_table(ws3, cur, "🍬 HIGHEST SUGAR (g per 100g)", "sugar", "Sugar (g)")
cur = rank_table(ws3, cur, "🌾 HIGHEST CARBS (g per 100g)", "carbs", "Carbs (g)")
cur = rank_table(ws3, cur, "✅ HIGHEST DIETARY FIBER (g per 100g) — Best for health claims", "fiber", "Fiber (g)")
cur = rank_table(ws3, cur, "✅ LOWEST SODIUM — Salt-conscious options", "sodium", "Sodium (mg)", reverse=False)

# Full comparison table
cur = section_header(ws3, cur, "FULL NUTRITION COMPARISON TABLE (per 100g)", 10)
full_hdrs = ["#","Product","Brand","Wt(g)","kcal","Fat(g)","SatFat","Carbs","Protein","Sodium(mg)"]
for ci, h in enumerate(full_hdrs, 1): hdr2(ws3, cur, ci, h)
cur += 1
for d in nut_data:
    rf = rfill(cur)
    vals = [d["num"], d["product"], d["brand"], d["wt"], d["cal"], d["fat"], d["sat"], d["carbs"], d["protein"], d["sodium"]]
    for ci, v in enumerate(vals, 1):
        c = wr(ws3, cur, ci, v, a=CTR if ci != 2 else WR, fl=rf)
        if ci in {5,6,7,8,9} and v: c.number_format = "0.0"
        if ci == 10 and v: c.number_format = "0"
    cur += 1

ws3.column_dimensions["B"].width = 32
for i in [1,3,4,5,6,7,8,9,10]:
    ws3.column_dimensions[get_column_letter(i)].width = 11

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — DESIGN ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Design Analysis")
ws4.freeze_panes = "C2"
d4h = ["#","Brand","Product","Local Name","Tagline","Primary Color","Secondary Color",
       "Accent Colors","Pack Format","Front Imagery","Design Style","Font Style",
       "Appeal /10","Shelf Vis /10","Target Audience","Texture Claims","Shape",
       "Claims on Pack","USP / Differentiator"]
for ci, h in enumerate(d4h, 1): hdr2(ws4, 1, ci, h)
set_widths(ws4, [4,16,28,22,26,16,16,18,14,24,18,14,11,11,18,18,14,32,30])

for ri, d in enumerate(data, 2):
    rf = rfill(ri)
    dv = [d.get("packet_num"), d.get("brand"),
          d.get("product_name_english") or d.get("product_name_local"),
          d.get("product_name_local"), d.get("tagline_or_slogan"),
          d.get("pack_primary_color"), d.get("pack_secondary_color"), d.get("pack_accent_colors"),
          d.get("pack_format"), d.get("front_imagery"), d.get("design_style"), d.get("font_style"),
          d.get("pack_appeal_score"), d.get("shelf_visibility_score"),
          d.get("target_audience"), d.get("texture_claims"), d.get("product_shape"),
          d.get("claims_on_pack"), d.get("usp_differentiator")]
    for ci, v in enumerate(dv, 1):
        wr(ws4, ri, ci, v, a=CTR if ci in {1,13,14} else WR, fl=rf)
ws4.auto_filter.ref = f"A1:{get_column_letter(len(d4h))}1"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — TRADE MARGINS
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Trade Margins")
ws5.freeze_panes = "C9"

ws5.merge_cells("A1:J1")
ws5["A1"] = "TRADE MARGIN ANALYSIS — Actual distributor data (source: local distributor interview, Mar 2026)"
ws5["A1"].font = fnt(True, "1F4E79", 11)
ws5["A1"].fill = fill("D6E4F0")
ws5["A1"].alignment = CTR

# Brand-level summary
ws5.merge_cells("A3:J3")
ws5["A3"] = "CONFIRMED TRADE CHAIN (per packet, all MRP = Rs 5)"
ws5["A3"].font = fnt(True, "1F4E79", 10)

trade_cols = ["Brand / SKU","Mfr → Dist (Rs)","Dist → Retail (Rs)","MRP (Rs)",
              "Dist Margin (Rs)","Dist Margin %","Retailer Margin (Rs)","Retailer Margin %",
              "Note"]
for ci, h in enumerate(trade_cols, 1): hdr2(ws5, 4, ci, h)
set_widths(ws5, [22,16,18,10,16,14,18,14,44])

brand_rows = [
    ["Balaji / Balaji Wafers", 4.25, 4.50, 5.00, "Standard — fast mover, lean margins"],
    ["Surya",                  3.50, 4.00, 5.00, "High retailer margin (20%) — retailer prefers pushing Surya; lower mfr realization"],
    ["Mung Dal (Balaji)",      4.50, 4.75, 5.00, "Lowest margin element — Rs 0.25 retailer margin (5%), least incentive to push"],
]
for ri, (br, mfr, dist, mrp, note) in enumerate(brand_rows, 5):
    dm = dist - mfr; rm = mrp - dist
    wr(ws5, ri, 1, br, f=fnt(True, size=9), a=CTR)
    wr(ws5, ri, 2, mfr,  f=INPF, a=CTR).number_format = '"Rs"0.00'
    wr(ws5, ri, 3, dist, f=INPF, a=CTR).number_format = '"Rs"0.00'
    wr(ws5, ri, 4, mrp,  f=INPF, a=CTR).number_format = '"Rs"0.00'
    wr(ws5, ri, 5, dm,   f=CALCF, a=CTR).number_format = '"Rs"0.00'
    wr(ws5, ri, 6, dm/dist, f=CALCF, a=CTR).number_format = '0.0%'
    wr(ws5, ri, 7, rm,   f=CALCF, a=CTR).number_format = '"Rs"0.00'
    wr(ws5, ri, 8, rm/mrp, f=CALCF, a=CTR).number_format = '0.0%'
    fl = REDF if "mung" in br.lower() else fill("FFFFFF")
    wr(ws5, ri, 9, note, a=WR, fl=fl)

# Per-SKU table
ws5.merge_cells("A8:J8")
ws5["A8"] = "PER SKU BREAKDOWN"
ws5["A8"].font = fnt(True, "1F4E79", 10)

t5h = ["#","Brand","Product","Net Wt (g)","MRP (Rs)","Price/g (Rs)",
       "Mfr → Dist (Rs)","Dist → Retail (Rs)","Retailer Margin (Rs)","Retailer Margin %"]
for ci, h in enumerate(t5h, 1): hdr2(ws5, 9, ci, h)

for ri, d in enumerate(data, 10):
    rf = rfill(ri)
    brand = d.get("brand","?")
    pname = d.get("product_name_english") or d.get("product_name_local","")
    mrp = d.get("mrp_inr"); wt = d.get("net_weight_g")
    t = get_trade(brand, pname)
    mfr_p = t["mfr"]; dist_p = t["dist"]

    is_moong = "mung" in (pname or "").lower() or "moong" in (pname or "").lower() or "dal" in (pname or "").lower()
    row_fl = REDF if is_moong else rf

    wr(ws5, ri, 1, d.get("packet_num"), a=CTR, fl=row_fl)
    wr(ws5, ri, 2, brand, a=CTR, fl=row_fl)
    wr(ws5, ri, 3, pname, a=WR, fl=row_fl)
    wr(ws5, ri, 4, wt, a=CTR, fl=row_fl).number_format = "0"
    wr(ws5, ri, 5, mrp, f=DATAF, a=CTR, fl=row_fl).number_format = '"Rs"0.00' if mrp else "@"
    if mrp and wt:
        wr(ws5, ri, 6, f"=E{ri}/D{ri}", f=CALCF, a=CTR, fl=row_fl).number_format = "0.000"
    else:
        wr(ws5, ri, 6, None, fl=WARNF)
    wr(ws5, ri, 7, mfr_p, f=DATAF, a=CTR, fl=row_fl).number_format = '"Rs"0.00' if mfr_p else "@"
    wr(ws5, ri, 8, dist_p, f=DATAF, a=CTR, fl=row_fl).number_format = '"Rs"0.00' if dist_p else "@"
    if mrp and dist_p:
        rm = mrp - dist_p
        wr(ws5, ri, 9, rm, f=CALCF, a=CTR, fl=row_fl).number_format = '"Rs"0.00'
        wr(ws5, ri, 10, rm/mrp, f=CALCF, a=CTR, fl=row_fl).number_format = "0.0%"
    else:
        wr(ws5, ri, 9, None, fl=row_fl); wr(ws5, ri, 10, None, fl=row_fl)

ws5.auto_filter.ref = f"A9:{get_column_letter(len(t5h))}9"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — MANUFACTURER REGISTER
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Manufacturer Register")
ws6.freeze_panes = "A2"
m6h = ["#","Brand","Product","Manufacturer","Full Address","City","State",
       "Phone","Email","FSSAI License","Pkg Material Mfg","Veg/Non-Veg","Country","Best Before (mo)","Certifications"]
for ci, h in enumerate(m6h, 1): hdr2(ws6, 1, ci, h)
set_widths(ws6, [4,16,28,26,44,14,14,14,24,22,22,10,10,12,18])
for ri, d in enumerate(data, 2):
    rf = rfill(ri)
    mv = [d.get("packet_num"), d.get("brand"),
          d.get("product_name_english") or d.get("product_name_local"),
          d.get("manufacturer_name"), d.get("manufacturer_address"),
          d.get("manufacturer_city"), d.get("manufacturer_state"),
          d.get("manufacturer_phone"), d.get("manufacturer_email"),
          d.get("fssai_license"), d.get("pkg_material_mfg"),
          d.get("veg_nonveg"), d.get("country_of_origin"),
          d.get("best_before_months"), d.get("certifications")]
    for ci, v in enumerate(mv, 1):
        wr(ws6, ri, ci, v, a=WR if ci in {4,5,10,11} else CTR if ci in {1,12,13,14} else WR, fl=rf)
ws6.auto_filter.ref = f"A1:{get_column_letter(len(m6h))}1"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — PM SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("PM Summary")
ws7.sheet_properties.tabColor = "FF6600"

ws7.merge_cells("A1:F1")
ws7["A1"] = "COMPETITOR RESEARCH — SNACK PACKETS  |  ₹5 Price Point  |  19 SKUs"
ws7["A1"].font = fnt(True, "1F4E79", 14)
ws7["A1"].fill = fill("D6E4F0")
ws7["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws7.row_dimensions[1].height = 28

from collections import defaultdict
brand_count = defaultdict(int)
for d in data: brand_count[d.get("brand","?")] += 1

surya_skus = brand_count.get("Surya", 0)
balaji_skus = sum(v for k,v in brand_count.items() if "balaji" in k.lower())

ws7["A3"] = "QUICK STATS"; ws7["A3"].font = fnt(True, size=11, color="1F4E79")
stats = [
    ("Total SKUs analysed", N),
    ("Distinct brands", "2  (Balaji / Balaji Wafers, Surya)"),
    ("Surya SKUs", surya_skus),
    ("Balaji SKUs", balaji_skus),
    ("MRP (all)", "Rs 5.00"),
    ("Balaji: Mfr → Dist", "Rs 4.25"),
    ("Balaji: Dist → Retail", "Rs 4.50  →  Retailer margin 10%"),
    ("Surya: Mfr → Dist", "Rs 3.50"),
    ("Surya: Dist → Retail", "Rs 4.00  →  Retailer margin 20%"),
    ("Mung Dal: Mfr → Dist", "Rs 4.50"),
    ("Mung Dal: Dist → Retail", "Rs 4.75  →  Retailer margin 5%  ⚠ LOWEST"),
]
for i, (k, v) in enumerate(stats, 4):
    ws7[f"A{i}"] = k; ws7[f"A{i}"].font = fnt(True, size=9)
    ws7[f"B{i}"] = str(v); ws7[f"B{i}"].font = fnt(size=9)

r = 4 + len(stats) + 2
ws7[f"A{r}"] = "KEY COMPETITIVE INSIGHTS"; ws7[f"A{r}"].font = fnt(True, size=11, color="1F4E79")
insights = [
    ("Margin war", "Surya gives retailer 2x the margin (20% vs 10%) — retailers push Surya harder. To compete, target ≥18% retailer margin."),
    ("Mung Dal weakness", "Mung Dal has the lowest retailer margin (5%) — least pushed product. Opportunity to capture with better margin to retailer."),
    ("Edible Oil dominates", "Edible Vegetable Oil / Palmolein is the universal base ingredient across virtually all SKUs — not a differentiator."),
    ("Calories cluster", "Most products 500-570 kcal/100g — tight band. Energy density is not a differentiator in this category."),
    ("Sodium variation", "Sodium varies widely (300-1300mg/100g) — high sodium products are a health risk; low-sodium is an unclaimed positioning."),
    ("Color gap", "Dominant colors: Yellow, Red, Orange. Blue/White/Premium metallic packs = 0 at this price point — shelf standout opportunity."),
    ("Claims gap", "Zero products claim: 'high protein', 'baked', 'no MSG', 'clean label', 'high fiber' — open positioning at Rs5."),
    ("Marathi presence", "Both Surya & Balaji use Devanagari/Marathi on pack — local script essential for this market."),
    ("Weight/Value play", "Balaji 16g packs = Rs 0.31/g vs 24g packs at Rs 0.21/g — 48% worse value. Launch 25g at Rs5 to own value positioning."),
    ("Surya positioning", "Surya leans into Maharashtra-specific themes (Ladki Bahin scheme) — hyper-local cultural marketing at Rs5 is unique."),
]
r += 1
hdr2(ws7, r, 1, "Theme"); hdr2(ws7, r, 2, "Insight")
r += 1
for theme, insight in insights:
    rf = rfill(r)
    ws7.cell(r, 1, theme).font = fnt(True, size=9); ws7.cell(r,1).fill=rf; ws7.cell(r,1).border=THIN; ws7.cell(r,1).alignment=WR
    ws7.cell(r, 2, insight).font = fnt(size=9); ws7.cell(r,2).fill=rf; ws7.cell(r,2).border=THIN; ws7.cell(r,2).alignment=WR
    ws7.row_dimensions[r].height = 30
    r += 1

set_widths(ws7, [22, 80])

# ── SAVE ──────────────────────────────────────────────────────────────────────
out = Path("chips_inventory_v5.xlsx")
wb.save(out)
print(f"Saved: {out}  ({N} SKUs, 7 sheets)")
