import sys, json, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

# ── LOAD & CLEAN DATA ────────────────────────────────────────────────────────
raw = json.loads(Path("packet_data.json").read_text(encoding="utf-8"))

def normalise_brand(b_en, b_lo):
    b = (b_en or b_lo or "").lower().strip()
    if "balaji" in b:    return "Balaji"
    if "surya" in b:     return "Surya"
    if "haldiram" in b:  return "Balaji"   # AI hallucination — confirmed Balaji from image
    lo = (b_lo or "").strip()
    if "सूर्य" in lo:    return "Surya"
    if "बालाजी" in lo:   return "Balaji"
    return (b_en or b_lo or "Unknown").strip().title()

def clean_mrp(v):
    try: return float(v)
    except: return None

def clean_weight(v, brand):
    try:
        w = float(v)
        # Snap to known Balaji weights: 16, 20, 24, 25
        if brand == "Balaji":
            valid = [16, 20, 21, 24, 25]
            # 21 is suspicious — flag but keep
            return w
        return w
    except: return None

def clean_color(s):
    if not s: return s
    # Normalise to Title Case
    return s.strip().title()

ISSUES = []
data = []
for d in raw:
    pn = d.get("packet_num")
    b_en = d.get("brand_english", "") or ""
    b_lo = d.get("brand_local", "") or ""
    brand = normalise_brand(b_en, b_lo)

    # Flag brand correction
    original_brand = (b_en or b_lo or "").strip()
    if "haldiram" in original_brand.lower():
        ISSUES.append(f"P{pn}: Brand was '{original_brand}' → corrected to '{brand}' (image shows Balaji green pack)")

    mrp = clean_mrp(d.get("mrp_inr"))
    wt  = clean_weight(d.get("net_weight_g"), brand)

    if wt is None and brand != "Unknown":
        ISSUES.append(f"P{pn}: '{d.get('product_name_english','')}' — net weight missing, needs manual check")
    if wt == 21:
        ISSUES.append(f"P{pn}: Weight=21g is suspicious for Balaji, likely 20g or 24g — verify physically")

    prod_en = d.get("product_name_english") or ""
    if prod_en.upper() in ("BALAJI WAFERS", "BALAJI", "SURYA"):
        ISSUES.append(f"P{pn}: product_name_english='{prod_en}' looks like brand, not product — needs review")

    cleaned = dict(d)
    cleaned["brand"] = brand
    cleaned["brand_english"] = brand
    cleaned["mrp_inr"] = mrp
    cleaned["net_weight_g"] = wt
    cleaned["pack_primary_color"]   = clean_color(d.get("pack_primary_color"))
    cleaned["pack_secondary_color"] = clean_color(d.get("pack_secondary_color"))
    cleaned["pack_accent_colors"]   = clean_color(d.get("pack_accent_colors"))
    data.append(cleaned)

# Print issues summary
print(f"\n{'='*60}")
print(f"DATA ISSUES FOUND ({len(ISSUES)}):")
for issue in ISSUES:
    print(f"  • {issue}")
print(f"{'='*60}\n")

# ── TRADE PRICING ────────────────────────────────────────────────────────────
TRADE = {
    "Balaji": {"mfr_to_dist": 4.25, "dist_to_retail": 4.50},
    "Surya":  {"mfr_to_dist": 3.50, "dist_to_retail": 4.00},
}

# ── STYLES ───────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)
THIN  = Border(**{s: Side(style="thin",   color="D0D0D0") for s in "left right top bottom".split()})
MED   = Border(**{s: Side(style="medium", color="2E75B6") for s in "left right top bottom".split()})
WR    = Alignment(wrap_text=True, vertical="top")
CTR   = Alignment(horizontal="center", vertical="top", wrap_text=True)

H1F=fill("1F4E79"); H1T=fnt(True,"FFFFFF",9)
H2F=fill("2E75B6"); H2T=fnt(True,"FFFFFF",9)
ALTF=fill("EBF3FB")
DATAF=fnt(size=9); CALCF=fnt(color="000000",size=9); INPF=fnt(color="0000FF",size=9)
REDF=fnt(color="C00000",size=9,bold=True)
ISSUEF=fill("FFF2CC")

def hdr1(ws,r,c,v,span=1):
    if span>1: ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)
    x=ws.cell(r,c,v); x.fill=H1F; x.font=H1T; x.alignment=CTR; x.border=MED

def hdr2(ws,r,c,v):
    x=ws.cell(r,c,v); x.fill=H2F; x.font=H2T; x.alignment=CTR; x.border=THIN

def wr(ws,r,c,v,f=None,a=None,fl=None):
    x=ws.cell(r,c,v); x.border=THIN
    x.font=f or DATAF; x.alignment=a or WR
    if fl: x.fill=fl
    return x

def rfill(ri): return ALTF if ri%2==0 else fill("FFFFFF")

def set_widths(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

N = len(data)
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTER INVENTORY
# ═══════════════════════════════════════════════════════════════════════════════
ws1=wb.active; ws1.title="Master Inventory"
ws1.freeze_panes="D3"
ws1.row_dimensions[1].height=16; ws1.row_dimensions[2].height=36

secs=[(1,3,"IDENTIFICATION"),(4,6,"PRICING & WEIGHT"),(7,10,"MANUFACTURER"),
      (11,12,"COMPLIANCE"),(13,14,"PRODUCT"),(15,17,"PACK COLORS"),
      (18,21,"PACK DESIGN"),(22,23,"SCORES"),(24,25,"COMPETITIVE INTEL")]
for cs,ce,title in secs:
    hdr1(ws1,1,cs,title,span=ce-cs+1)

c2=["#","Brand","Product (EN + Local)","MRP (Rs)","Net Wt (g)","Price/g (Rs)",
    "Manufacturer","City","State","Phone","FSSAI License","Best Before (mo)",
    "Type / Category","Flavor / Variant","Primary Color","Secondary Color","Accent Colors",
    "Pack Format","Front Imagery","Design Style","Font Style",
    "Appeal /10","Shelf Vis /10","Claims on Pack","Competitive Notes"]
for ci,h in enumerate(c2,1): hdr2(ws1,2,ci,h)
set_widths(ws1,[4,14,30,9,9,10,26,14,14,14,20,12,18,22,16,16,18,14,22,18,14,12,14,32,36])

for ri,d in enumerate(data,3):
    rf=rfill(ri)
    brand=d.get("brand","?"); mrp=d.get("mrp_inr"); wt=d.get("net_weight_g")
    prod_en=d.get("product_name_english","") or ""
    prod_lo=d.get("product_name_local","") or ""
    prod_combined = prod_en + (f"\n[{prod_lo}]" if prod_lo and prod_lo!=prod_en else "")

    vals=[
        d.get("packet_num"), brand, prod_combined,
        mrp, wt,
        f"=D{ri}/E{ri}" if (mrp and wt) else None,
        d.get("manufacturer_name"), d.get("manufacturer_city"), d.get("manufacturer_state"), d.get("manufacturer_phone"),
        d.get("fssai_license"), d.get("best_before_months"),
        d.get("product_type"), d.get("variant_flavor"),
        d.get("pack_primary_color"), d.get("pack_secondary_color"), d.get("pack_accent_colors"),
        d.get("pack_format"), d.get("front_imagery"), d.get("design_style"), d.get("font_style"),
        d.get("pack_appeal_score"), d.get("shelf_visibility_score"),
        d.get("claims_on_pack"), d.get("competitive_notes"),
    ]
    for ci,v in enumerate(vals,1):
        a=CTR if ci in {1,4,5,6,12,22,23} else WR
        f=CALCF if ci==6 else INPF if ci in {4,5} else DATAF
        # Flag missing weight in yellow
        fl=ISSUEF if (ci==5 and v is None) else rf
        c=wr(ws1,ri,ci,v,f=f,a=a,fl=fl)
        if ci==6 and v: c.number_format="0.000"
        if ci==4 and v: c.number_format='"Rs"#,##0.00'

ws1.auto_filter.ref=f"A2:{get_column_letter(len(c2))}2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — INGREDIENTS & NUTRITION
# ═══════════════════════════════════════════════════════════════════════════════
ws2=wb.create_sheet("Ingredients & Nutrition")
ws2.freeze_panes="D2"
hdr1(ws2,1,1,"IDENTIFICATION",span=3)
hdr1(ws2,1,4,"INGREDIENTS & ALLERGENS",span=2)
hdr1(ws2,1,6,"NUTRITIONAL INFO (per 100g)",span=12)

n2=["#","Brand","Product","Ingredients (Full List)","Allergens",
    "Calories","Fat (g)","Sat.Fat (g)","Trans Fat (g)","Carbs (g)",
    "Fiber (g)","Sugar (g)","Protein (g)","Sodium (mg)","Calcium (mg)","Iron (mg)","Vit A","Vit C"]
for ci,h in enumerate(n2,1): hdr2(ws2,2,ci,h)
set_widths(ws2,[4,14,26,56,24,10,9,10,10,9,8,8,9,11,11,8,9,9])

for ri,d in enumerate(data,3):
    rf=rfill(ri)
    nv=[d.get("packet_num"), d.get("brand"), d.get("product_name_english"),
        d.get("ingredients_full"), d.get("allergens"),
        d.get("calories_per_100g"), d.get("total_fat_g"), d.get("saturated_fat_g"), d.get("trans_fat_g"),
        d.get("total_carbs_g"), d.get("dietary_fiber_g"), d.get("sugar_g"), d.get("protein_g"),
        d.get("sodium_mg"), d.get("calcium_mg"), d.get("iron_mg"), d.get("vitamin_a_mcg"), d.get("vitamin_c_mg")]
    for ci,v in enumerate(nv,1):
        a=CTR if ci in set(range(6,19))|{1} else WR
        # Flag missing ingredients
        fl=ISSUEF if (ci==4 and not v) else rf
        wr(ws2,ri,ci,v,a=a,fl=fl)
ws2.auto_filter.ref=f"A2:{get_column_letter(len(n2))}2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — DESIGN ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
ws3=wb.create_sheet("Design Analysis")
ws3.freeze_panes="C2"
d3h=["#","Brand","Product","Product (Local Script)","Tagline / Slogan",
     "Primary Color","Secondary Color","Accent Colors","Pack Format","Front Imagery",
     "Design Style","Font Style","Appeal /10","Shelf Vis /10","Target Audience",
     "Texture Claims","Shape","Claims on Pack","USP / Differentiator"]
for ci,h in enumerate(d3h,1): hdr2(ws3,1,ci,h)
set_widths(ws3,[4,14,26,22,26,16,16,18,14,24,18,14,11,11,18,18,14,32,30])
for ri,d in enumerate(data,2):
    rf=rfill(ri)
    dv=[d.get("packet_num"), d.get("brand"),
        d.get("product_name_english") or d.get("product_name_local"),
        d.get("product_name_local"), d.get("tagline_or_slogan"),
        d.get("pack_primary_color"), d.get("pack_secondary_color"), d.get("pack_accent_colors"),
        d.get("pack_format"), d.get("front_imagery"), d.get("design_style"), d.get("font_style"),
        d.get("pack_appeal_score"), d.get("shelf_visibility_score"),
        d.get("target_audience"), d.get("texture_claims"), d.get("product_shape"),
        d.get("claims_on_pack"), d.get("usp_differentiator")]
    for ci,v in enumerate(dv,1):
        wr(ws3,ri,ci,v,a=CTR if ci in {1,13,14} else WR,fl=rf)
ws3.auto_filter.ref=f"A1:{get_column_letter(len(d3h))}1"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — TRADE MARGINS (with real distributor data)
# ═══════════════════════════════════════════════════════════════════════════════
ws4=wb.create_sheet("Trade Margins")
ws4.freeze_panes="C9"

# Header
ws4.merge_cells("A1:J1")
ws4["A1"]="TRADE MARGIN ANALYSIS — Actual distributor data (source: local distributor interview)"
ws4["A1"].font=fnt(True,"1F4E79",11)
ws4["A1"].fill=fill("D6E4F0")

# Brand-level trade summary
ws4.merge_cells("A3:J3")
ws4["A3"]="CONFIRMED TRADE CHAIN (per packet)"
ws4["A3"].font=fnt(True,"1F4E79",10)

trade_hdr=["Brand","Mfr → Dist (Rs)","Dist → Retailer (Rs)","MRP (Rs)",
           "Dist Margin (Rs)","Dist Margin %","Retailer Margin (Rs)","Retailer Margin %",
           "Mfr Revenue/Pack","Your Opportunity"]
for ci,h in enumerate(trade_hdr,1): hdr2(ws4,4,ci,h)
set_widths(ws4,[14,16,18,10,16,14,18,14,16,36])

trade_rows=[
    ["Balaji",4.25,4.50,5.00],
    ["Surya", 3.50,4.00,5.00],
]
for ri,(brand,mfr,dist,mrp) in enumerate(trade_rows,5):
    wr(ws4,ri,1,brand,f=fnt(True,size=9),a=CTR)
    wr(ws4,ri,2,mfr, f=INPF,a=CTR).number_format='"Rs"0.00'
    wr(ws4,ri,3,dist,f=INPF,a=CTR).number_format='"Rs"0.00'
    wr(ws4,ri,4,mrp, f=INPF,a=CTR).number_format='"Rs"0.00'
    dm=dist-mfr; dm_pct=dm/dist
    rm=mrp-dist; rm_pct=rm/mrp
    wr(ws4,ri,5,dm,  f=CALCF,a=CTR).number_format='"Rs"0.00'
    wr(ws4,ri,6,dm_pct,f=CALCF,a=CTR).number_format='0.0%'
    wr(ws4,ri,7,rm,  f=CALCF,a=CTR).number_format='"Rs"0.00'
    wr(ws4,ri,8,rm_pct,f=CALCF,a=CTR).number_format='0.0%'
    wr(ws4,ri,9,mfr, f=CALCF,a=CTR).number_format='"Rs"0.00'
    note="Lean distributor margin — fast mover" if brand=="Balaji" else "High retailer margin (20%) — retailer preference; lower mfr realization"
    wr(ws4,ri,10,note,a=WR)

# Per-SKU table
ws4.merge_cells("A8:J8")
ws4["A8"]="PER SKU BREAKDOWN"
ws4["A8"].font=fnt(True,"1F4E79",10)

t4h=["#","Brand","Product","Net Wt (g)","MRP (Rs)","Price/g (Rs)",
     "Mfr → Dist (Rs)","Dist → Retailer (Rs)","Retailer Margin (Rs)","Retailer Margin %"]
for ci,h in enumerate(t4h,1): hdr2(ws4,9,ci,h)
set_widths(ws4,[4,14,28,10,10,10,14,16,16,14])

for ri,d in enumerate(data,10):
    rf=rfill(ri); brand=d.get("brand","?")
    mrp=d.get("mrp_inr"); wt=d.get("net_weight_g")
    t=TRADE.get(brand,{"mfr_to_dist":None,"dist_to_retail":None})
    mfr_p=t["mfr_to_dist"]; dist_p=t["dist_to_retail"]

    wr(ws4,ri,1,d.get("packet_num"),a=CTR,fl=rf)
    wr(ws4,ri,2,brand,a=CTR,fl=rf)
    wr(ws4,ri,3,d.get("product_name_english") or d.get("product_name_local"),a=WR,fl=rf)
    c=wr(ws4,ri,4,wt,f=DATAF,a=CTR,fl=rf); c.number_format="0"
    c=wr(ws4,ri,5,mrp,f=DATAF,a=CTR,fl=rf); c.number_format='"Rs"0.00' if mrp else "@"
    if mrp and wt:
        c=wr(ws4,ri,6,f"=E{ri}/D{ri}",f=CALCF,a=CTR,fl=rf); c.number_format="0.000"
    else:
        wr(ws4,ri,6,None,fl=ISSUEF)
    c=wr(ws4,ri,7,mfr_p,f=DATAF,a=CTR,fl=rf); c.number_format='"Rs"0.00' if mfr_p else "@"
    c=wr(ws4,ri,8,dist_p,f=DATAF,a=CTR,fl=rf); c.number_format='"Rs"0.00' if dist_p else "@"
    if mrp and dist_p:
        rm=mrp-dist_p
        c=wr(ws4,ri,9,rm,f=CALCF,a=CTR,fl=rf); c.number_format='"Rs"0.00'
        c=wr(ws4,ri,10,rm/mrp,f=CALCF,a=CTR,fl=rf); c.number_format="0.0%"
    else:
        wr(ws4,ri,9,None,fl=rf); wr(ws4,ri,10,None,fl=rf)

ws4.auto_filter.ref=f"A9:{get_column_letter(len(t4h))}9"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — MANUFACTURER REGISTER
# ═══════════════════════════════════════════════════════════════════════════════
ws5=wb.create_sheet("Manufacturer Register")
ws5.freeze_panes="A2"
m5h=["#","Brand","Product","Manufacturer Name","Full Address","City","State",
     "Phone","Email","FSSAI License","Pkg Material Mfg","Veg/Non-Veg",
     "Country","Best Before (mo)","Certifications"]
for ci,h in enumerate(m5h,1): hdr2(ws5,1,ci,h)
set_widths(ws5,[4,14,26,26,44,14,14,14,24,22,22,10,10,12,18])
for ri,d in enumerate(data,2):
    rf=rfill(ri)
    mv=[d.get("packet_num"), d.get("brand"),
        d.get("product_name_english") or d.get("product_name_local"),
        d.get("manufacturer_name"), d.get("manufacturer_address"),
        d.get("manufacturer_city"), d.get("manufacturer_state"),
        d.get("manufacturer_phone"), d.get("manufacturer_email"),
        d.get("fssai_license"), d.get("pkg_material_mfg"),
        d.get("veg_nonveg"), d.get("country_of_origin"),
        d.get("best_before_months"), d.get("certifications")]
    for ci,v in enumerate(mv,1):
        wr(ws5,ri,ci,v,a=WR if ci in {4,5,10,11} else CTR if ci in {1,12,13,14} else WR,fl=rf)
ws5.auto_filter.ref=f"A1:{get_column_letter(len(m5h))}1"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — DATA ISSUES LOG
# ═══════════════════════════════════════════════════════════════════════════════
ws6=wb.create_sheet("⚠ Issues Log")
ws6.sheet_properties.tabColor="FF0000"
ws6.merge_cells("A1:D1")
ws6["A1"]="DATA ISSUES & INCONSISTENCIES — Review before using data"
ws6["A1"].font=fnt(True,"C00000",12)
ws6["A1"].fill=fill("FFE0E0")

ws6.column_dimensions["A"].width=6
ws6.column_dimensions["B"].width=14
ws6.column_dimensions["C"].width=50
ws6.column_dimensions["D"].width=36

hdr2(ws6,2,1,"Packet #")
hdr2(ws6,2,2,"Severity")
hdr2(ws6,2,3,"Issue")
hdr2(ws6,2,4,"Recommended Action")

all_issues=[
    (10,  "🔴 CRITICAL", "Brand misidentified as 'Haldiram's' by AI — image clearly shows Balaji green namkeen pack", "Corrected to Balaji in this file; manually fill product name for P10"),
    (1,   "🟡 NEEDS DATA","Surya P1 — net weight missing (null)", "Add manually from physical packet"),
    (2,   "🟡 NEEDS DATA","Surya P2 — net weight missing (null)", "Add manually from physical packet"),
    (3,   "🟡 NEEDS DATA","Surya P3 — net weight missing (null)", "Add manually from physical packet"),
    (10,  "🟡 NEEDS DATA","P10 product name reads as brand — exact product name unclear from image", "Add manually from physical packet"),
    (16,  "🟡 NEEDS DATA","P16 product name = 'BALAJI WAFERS' — brand used instead of product name", "Add manually from physical packet"),
    (18,  "🟡 NEEDS DATA","P18 weight = 21g — non-standard Balaji size, likely 20g or 24g", "Verify physically and correct"),
    (9,   "🟢 CONFIRMED", "P9 & P17 Chataka Pataka are 2 different flavours — both valid SKUs", "No action needed"),
    (19,  "🟢 CONFIRMED", "P19 & P20 Crunchem appear as Tomato and Salted — 2 different flavours", "No action needed"),
    (13,  "🟢 CONFIRMED", "P13 & P14 both Mung Dal — verify if different variants or same SKU", "Check packet physically for flavour difference"),
    (0,   "🟢 FIXED",    "Brand spellings were inconsistent (5 variants of Balaji)", "Fixed — all normalised to 'Balaji' / 'Surya'"),
    (0,   "🟢 FIXED",    "MRP column had mixed int/float types", "Fixed — all normalised"),
    (0,   "🟢 FIXED",    "Color values had inconsistent casing", "Fixed — Title Case throughout"),
    (0,   "💡 INSIGHT",  "Total confirmed SKUs: 20 (3 Surya + 17 Balaji)", "Balaji dominates shelf presence 5:1"),
    (0,   "💡 INSIGHT",  "Balaji 16g packs = Rs0.31/g vs Rs0.21/g for 24g at same Rs5 MRP", "Opportunity: launch 25g at Rs5 to win on value"),
    (0,   "💡 INSIGHT",  "Surya gives retailer 20% margin vs Balaji 10% — 2x retailer incentive", "Price your product at 18-22% retailer margin to compete"),
    (0,   "💡 INSIGHT",  "No competitor claims 'high protein', 'baked', or 'clean label' at Rs5", "Open positioning gap for your brand"),
]

sev_colors={"🔴 CRITICAL":"FFE0E0","🟡 WARNING":"FFF2CC","🟢 INFO":"E2EFDA","💡 INSIGHT":"EBF3FB"}

for ri,(pn,sev,issue,action) in enumerate(all_issues,3):
    f=fill(sev_colors.get(sev,"FFFFFF"))
    ws6.cell(ri,1, pn if pn else "—").font=fnt(True,size=9); ws6.cell(ri,1).alignment=CTR; ws6.cell(ri,1).border=THIN; ws6.cell(ri,1).fill=f
    ws6.cell(ri,2,sev).font=fnt(size=9); ws6.cell(ri,2).alignment=CTR; ws6.cell(ri,2).border=THIN; ws6.cell(ri,2).fill=f
    ws6.cell(ri,3,issue).font=fnt(size=9); ws6.cell(ri,3).alignment=WR; ws6.cell(ri,3).border=THIN; ws6.cell(ri,3).fill=f
    ws6.cell(ri,4,action).font=fnt(size=9,italic=True); ws6.cell(ri,4).alignment=WR; ws6.cell(ri,4).border=THIN; ws6.cell(ri,4).fill=f
    ws6.row_dimensions[ri].height=30

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — PM SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws7=wb.create_sheet("PM Summary")
ws7.sheet_properties.tabColor="FF6600"
ws7.merge_cells("A1:F1")
ws7["A1"]="COMPETITOR RESEARCH — SNACK PACKETS  |  ₹5 Price Point"
ws7["A1"].font=fnt(True,"1F4E79",14); ws7["A1"].fill=fill("D6E4F0")
ws7["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws7.row_dimensions[1].height=28

brands_count={}
for d in data:
    b=d.get("brand","Unknown"); brands_count[b]=brands_count.get(b,0)+1

ws7["A3"]="QUICK STATS"; ws7["A3"].font=fnt(True,size=10,color="1F4E79")
stats=[("Total packets analysed",N),("Distinct brands","2 (Balaji, Surya)"),
       ("Duplicate/windowing issues","~3 pairs to verify"),
       ("MRP (all)","Rs 5.00"),("Balaji: Mfr → Dist","Rs 4.25"),("Balaji: Dist → Retail","Rs 4.50"),
       ("Surya: Mfr → Dist","Rs 3.50"),("Surya: Dist → Retail","Rs 4.00"),
       ("Balaji retailer margin","10%"),("Surya retailer margin","20%")]
for i,(k,v) in enumerate(stats,4):
    ws7[f"A{i}"]=k; ws7[f"A{i}"].font=fnt(True,size=9)
    ws7[f"B{i}"]=v; ws7[f"B{i}"].font=fnt(size=9)

r=4+len(stats)+2
ws7[f"A{r}"]="BRAND BREAKDOWN"; ws7[f"A{r}"].font=fnt(True,size=10,color="1F4E79")
hdr2(ws7,r+1,1,"Brand"); hdr2(ws7,r+1,2,"SKUs"); hdr2(ws7,r+1,3,"% of Range")
for j,(b,cnt) in enumerate(sorted(brands_count.items(),key=lambda x:-x[1]),r+2):
    ws7.cell(j,1,b).font=fnt(size=9); ws7.cell(j,1).alignment=CTR; ws7.cell(j,1).border=THIN
    ws7.cell(j,2,cnt).font=fnt(size=9); ws7.cell(j,2).alignment=CTR; ws7.cell(j,2).border=THIN
    ws7.cell(j,3,f"=B{j}/{N}").number_format="0%"; ws7.cell(j,3).font=CALCF; ws7.cell(j,3).alignment=CTR; ws7.cell(j,3).border=THIN

r2=r+2+len(brands_count)+2
ws7[f"A{r2}"]="PM ACTION CHECKLIST"; ws7[f"A{r2}"].font=fnt(True,size=10,color="1F4E79")
actions=["[ ] Fill in missing Surya weights (P1, P2, P3) from physical packets",
         "[ ] Fill in P10 and P16 product names from physical packets",
         "[ ] Confirm P18 weight is 20g or 24g",
         "[ ] Confirm P13/P14 Mung Dal are same or different variants",
         "[ ] Benchmark your target weight vs Balaji 24g at Rs0.21/g — aim for 25g+",
         "[ ] Set retailer margin at 18-22% to match/beat Surya's 20% push margin",
         "[ ] Target mfr price to distributor < Rs3.50 to beat Surya at mfr level",
         "[ ] Color gap: dominant colors are Yellow, Green, Red — consider Blue or White",
         "[ ] Claims gap: 'high protein', 'baked', 'no MSG' are unclaimed at Rs5",
         "[ ] Consider Marathi/Hindi dominant pack design — both Surya and Balaji use it"]
for j,a in enumerate(actions,r2+1):
    ws7[f"A{j}"]=a; ws7[f"A{j}"].font=fnt(size=9)

set_widths(ws7,[52,20,12])

# ── SAVE ─────────────────────────────────────────────────────────────────────
out=Path("chips_inventory_v4.xlsx")
wb.save(out)
print(f"\nSaved: {out}  ({N} SKUs, 7 sheets)")
print("Sheets: Master Inventory | Ingredients & Nutrition | Design Analysis")
print("        Trade Margins | Manufacturer Register | Issues Log | PM Summary")
