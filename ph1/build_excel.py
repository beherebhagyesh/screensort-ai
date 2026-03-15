import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

data = json.loads(Path("packet_data.json").read_text(encoding="utf-8"))
N = len(data)

wb = Workbook()

# ── STYLES ───────────────────────────────────────────────────────────────────
def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)
THIN  = Border(**{s: Side(style="thin", color="D0D0D0")
                  for s in ("left","right","top","bottom")})
MED   = Border(**{s: Side(style="medium", color="2E75B6")
                  for s in ("left","right","top","bottom")})

H1F   = fill("1F4E79"); H1FT  = font(True, "FFFFFF", 9)
H2F   = fill("2E75B6"); H2FT  = font(True, "FFFFFF", 9)
SECF  = fill("D6E4F0"); SECFT = font(True, "1F4E79", 9)
ALTF  = fill("EBF3FB")
INPF  = font(color="0000FF", size=9)
CALCF = font(color="000000", size=9)
DATAF = font(size=9)
WR    = Alignment(wrap_text=True, vertical="top")
CTR   = Alignment(horizontal="center", vertical="top", wrap_text=True)

def hdr1(ws, row, col, val, span=1):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=val)
    c.fill=H1F; c.font=H1FT; c.alignment=CTR; c.border=MED

def hdr2(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.fill=H2F; c.font=H2FT; c.alignment=CTR; c.border=THIN

def wr(ws, row, col, val, f=None, a=None, fl=None):
    c = ws.cell(row=row, column=col, value=val)
    c.border = THIN
    c.font   = f  or DATAF
    c.alignment = a or WR
    if fl: c.fill = fl
    return c

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def row_fill(ri):
    return ALTF if ri % 2 == 0 else fill("FFFFFF")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTER INVENTORY
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Master Inventory"
ws1.freeze_panes = "D3"
ws1.row_dimensions[1].height = 16
ws1.row_dimensions[2].height = 36

# Section header row 1
secs = [
    (1,  3,  "IDENTIFICATION"),
    (4,  4,  "PRICING"),
    (5,  5,  "WEIGHT"),
    (6,  6,  "PRICE/g"),
    (7, 10,  "MANUFACTURER"),
    (11, 12, "COMPLIANCE"),
    (13, 14, "PRODUCT DETAIL"),
    (15, 17, "PACK DESIGN — COLORS"),
    (18, 21, "PACK DESIGN — VISUAL"),
    (22, 23, "SHELF SCORES"),
    (24, 25, "COMPETITIVE INTEL"),
]
for (cs, ce, title) in secs:
    hdr1(ws1, 1, cs, title, span=(ce-cs+1))

# Column headers row 2
c2 = [
    "#", "Brand (EN)", "Product (EN)",               # 1-3
    "MRP (Rs)", "Net Wt (g)", "Price/g (Rs)",         # 4-6
    "Manufacturer", "City", "State", "Phone",          # 7-10
    "FSSAI License", "Best Before (mo)",              # 11-12
    "Type / Category", "Flavor / Variant",             # 13-14
    "Primary Color", "Secondary Color", "Accent Colors",# 15-17
    "Pack Format", "Front Imagery", "Design Style", "Font Style", # 18-21
    "Appeal Score /10", "Shelf Visibility /10",        # 22-23
    "Claims on Pack", "Competitive Notes",             # 24-25
]
for ci, h in enumerate(c2, 1):
    hdr2(ws1, 2, ci, h)

set_widths(ws1, [4,18,28,9,9,10,26,14,14,14,20,12,18,22,16,16,16,14,22,18,14,13,15,32,36])

for ri, d in enumerate(data, 3):
    rf = row_fill(ri)
    mrp = d.get("mrp_inr"); wt = d.get("net_weight_g")
    vals = [
        d.get("packet_num"),
        d.get("brand_english") or d.get("brand_local"),
        (d.get("product_name_english") or "") + ("\n["+d.get("product_name_local")+"]" if d.get("product_name_local") else ""),
        mrp, wt,
        f"=D{ri}/E{ri}" if (mrp and wt) else None,
        d.get("manufacturer_name"), d.get("manufacturer_city"), d.get("manufacturer_state"), d.get("manufacturer_phone"),
        d.get("fssai_license"), d.get("best_before_months"),
        d.get("product_type"), d.get("variant_flavor"),
        d.get("pack_primary_color"), d.get("pack_secondary_color"), d.get("pack_accent_colors"),
        d.get("pack_format"), d.get("front_imagery"), d.get("design_style"), d.get("font_style"),
        d.get("pack_appeal_score"), d.get("shelf_visibility_score"),
        d.get("claims_on_pack"), d.get("competitive_notes"),
    ]
    num_cols = {4,5,6,12,22,23}
    for ci, v in enumerate(vals, 1):
        aln = CTR if ci in num_cols | {1} else WR
        f = CALCF if ci == 6 else DATAF
        c = wr(ws1, ri, ci, v, f=f, a=aln, fl=rf)
        if ci == 6 and v: c.number_format = "0.000"
        if ci == 4 and v: c.number_format = '"₹"#,##0.00'

ws1.auto_filter.ref = f"A2:{get_column_letter(len(c2))}2"
ws1.row_dimensions[1].height = 16

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — INGREDIENTS & NUTRITION
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Ingredients & Nutrition")
ws2.freeze_panes = "D2"

hdr1(ws2, 1, 1, "IDENTIFICATION", span=3)
hdr1(ws2, 1, 4, "INGREDIENTS", span=2)
hdr1(ws2, 1, 6, "NUTRITIONAL INFO (per 100g)", span=12)

n2 = ["#","Brand","Product","Ingredients (Full)","Allergens",
      "Calories","Fat (g)","Sat.Fat (g)","Trans Fat (g)",
      "Carbs (g)","Fiber (g)","Sugar (g)","Protein (g)",
      "Sodium (mg)","Calcium (mg)","Iron (mg)","Vit A (mcg)","Vit C (mg)"]
for ci, h in enumerate(n2, 1):
    hdr2(ws2, 2, ci, h)
set_widths(ws2, [4,16,26,52,24,10,9,10,10,9,8,8,9,11,11,8,9,9])

for ri, d in enumerate(data, 3):
    rf = row_fill(ri)
    nv = [
        d.get("packet_num"),
        d.get("brand_english") or d.get("brand_local"),
        d.get("product_name_english") or d.get("product_name_local"),
        d.get("ingredients_full"),
        d.get("allergens"),
        d.get("calories_per_100g"), d.get("total_fat_g"), d.get("saturated_fat_g"), d.get("trans_fat_g"),
        d.get("total_carbs_g"), d.get("dietary_fiber_g"), d.get("sugar_g"), d.get("protein_g"),
        d.get("sodium_mg"), d.get("calcium_mg"), d.get("iron_mg"), d.get("vitamin_a_mcg"), d.get("vitamin_c_mg"),
    ]
    num_c = set(range(6, 19))
    for ci, v in enumerate(nv, 1):
        aln = CTR if ci in num_c | {1} else WR
        wr(ws2, ri, ci, v, a=aln, fl=rf)

ws2.auto_filter.ref = f"A2:{get_column_letter(len(n2))}2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — DESIGN ANALYSIS (PM competitor view)
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Design Analysis")
ws3.freeze_panes = "C2"

d3h = ["#","Brand","Product","Brand (Local)","Product (Local)","Tagline / Slogan",
       "Primary Color","Secondary Color","Accent Colors","Pack Format","Front Imagery",
       "Design Style","Font Style","Appeal /10","Shelf Vis. /10","Target Audience",
       "Texture Claims","Shape","Claims on Pack","USP / Differentiator"]
for ci, h in enumerate(d3h, 1):
    hdr2(ws3, 1, ci, h)
set_widths(ws3, [4,16,26,18,22,26,16,16,16,14,24,18,14,11,11,18,18,14,32,30])

for ri, d in enumerate(data, 2):
    rf = row_fill(ri)
    dv = [
        d.get("packet_num"),
        d.get("brand_english") or d.get("brand_local"),
        d.get("product_name_english") or d.get("product_name_local"),
        d.get("brand_local"), d.get("product_name_local"),
        d.get("tagline_or_slogan"),
        d.get("pack_primary_color"), d.get("pack_secondary_color"), d.get("pack_accent_colors"),
        d.get("pack_format"), d.get("front_imagery"),
        d.get("design_style"), d.get("font_style"),
        d.get("pack_appeal_score"), d.get("shelf_visibility_score"),
        d.get("target_audience"),
        d.get("texture_claims"), d.get("product_shape"),
        d.get("claims_on_pack"), d.get("usp_differentiator"),
    ]
    num_c = {1, 14, 15}
    for ci, v in enumerate(dv, 1):
        aln = CTR if ci in num_c else WR
        wr(ws3, ri, ci, v, a=aln, fl=rf)

ws3.auto_filter.ref = f"A1:{get_column_letter(len(d3h))}1"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — PRICE & TRADE MARGINS
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Price & Trade Margins")
ws4.freeze_panes = "C9"

# Assumptions block
asmp = [
    ("Retailer Margin %",     0.10),
    ("Distributor Margin %",  0.07),
    ("Manufacturer Margin %", 0.20),
    ("GST / Tax Rate %",      0.12),
]
ws4.merge_cells("A1:F1")
ws4["A1"] = "TRADE MARGIN ASSUMPTIONS  (blue cells = change these)"
ws4["A1"].font = Font(bold=True, size=10, color="1F4E79")
for i, (label, val) in enumerate(asmp, 2):
    ws4[f"A{i}"] = label
    ws4[f"A{i}"].font = font(bold=True, size=9)
    ws4[f"B{i}"] = val
    ws4[f"B{i}"].font = INPF
    ws4[f"B{i}"].number_format = "0%"
    ws4[f"B{i}"].fill = fill("FFF2CC")

ws4["A7"] = "NOTE: Margins are estimates. Update with actual trade terms."
ws4["A7"].font = font(italic=True, color="808080", size=8)

# Table headers row 9
t4h = ["#","Brand","Product","MRP (Rs)","Net Wt (g)","Price/g",
       "Retailer Gets (Rs)","Distributor Gets (Rs)","Mfr Revenue (Rs)",
       "GST (Rs)","Mfr Net (Rs)","Mfr Cost Est (Rs)","Gross Margin %"]
for ci, h in enumerate(t4h, 1):
    hdr2(ws4, 9, ci, h)
set_widths(ws4, [4,16,26,10,10,10,16,18,16,10,14,16,14])

for ri, d in enumerate(data, 10):
    rf = row_fill(ri)
    mrp = d.get("mrp_inr"); wt = d.get("net_weight_g")
    base_vals = [
        d.get("packet_num"),
        d.get("brand_english") or d.get("brand_local"),
        d.get("product_name_english") or d.get("product_name_local"),
        mrp, wt,
        f"=D{ri}/E{ri}" if (mrp and wt) else None,
    ]
    for ci, v in enumerate(base_vals, 1):
        c = wr(ws4, ri, ci, v, f=INPF if ci in {4,5} else CALCF if ci==6 else DATAF, a=CTR if ci in {1,4,5,6} else WR, fl=rf)
        if ci == 6 and v: c.number_format = "0.000"
        if ci == 4 and v: c.number_format = '"₹"#,##0.00'

    if mrp:
        formulas = [
            f"=D{ri}*$B$2",          # retailer
            f"=D{ri}*$B$3",          # distributor
            f"=D{ri}*(1-$B$2-$B$3)", # mfr revenue
            f"=H{ri}*$B$4",          # gst
            f"=H{ri}-I{ri}",         # mfr net
            f"=J{ri}*(1-$B$4)",      # mfr cost est
            f"=(J{ri}-L{ri})/J{ri}", # gross margin %
        ]
        for ci, formula in enumerate(formulas, 7):
            c = wr(ws4, ri, ci, formula, f=CALCF, a=CTR, fl=rf)
            if ci == 13:
                c.number_format = "0.0%"
            else:
                c.number_format = '"₹"0.00'

ws4.auto_filter.ref = f"A9:{get_column_letter(len(t4h))}9"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — MANUFACTURER REGISTER
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Manufacturer Register")
ws5.freeze_panes = "A2"

m5h = ["#","Brand","Product","Manufacturer Name","Full Address","City","State",
       "Phone","Email","FSSAI License","Pkg Material Mfg","Veg/Non-Veg",
       "Country","Best Before (mo)","Batch No.","Certifications"]
for ci, h in enumerate(m5h, 1):
    hdr2(ws5, 1, ci, h)
set_widths(ws5, [4,16,26,26,44,14,14,14,24,22,22,10,10,12,14,18])

for ri, d in enumerate(data, 2):
    rf = row_fill(ri)
    mv = [
        d.get("packet_num"),
        d.get("brand_english") or d.get("brand_local"),
        d.get("product_name_english") or d.get("product_name_local"),
        d.get("manufacturer_name"), d.get("manufacturer_address"),
        d.get("manufacturer_city"), d.get("manufacturer_state"),
        d.get("manufacturer_phone"), d.get("manufacturer_email"),
        d.get("fssai_license"), d.get("pkg_material_mfg"),
        d.get("veg_nonveg"), d.get("country_of_origin"),
        d.get("best_before_months"), d.get("batch_no"), d.get("certifications"),
    ]
    wide = {4,5,10,11}
    for ci, v in enumerate(mv, 1):
        wr(ws5, ri, ci, v, a=WR if ci in wide else CTR if ci in {1,12,13,14} else WR, fl=rf)

ws5.auto_filter.ref = f"A1:{get_column_letter(len(m5h))}1"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — PM SUMMARY DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("PM Summary")
ws6.sheet_properties.tabColor = "FF6600"

ws6.merge_cells("A1:H1")
ws6["A1"] = "COMPETITOR RESEARCH — SNACK PACKETS SUMMARY"
ws6["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws6["A1"].fill = fill("D6E4F0")
ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 28

# Quick stats
brands = {}
for d in data:
    b = d.get("brand_english") or d.get("brand_local") or "Unknown"
    brands[b] = brands.get(b, 0) + 1

ws6["A3"] = "QUICK STATS"; ws6["A3"].font = font(bold=True, size=10, color="1F4E79")
stats = [
    ("Total SKUs analysed", N),
    ("Unique brands", len(brands)),
    ("Price point (all)", "₹5 MRP"),
    ("Avg pack weight (Balaji)", "~24g"),
]
for i, (k, v) in enumerate(stats, 4):
    ws6[f"A{i}"] = k; ws6[f"A{i}"].font = font(bold=True, size=9)
    ws6[f"B{i}"] = v; ws6[f"B{i}"].font = font(size=9)

# Brand breakdown
ws6["A9"] = "BRAND BREAKDOWN"; ws6["A9"].font = font(bold=True, size=10, color="1F4E79")
hdr2(ws6, 10, 1, "Brand")
hdr2(ws6, 10, 2, "# SKUs")
hdr2(ws6, 10, 3, "% of Range")
for i, (b, cnt) in enumerate(sorted(brands.items(), key=lambda x: -x[1]), 11):
    ws6.cell(i, 1, b).font = font(size=9)
    ws6.cell(i, 2, cnt).font = font(size=9)
    ws6.cell(i, 2).alignment = CTR
    ws6.cell(i, 3, f"=B{i}/{N}").number_format = "0%"
    ws6.cell(i, 3).font = CALCF
    ws6.cell(i, 3).alignment = CTR

# PM action checklist
r = 11 + len(brands) + 2
ws6[f"A{r}"] = "PM ACTIONS / GAP ANALYSIS"
ws6[f"A{r}"].font = font(bold=True, size=10, color="1F4E79")
actions = [
    "[ ] Map all Rs5 SKUs — check if any competitor has 25g+ at same price",
    "[ ] Color audit — identify under-used colors for differentiation",
    "[ ] Nutrition benchmarking — identify low-sodium / high-protein gap",
    "[ ] Ingredients — check for any 'clean label' gaps (no MSG, no artificial flavors)",
    "[ ] Design — note which brands use Hindi/Marathi vs English dominant packs",
    "[ ] FSSAI check — verify all manufacturers have valid licenses",
    "[ ] Manufacturer mapping — identify if Balaji uses contract manufacturers",
    "[ ] Price/gram — set target at 10-15% better value than market avg",
    "[ ] Pack appeal scores — design brief should target score 8+",
    "[ ] Claims gap — identify any claim no competitor is making",
]
for j, a in enumerate(actions, r+1):
    ws6[f"A{j}"] = a
    ws6[f"A{j}"].font = font(size=9)

set_widths(ws6, [50, 12, 12])

# ── SAVE ─────────────────────────────────────────────────────────────────────
out = Path("chips_inventory_v2.xlsx")
wb.save(out)
print(f"Saved: {out}  ({N} products, 6 sheets)")
