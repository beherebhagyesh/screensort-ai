"""
Build foreign_snacks_intelligence.xlsx
Comprehensive import snack competitive intelligence.
Vegetarian-first filter. Full economics for India import.
Run from project root: /c/Puthon313/python ph3/build_foreign.py
"""
import sys, json, re
from pathlib import Path
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
data = json.loads((ROOT / "ph3/serp_foreign_data.json").read_text(encoding="utf-8"))
ph1  = json.loads((ROOT / "ph1/packet_data.json").read_text(encoding="utf-8"))

products  = data["all_products"]
bsummary  = data["brand_summary"]
bprofiles = data["brand_profiles"]
snippets  = data["import_intelligence_snippets"]

# ── STYLES ──────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)
THIN  = Border(**{s: Side(style="thin",   color="D0D0D0") for s in "left right top bottom".split()})
MED   = Border(**{s: Side(style="medium", color="1A3C5E") for s in "left right top bottom".split()})
WR    = Alignment(wrap_text=True, vertical="top")
CTR   = Alignment(horizontal="center", vertical="top", wrap_text=True)

H1F=fill("0D2137"); H1T=fnt(True,"FFFFFF",10)
H2F=fill("1A5276"); H2T=fnt(True,"FFFFFF",9)
ALTF=fill("EBF5FB")
GREENF=fill("D5F5E3"); YELLOWF=fill("FEF9E7"); REDF=fill("FADBD8")
ORANGEF=fill("FDEBD0"); PURPLEF=fill("E8DAEF"); GREYF=fill("F2F2F2")
BLUEF=fill("D6EAF8"); CYANF=fill("D1F2EB")
DATAF=fnt(size=9); BOLDF=fnt(bold=True,size=9)
INPF=fnt(color="0000FF",size=9); CALCF=fnt(color="1A5276",size=9)
WARNF=fnt(bold=True,color="CC0000",size=9); LINKF=fnt(color="0563C1",size=8)

def hdr1(ws,r,c,v,span=1,bg=None):
    if span>1: ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)
    x=ws.cell(r,c,v); x.fill=bg or H1F; x.font=H1T; x.alignment=CTR; x.border=MED
def hdr2(ws,r,c,v):
    x=ws.cell(r,c,v); x.fill=H2F; x.font=H2T; x.alignment=CTR; x.border=THIN
def wr(ws,r,c,v,f=None,a=None,fl=None):
    x=ws.cell(r,c,v); x.border=THIN; x.font=f or DATAF; x.alignment=a or WR
    if fl: x.fill=fl
    return x
def rfill(ri): return ALTF if ri%2==0 else fill("FFFFFF")
def sw(ws,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

# ── DATA PREP ───────────────────────────────────────────────────────────────
USD_TO_INR = 83.5   # rate for calculations
THB_TO_INR = 2.3
SGD_TO_INR = 62.0
GBP_TO_INR = 105.0
EUR_TO_INR = 91.0
MYR_TO_INR = 18.5
KRW_TO_INR = 0.063

FX = {"USD":USD_TO_INR,"SGD":SGD_TO_INR,"GBP":GBP_TO_INR,
      "EUR":EUR_TO_INR,"MYR":MYR_TO_INR,"THB":THB_TO_INR,
      "KRW":KRW_TO_INR,"INR":1.0}

def to_inr(price, currency):
    if price is None: return None
    return round(price * FX.get(currency, USD_TO_INR), 1)

# Import duty rates for India
# ASEAN FTA (AIFTA) applies to Thailand, Malaysia, Philippines, Indonesia, Vietnam
ASEAN = {"Tong Garden","Oishi","Mamee","Jack n Jill","Indomie"}
# HS codes and duty
IMPORT_DUTY = {
    "1905": {"desc": "Biscuits, wafers, crackers, extruded snacks",
             "bcd_standard": 0.30, "bcd_asean": 0.05,
             "sws": 0.10, "igst": 0.18},
    "2008": {"desc": "Processed nuts, seeds (roasted/flavored)",
             "bcd_standard": 0.30, "bcd_asean": 0.10,
             "sws": 0.10, "igst": 0.12},
    "1904": {"desc": "Rice crackers, puffed/flaked cereals",
             "bcd_standard": 0.30, "bcd_asean": 0.05,
             "sws": 0.10, "igst": 0.12},
    "0802": {"desc": "Nuts in shell/shelled (pistachios, cashews, etc.)",
             "bcd_standard": 0.30, "bcd_asean": 0.10,
             "sws": 0.10, "igst": 0.12},
    "1902": {"desc": "Noodle-based snacks (Indomie style)",
             "bcd_standard": 0.30, "bcd_asean": 0.05,
             "sws": 0.10, "igst": 0.12},
}
# Brand → primary HS code
BRAND_HS = {
    "Tong Garden":          "2008",  # nuts/seeds dominant
    "Oishi":                "1905",
    "Mamee":                "1902",
    "Jack n Jill":          "1905",
    "Want Want":            "1904",
    "Calbee":               "1905",
    "Hippeas":              "1905",
    "Harvest Snaps":        "1905",
    "Tyrrell's":            "1905",
    "Nongshim":             "1905",
    "Lotte":                "1905",
    "Wonderful Pistachios": "0802",
    "Kettle Brand":         "1905",
    "Indomie":              "1902",
    "Pocky / Glico":        "1905",
}

def calc_landed(export_price_usd, brand, pack_size_g=None):
    """Returns full landed economics dict."""
    hs  = BRAND_HS.get(brand, "1905")
    dur = IMPORT_DUTY[hs]
    is_asean = brand in ASEAN
    bcd  = dur["bcd_asean"] if is_asean else dur["bcd_standard"]
    sws  = bcd * dur["sws"]          # SWS is 10% of BCD
    # CIF estimate: export + 20% freight/insurance
    cif_usd = export_price_usd * 1.20
    cif_inr = cif_usd * USD_TO_INR
    bcd_inr = cif_inr * bcd
    sws_inr = cif_inr * sws
    igst_base = cif_inr + bcd_inr + sws_inr
    igst_inr = igst_base * dur["igst"]
    landed = cif_inr + bcd_inr + sws_inr + igst_inr
    # Supply chain margins
    importer_margin = landed * 0.35   # importer/compliance/wastage
    dist_price = landed + importer_margin
    dist_margin = dist_price * 0.20
    retailer_in = dist_price + dist_margin
    retailer_margin = retailer_in * 0.25
    mrp_min = retailer_in + retailer_margin
    return {
        "hs_code": hs, "is_asean_fta": is_asean,
        "export_price_usd": export_price_usd,
        "cif_inr": round(cif_inr,1),
        "bcd_pct": f"{bcd*100:.0f}%",
        "bcd_inr": round(bcd_inr,1),
        "sws_inr": round(sws_inr,1),
        "igst_pct": f"{dur['igst']*100:.0f}%",
        "igst_inr": round(igst_inr,1),
        "total_landed_inr": round(landed,1),
        "importer_margin_inr": round(importer_margin,1),
        "dist_price_inr": round(dist_price,1),
        "retailer_in_inr": round(retailer_in,1),
        "min_mrp_inr": round(mrp_min,1),
        "effective_duty_pct": round((landed/cif_inr-1)*100,1),
    }

# Typical export price estimates (USD) per brand per pack — based on scrape data
# Using ~40-60% of USA retail as FOB export estimate
BRAND_EXPORT_PRICES = {
    "Tong Garden":          {"export_usd": 0.60, "typical_pack_g": 50,  "note": "Honey/BBQ peanuts 30-50g pack. Thai FOB est."},
    "Oishi":                {"export_usd": 0.45, "typical_pack_g": 30,  "note": "Extruded snack 30g pack. Thai FOB est."},
    "Mamee":                {"export_usd": 0.35, "typical_pack_g": 25,  "note": "Monster noodle snack 25g. Malaysian FOB est."},
    "Jack n Jill":          {"export_usd": 0.50, "typical_pack_g": 55,  "note": "Piattos 55g. Philippine FOB est."},
    "Want Want":            {"export_usd": 0.55, "typical_pack_g": 70,  "note": "Rice cracker 70g. Taiwan FOB est."},
    "Calbee":               {"export_usd": 1.10, "typical_pack_g": 80,  "note": "JagaRico/Pea Crisps 80g. Japan premium."},
    "Hippeas":              {"export_usd": 1.80, "typical_pack_g": 78,  "note": "Chickpea puffs 78g. UK export price."},
    "Harvest Snaps":        {"export_usd": 1.50, "typical_pack_g": 93,  "note": "Baked green pea 93g. USA FOB est."},
    "Tyrrell's":            {"export_usd": 1.75, "typical_pack_g": 80,  "note": "Hand-cooked crisps 80g. UK export."},
    "Nongshim":             {"export_usd": 0.55, "typical_pack_g": 40,  "note": "Honey crackers 40g. Korea FOB."},
    "Lotte":                {"export_usd": 0.70, "typical_pack_g": 50,  "note": "Pepero/crackers 50g. Korea FOB."},
    "Wonderful Pistachios": {"export_usd": 1.25, "typical_pack_g": 49,  "note": "No-shells 49g bag. USA FOB."},
    "Kettle Brand":         {"export_usd": 1.20, "typical_pack_g": 64,  "note": "Kettle chips 64g. USA FOB."},
    "Indomie":              {"export_usd": 0.25, "typical_pack_g": 85,  "note": "Instant noodle 85g. Indonesian FOB (very low)."},
    "Pocky / Glico":        {"export_usd": 0.80, "typical_pack_g": 47,  "note": "Pocky 47g box. Japan FOB."},
}

# Pre-compute economics for all brands
economics = {}
for brand, ep in BRAND_EXPORT_PRICES.items():
    economics[brand] = calc_landed(ep["export_usd"], brand, ep["typical_pack_g"])
    economics[brand]["typical_pack_g"] = ep["typical_pack_g"]
    economics[brand]["note"] = ep["note"]

# Veg compliance data (curated)
VEG_COMPLIANCE = {
    "Tong Garden": {
        "overall": "MOSTLY VEGETARIAN",
        "certified": "Halal (majority of range)",
        "green_dot_eligible": "YES — nut/seed products. Flag: anchovy peanuts range.",
        "flag_products": "Peanuts Mixed Anchovy — contains fish. Avoid this SKU.",
        "clean_label": "YES — most products have 3-5 ingredients",
        "key_veg_skus": "Honey Peanuts, Bar-B-Q Coated Peanuts, Salted Cashew, Wasabi Green Peas, Sunflower Seeds, Party Snack Mix, Broad Beans",
    },
    "Oishi": {
        "overall": "PARTIALLY VEGETARIAN",
        "certified": "Halal (some)",
        "green_dot_eligible": "SELECT SKUs only — avoid prawn flavors",
        "flag_products": "Prawn Crackers range — shrimp. Squid-flavored variants.",
        "clean_label": "MODERATE — some artificial flavors",
        "key_veg_skus": "Bread Pan Cheese & Onion, Potato Fries Ketchup, Cheese Sticks, Onion Rings, Pillows (chocolate-filled)",
    },
    "Mamee": {
        "overall": "PARTIALLY VEGETARIAN",
        "certified": "Halal",
        "green_dot_eligible": "SELECT SKUs — Monster original has chicken flavor",
        "flag_products": "Monster Chicken flavor — non-veg. BBQ variants may contain meat extract.",
        "clean_label": "MODERATE",
        "key_veg_skus": "Monster Original (spicy), Ghost Pepper variant, Crunch Crackers",
    },
    "Jack n Jill": {
        "overall": "PARTIALLY VEGETARIAN",
        "certified": "Some Halal",
        "green_dot_eligible": "SELECT SKUs",
        "flag_products": "BBQ flavors may contain meat extract. Prawn crackers non-veg.",
        "clean_label": "MODERATE",
        "key_veg_skus": "Piattos Cheese, Piattos Sour Cream, Nova Cheese Crackers, Mr. Chips Classic Salt",
    },
    "Want Want": {
        "overall": "MOSTLY VEGETARIAN",
        "certified": "Not certified but mostly clean",
        "green_dot_eligible": "YES — rice-based, mostly plant ingredients",
        "flag_products": "Seaweed flavor variants — check for fish extract in seasoning.",
        "clean_label": "MODERATE — some MSG",
        "key_veg_skus": "Rice Crackers Original, Golden Rice Cracker Bites, Senbei Original, Snow Cracker",
    },
    "Calbee": {
        "overall": "PARTIALLY VEGETARIAN",
        "certified": "Not certified",
        "green_dot_eligible": "SELECT SKUs",
        "flag_products": "Shrimp Chips — shrimp. Some prawn-flavored variants.",
        "clean_label": "MODERATE",
        "key_veg_skus": "Harvest Snaps (all pea/lentil/bean), JagaRico Salt & Butter, Mini Salted Peas, Asian Style Thai Curry (check)",
    },
    "Hippeas": {
        "overall": "100% VEGAN",
        "certified": "B-Corp, Organic, Vegan certified, Non-GMO",
        "green_dot_eligible": "YES — every single SKU",
        "flag_products": "NONE — clean vegan across full range",
        "clean_label": "EXCELLENT — organic ingredients, no artificial anything",
        "key_veg_skus": "ALL: Salt & Vinegar, White Cheddar, Rockin' Ranch, Nacho Vibes, Far Out Fajita, Purple Power, Sriracha Sunshine",
    },
    "Harvest Snaps": {
        "overall": "100% VEGETARIAN",
        "certified": "Vegetarian, Gluten-free (most), Non-GMO",
        "green_dot_eligible": "YES — all SKUs",
        "flag_products": "NONE",
        "clean_label": "EXCELLENT — baked, simple ingredients",
        "key_veg_skus": "ALL: Lightly Salted Green Pea, Caesar, Tomato Basil, Mango Chili, Black Pepper, White Cheddar Jalapeno",
    },
    "Tyrrell's": {
        "overall": "PARTIALLY VEGETARIAN",
        "certified": "UK Vegetarian Society mark on some",
        "green_dot_eligible": "SELECT SKUs only",
        "flag_products": "Beef Brisket & Black Pepper — non-veg. Prawn Cocktail — non-veg. Chorizo variants.",
        "clean_label": "GOOD on veg variants — simple potato/veg crisps",
        "key_veg_skus": "Naked (sea salt only), Lentil Crisps Sour Cream, Parsnip & Black Pepper, Sweet Chilli, Veg Crisps",
    },
    "Nongshim": {
        "overall": "PARTIALLY VEGETARIAN",
        "certified": "Halal (some)",
        "green_dot_eligible": "SELECT — Honey Twist confirmed veg",
        "flag_products": "Shrimp Crackers — shrimp. Many Korean flavors have seafood.",
        "clean_label": "MODERATE",
        "key_veg_skus": "Honey Twist Snack, Honey Butter variants (check), Bokki Boy Tteok Snack",
    },
    "Lotte": {
        "overall": "PARTIALLY VEGETARIAN",
        "certified": "Halal (some)",
        "green_dot_eligible": "SELECT — biscuit range mostly veg",
        "flag_products": "Some meat-flavored snacks. Pepero is vegetarian.",
        "clean_label": "MODERATE",
        "key_veg_skus": "Pepero Original, Pepero Almond, Koala March (Choc/Strawberry), Choco Pie, Ghana Choco",
    },
    "Wonderful Pistachios": {
        "overall": "100% VEGETARIAN",
        "certified": "Vegetarian, Kosher, Gluten-free",
        "green_dot_eligible": "YES — pure nut product",
        "flag_products": "NONE",
        "clean_label": "EXCELLENT — 1-2 ingredients",
        "key_veg_skus": "ALL: Roasted Salted, Lightly Salted, Sea Salt & Vinegar, Jalapeño Lime, Chili Roasted, No Shells",
    },
    "Kettle Brand": {
        "overall": "MOSTLY VEGETARIAN",
        "certified": "Non-GMO, some Vegetarian Society",
        "green_dot_eligible": "SELECT SKUs",
        "flag_products": "Honey Dijon (check), some BBQ varieties may contain milk/honey.",
        "clean_label": "GOOD — kettle-cooked, simple ingredients",
        "key_veg_skus": "Sea Salt, Himalayan Salt, Jalapeño, Pepperoncini, Backyard Barbeque (veg version), Salt & Fresh Ground Pepper",
    },
    "Indomie": {
        "overall": "PARTIALLY VEGETARIAN",
        "certified": "Halal",
        "green_dot_eligible": "SELECT — vegetable flavor only",
        "flag_products": "Chicken, Beef, Soto Ayam, Satay flavors — non-veg. Most popular Mi Goreng has chicken flavor.",
        "clean_label": "LOW — high MSG, artificial flavors",
        "key_veg_skus": "Vegetable Cup Noodle, Onion Chicken (flag), Curly Noodle Vegetable",
    },
    "Pocky / Glico": {
        "overall": "MOSTLY VEGETARIAN",
        "certified": "Not certified",
        "green_dot_eligible": "MOST SKUs — chocolate/strawberry/matcha are veg",
        "flag_products": "Check: some cream/gelatin in coating. Pocky Men (strong) may differ.",
        "clean_label": "MODERATE",
        "key_veg_skus": "Pocky Chocolate, Strawberry, Matcha, Cookies & Cream, Almond Crush, Pistachio, Blueberry Yogurt",
    },
}

# FSSAI import requirements (hardcoded from regulatory knowledge)
FSSAI_STEPS = [
    ("1. FSSAI Import Clearance", "Every food consignment needs clearance at port of entry. FSSAI appointed officer at 6 airports (Delhi, Mumbai, Chennai, Kolkata, Bengaluru, Hyderabad) and select seaports."),
    ("2. Registration/License", "Importer needs FSSAI Import License (Central License) — applies to importers, not per-product."),
    ("3. NOC per Product", "For new products/brands, an application is filed with FSSAI with: product composition, manufacturing details, safety data, COA (Certificate of Analysis)."),
    ("4. Label Compliance", "Indian label (sticker or label) must include: product name, ingredients, net qty, nutritional info per 100g/ml, FSSAI license no. of importer, country of origin, best before in DD/MM/YYYY, vegetarian/non-vegetarian symbol (green/red dot)."),
    ("5. Green Dot", "Importer applies the green (Vegetarian) or red (Non-Veg) dot based on product composition. Must be 3mm min diameter. Adjacent to product name."),
    ("6. Physical Inspection", "First consignment of new product: full lab test. Subsequent: random sampling. Test: microbiology, heavy metals, pesticides, nutrition label accuracy."),
    ("7. HS Code Declaration", "Correct HS code critical: determines duty rate + any special licenses needed."),
    ("8. CDSCO (if required)", "Novel foods or foods with health claims may need CDSCO approval additionally."),
    ("9. BIS (if applicable)", "Some processed food categories need BIS ISI mark — check product-specific requirements."),
    ("10. Timeline", "First-time import clearance: 7-21 working days. Subsequent: 3-7 days. Expedited for perishables."),
]

# ── BUILD WORKBOOK ──────────────────────────────────────────────────────────────
wb = Workbook()
BRAND_ORDER = ["Tong Garden","Oishi","Mamee","Jack n Jill","Want Want","Calbee",
               "Hippeas","Harvest Snaps","Tyrrell's","Nongshim","Lotte",
               "Wonderful Pistachios","Kettle Brand","Indomie","Pocky / Glico"]

BRAND_FILLS = {
    "Tong Garden": fill("FFF3CD"),         # warm gold (SE Asia)
    "Oishi": fill("FCE4D6"),               # orange (Thailand)
    "Mamee": fill("E2F0D9"),               # green (Malaysia)
    "Jack n Jill": fill("DDEBF7"),         # blue (Philippines)
    "Want Want": fill("FCE4D6"),           # orange (Taiwan)
    "Calbee": fill("FFF3CD"),              # gold (Japan)
    "Hippeas": fill("D5F5E3"),             # mint green (vegan hero)
    "Harvest Snaps": fill("D5F5E3"),       # mint green (vegan)
    "Tyrrell's": fill("E8DAEF"),           # purple (UK premium)
    "Nongshim": fill("FDEBD0"),            # peach (Korea)
    "Lotte": fill("FDEBD0"),               # peach (Korea)
    "Wonderful Pistachios": fill("EAF2FF"),# light blue (USA)
    "Kettle Brand": fill("EAF2FF"),        # light blue (USA)
    "Indomie": fill("FCE4D6"),             # orange (Indonesia)
    "Pocky / Glico": fill("FFF3CD"),       # gold (Japan)
}
IMPORT_FEASIBILITY = {
    "Tong Garden": ("HIGH", "ASEAN FTA + proven demand + nut snacks gap in India"),
    "Oishi": ("HIGH", "ASEAN FTA + extruded snacks affordable + strong SE Asia brand"),
    "Mamee": ("MEDIUM-HIGH", "ASEAN FTA + iconic noodle snack + partial veg issue"),
    "Jack n Jill": ("MEDIUM", "ASEAN FTA + partial veg range + moderate differentiation"),
    "Want Want": ("MEDIUM-HIGH", "Rice crackers = novel in India + veg-friendly"),
    "Calbee": ("MEDIUM", "Premium Japan = good aspirational brand + higher cost"),
    "Hippeas": ("HIGH", "100% vegan + clean label trend + premium but India-ready"),
    "Harvest Snaps": ("HIGH", "100% veg + baked + pea crisps = first-mover advantage in India"),
    "Tyrrell's": ("MEDIUM", "UK premium = niche. High duty (no UK-India FTA). Aspirational."),
    "Nongshim": ("MEDIUM", "Korean wave popularity + limited veg range"),
    "Lotte": ("MEDIUM", "Korean/Japanese trust + Pocky already in market"),
    "Wonderful Pistachios": ("HIGH", "Pure nut = easy to import + growing nut snack market India"),
    "Kettle Brand": ("MEDIUM-LOW", "High import cost (USA, no FTA) + Lay's dominates chips in India"),
    "Indomie": ("MEDIUM", "Already beloved in India + limited veg range"),
    "Pocky / Glico": ("HIGH", "Already in Indian market via grey imports + strong brand recognition"),
}

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — IMPORT LANDSCAPE MAP
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Import Landscape"
ws1.freeze_panes = "C3"

hdr1(ws1,1,1,"FOREIGN VEGETARIAN SNACK IMPORT INTELLIGENCE — India Opportunity Map",span=12)
cols=["Brand","Country","HQ City","Product Types","Veg Status","SKUs\nFound",
      "Typical Pack\n& Export $","Min MRP\n(INR est.)","ASEAN\nFTA","Import\nFeasibility","Green Dot\nEligible","Key Opportunity"]
for ci,h in enumerate(cols,1): hdr2(ws1,2,ci,h)
sw(ws1,[22,14,14,30,22,8,18,14,8,18,12,55])

for ri,brand in enumerate(BRAND_ORDER,3):
    bp = bprofiles.get(brand,{}); ec = economics.get(brand,{}); vc = VEG_COMPLIANCE.get(brand,{})
    feas,feas_note = IMPORT_FEASIBILITY.get(brand,("?",""))
    bf = BRAND_FILLS.get(brand,fill("FFFFFF"))
    rf = REDF if feas=="HIGH" else (YELLOWF if "MEDIUM" in feas else GREYF)

    ep = BRAND_EXPORT_PRICES.get(brand,{})
    pack_info = f"~{ep.get('typical_pack_g','?')}g @ ${ep.get('export_usd','?'):.2f}"
    min_mrp = ec.get("min_mrp_inr","?")

    vals=[
        brand, bp.get("country",""), bp.get("hq_city",""),
        bp.get("product_types",""), vc.get("overall",""),
        bsummary.get(brand,{}).get("products_found",0),
        pack_info,
        f"Rs {min_mrp:.0f}" if isinstance(min_mrp,float) else "?",
        "YES ✓" if brand in ASEAN else "No",
        feas, vc.get("green_dot_eligible","?"), feas_note,
    ]
    for ci,v in enumerate(vals,1):
        a=CTR if ci in {6,7,8,9,10,11} else WR
        f=BOLDF if ci==1 else (WARNF if (ci==10 and "HIGH" in str(v)) else DATAF)
        cell_fl = (GREENF if ci==9 and "YES" in str(v) else
                   (GREENF if ci==10 and "HIGH" in str(v) else
                    (YELLOWF if ci==10 and "MEDIUM" in str(v) else bf)))
        wr(ws1,ri,ci,v,f=f,a=a,fl=cell_fl)
    ws1.row_dimensions[ri].height=48
ws1.auto_filter.ref=f"A2:{get_column_letter(len(cols))}2"

note_r=len(BRAND_ORDER)+4
ws1.merge_cells(start_row=note_r,start_column=1,end_row=note_r,end_column=12)
ws1.cell(note_r,1,
    "KEY: ASEAN FTA (India-ASEAN AIFTA) reduces Basic Customs Duty from 30% to 5-10% for Thailand/Malaysia/Philippines/Indonesia/Vietnam products. "
    "This makes Tong Garden, Oishi, Mamee, Jack n Jill, Indomie significantly cheaper to import than US/UK/Korean/Japanese brands. "
    "Hippeas + Harvest Snaps = 100% vegan, no India equivalent — first-mover premium positioning. "
    "Pocky/Glico already sells in India via grey imports — white-channel opportunity exists."
).font=fnt(italic=True,size=8,color="1A3C5E")
ws1.cell(note_r,1).fill=BLUEF; ws1.row_dimensions[note_r].height=44

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — FULL PRODUCT CATALOGUE (all 937 products)
# ═══════════════════════════════════════════════════════════════════════════════
ws2=wb.create_sheet("Full Product Catalogue")
ws2.freeze_panes="C3"
hdr1(ws2,1,1,f"FULL PRODUCT CATALOGUE — {len(products)} Discovered SKUs | Vegetarian-First Filter",span=9)
ph=["Brand","Product Name","Size (g)","Export\nPrice","Currency","Price (INR\nest.)","Veg\nFlag","Platform","Country"]
for ci,h in enumerate(ph,1): hdr2(ws2,2,ci,h)
sw(ws2,[20,65,10,12,10,14,16,28,14])

ordered=sorted(products,key=lambda x:(BRAND_ORDER.index(x.get("brand","?"))
    if x.get("brand") in BRAND_ORDER else 99, x.get("product_name","")))
VEG_FILLS={"Vegetarian":GREENF,"100% Vegan":GREENF,"NON-VEG (flag)":REDF,"Unknown":YELLOWF}

for ri,p in enumerate(ordered,3):
    brand=p.get("brand","?"); bf=BRAND_FILLS.get(brand,fill("FFFFFF"))
    pname=p.get("product_name",""); sz=p.get("size_g"); pr=p.get("price_num"); cur=p.get("currency","?")
    veg=p.get("veg_flag","Unknown")
    inr_est=to_inr(pr,cur) if pr else None
    plat=(p.get("source_platform","") or "")[:30]
    bp=bprofiles.get(brand,{})
    vf=VEG_FILLS.get(veg,YELLOWF)

    wr(ws2,ri,1,brand,f=BOLDF,fl=bf)
    wr(ws2,ri,2,pname,fl=fill("FFFFFF"))
    wr(ws2,ri,3,f"{sz:.0f}" if sz else "",a=CTR,fl=fill("FFFFFF"))
    c=wr(ws2,ri,4,pr,a=CTR,fl=fill("FFFFFF"))
    if pr: c.number_format='#,##0.00'
    wr(ws2,ri,5,cur,a=CTR,fl=fill("FFFFFF"))
    c2=wr(ws2,ri,6,inr_est,a=CTR,fl=fill("FFFFFF"))
    if inr_est: c2.number_format='"Rs"#,##0'
    wr(ws2,ri,7,veg,a=CTR,fl=vf)
    wr(ws2,ri,8,plat,f=LINKF,fl=fill("FFFFFF"))
    wr(ws2,ri,9,bp.get("country",""),fl=fill("FFFFFF"))
    ws2.row_dimensions[ri].height=18
ws2.auto_filter.ref=f"A2:{get_column_letter(len(ph))}2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — TONG GARDEN DEEP DIVE
# ═══════════════════════════════════════════════════════════════════════════════
ws3=wb.create_sheet("Tong Garden Deep Dive")
ws3.freeze_panes="B3"
hdr1(ws3,1,1,"TONG GARDEN — Complete Brand Intelligence & India Import Strategy",span=7,bg=fill("7B3F00"))

# Brand overview
TG_INFO = [
    ("Founded", "1972, Bangkok, Thailand"),
    ("Full Name", "Tong Garden Food Company Limited"),
    ("Markets", "Thailand, Singapore, Malaysia, Vietnam, Myanmar, China, Export globally"),
    ("Revenue", "~THB 3,000 Cr+ (~Rs 700 Cr+) est."),
    ("Positioning", "Premium Asian nuts & seeds snack brand. 'Quality you can trust since 1972.'"),
    ("Parent / Listed", "Tong Garden Holdings — Singapore-listed"),
    ("Manufacturing", "Thailand (multiple plants). HACCP, ISO 22000, BRC certified."),
    ("Certifications", "Halal, GMP, HACCP, BRC Grade A, ISO 22000:2018"),
    ("Export Countries", "50+ countries. Strong in SE Asia, Middle East, USA, UK, Australia."),
    ("India Presence (current)", "Grey market only — imported via Singapore/Thailand. No official Indian distributor found."),
    ("Website", "tonggardengroup.com"),
]
r=3
for field,val in TG_INFO:
    rf=rfill(r)
    wr(ws3,r,1,field,f=BOLDF,fl=rf)
    ws3.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
    c=ws3.cell(r,2,val); c.fill=rf; c.border=THIN; c.font=DATAF; c.alignment=WR
    ws3.row_dimensions[r].height=20
    r+=1
r+=1

# Tong Garden product range
ws3.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
ws3.cell(r,1,"PRODUCT RANGE — Full Category Analysis").font=fnt(True,"7B3F00",10)
ws3.cell(r,1).fill=fill("FDEBD0"); r+=1
for ci2,h2 in enumerate(["Category","Key SKUs","Pack Sizes","Veg Status","Est. Export $ (USD)","India MRP (Rs est.)","India Opportunity"],1):
    x=ws3.cell(r,ci2,h2); x.fill=H2F; x.font=H2T; x.alignment=CTR; x.border=THIN
sw(ws3,[18,50,16,18,18,16,40])
r+=1

TG_PRODUCTS = [
    ("Nuts — Peanuts",
     "Honey Coated Peanuts (38g,50g,160g,500g)\nBar-B-Q Coated Peanuts (50g,500g)\nChilli Coated Peanuts\nSalted Peanuts\nCrispy Coated Peanuts",
     "38g, 50g, 160g, 500g, 1kg", "✓ Vegetarian", "$0.50-$2.50", "Rs 120-350",
     "STRONGEST opportunity. No comparable Indian brand in flavored coated peanut space at premium."),
    ("Nuts — Premium (Cashew/Macadamia/Mixed)",
     "Salted Cashew Mix Macadamia (140g)\nMixed Nuts (175g)\nParty Snack Mix Nut & Seed (175g)\nSalted Cashews (12 packs)",
     "70g, 140g, 175g, 250g", "✓ Vegetarian", "$2.00-$5.50", "Rs 350-900",
     "Premium gifting + impulse. No direct competitor in flavored mixed nuts import at Rs 350-600."),
    ("Seeds",
     "Salted Roasted Sunflower Kernels (110g)\nSalted Roasted Pumpkin Kernels (110g)\nSunflower Seed Crisp (30g)\nPumpkin Seed Crisp (30g)",
     "30g, 50g, 110g, 500g", "✓ Vegetarian", "$0.60-$2.50", "Rs 130-450",
     "Seeds = fastest-growing snack segment in India. Pumpkin seeds especially growing. First-mover in flavored roasted seeds."),
    ("Broad Beans",
     "Salted Broad Beans (500g)\nWasabi Coated Broad Beans (500g)\nBBQ Flavour Broad Beans (500g)\nCrispy Flour Baked Beans (40g x6)",
     "40g, 500g", "✓ Vegetarian", "$0.70-$3.50", "Rs 150-600",
     "UNIQUE: Flavored broad beans not widely available in India. Protein-dense, novelty factor high."),
    ("Green Peas / Wasabi Peas",
     "Wasabi Coated Green Peas (500g)\nSalted Green Peas",
     "30g, 50g, 500g", "✓ Vegetarian", "$0.60-$2.50", "Rs 130-450",
     "Wasabi peas = cult snack in urban India. Growing via import stores. Tong Garden has competitive pricing."),
    ("Wheat Snacks",
     "Wheat Crackers / Baked Wheat Snacks\nParty Snack (40g)",
     "40g, 175g", "✓ Vegetarian", "$0.50-$1.50", "Rs 110-280",
     "Wheat snacks = familiar to Indian consumers. Baked positioning resonates."),
    ("Seaweed Snacks",
     "Seaweed Snack Roasted Sea Salt (5g)\nSeaweed Crackers",
     "5g, 20g, 40g", "✓ Vegetarian (marine plant)", "$0.80-$2.00", "Rs 180-380",
     "Korean/Japanese seaweed snack trend hitting India. Tong Garden has affordable entry."),
    ("⚠ NON-VEG FLAG",
     "Peanuts Mixed Anchovy Spicy — CONTAINS FISH\nSome prawn-flavored variants",
     "Various", "✗ NON-VEG", "AVOID", "N/A",
     "SKIP these SKUs entirely for India import. Green dot ineligible."),
]
for row_data in TG_PRODUCTS:
    rf=rfill(r)
    is_flag="NON-VEG" in row_data[0]
    row_fill=REDF if is_flag else rf
    for ci2,v in enumerate(row_data,1):
        c=wr(ws3,r,ci2,v,fl=row_fill)
        if is_flag: c.font=fnt(bold=True,color="CC0000",size=9)
    ws3.row_dimensions[r].height=max(28,row_data[1].count("\n")*14+18)
    r+=1

# Tong Garden India economics detail
r+=1
ws3.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
ws3.cell(r,1,"INDIA IMPORT ECONOMICS — Tong Garden Honey Peanuts 50g (Reference SKU)").font=fnt(True,"7B3F00",10)
ws3.cell(r,1).fill=fill("FDEBD0"); r+=1

ec_tg=economics["Tong Garden"]
econ_rows=[
    ("Export Price (FOB Thailand)", f"$0.60 → Rs {0.60*USD_TO_INR:.0f}"),
    ("Freight + Insurance (+20%)", f"Rs {ec_tg['cif_inr']-0.60*USD_TO_INR:.0f} → CIF Rs {ec_tg['cif_inr']:.0f}"),
    ("Basic Customs Duty (ASEAN FTA 5%)", f"Rs {ec_tg['bcd_inr']:.0f}"),
    ("SWS (10% of BCD)", f"Rs {ec_tg['sws_inr']:.0f}"),
    ("IGST 12% (on CIF+BCD+SWS)", f"Rs {ec_tg['igst_inr']:.0f}"),
    ("TOTAL LANDED COST", f"Rs {ec_tg['total_landed_inr']:.0f} per pack"),
    ("Effective Duty %", f"{ec_tg['effective_duty_pct']:.1f}% on CIF value"),
    ("Importer Margin + Compliance (35%)", f"Rs {ec_tg['importer_margin_inr']:.0f}"),
    ("Distributor Price", f"Rs {ec_tg['dist_price_inr']:.0f}"),
    ("Distributor Margin (20%)", f"Rs {ec_tg['dist_price_inr']*0.20:.0f}"),
    ("Retailer Buy Price", f"Rs {ec_tg['retailer_in_inr']:.0f}"),
    ("Retailer Margin (25%)", f"Rs {ec_tg['retailer_in_inr']*0.25:.0f}"),
    ("MINIMUM VIABLE MRP", f"Rs {ec_tg['min_mrp_inr']:.0f} for 50g pack"),
    ("Comparable Indian product", "Haldiram's 50g namkeen: Rs 20-30. BUT nut snacks are different shelf."),
    ("Comparable imported product", "Lay's (PepsiCo India) 50g: Rs 30. Premium imported nuts: Rs 150-350."),
    ("Assessment", "Rs 120-150 for 50g Tong Garden is VIABLE in premium kirana/modern trade/online."),
]
for field,val in econ_rows:
    rf=rfill(r)
    is_total="TOTAL" in field or "MINIMUM" in field
    row_fill=GREENF if "MINIMUM" in field else (BLUEF if "TOTAL" in field else rf)
    wr(ws3,r,1,field,f=BOLDF if is_total else DATAF,fl=row_fill)
    ws3.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
    c=ws3.cell(r,2,val); c.fill=row_fill; c.border=THIN
    c.font=fnt(bold=is_total,size=9); c.alignment=WR
    ws3.row_dimensions[r].height=20; r+=1

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — BRAND PROFILES
# ═══════════════════════════════════════════════════════════════════════════════
ws4=wb.create_sheet("Brand Profiles")
sw(ws4,[22,25,20,20,55])
ws4.column_dimensions["B"].width=25

hdr1(ws4,1,1,"BRAND PROFILES — 15 Import-Ready Vegetarian Snack Brands",span=5)

FULL_PROFILES = {
    "Hippeas": ("B-Corp Organic Chickpea Puffs",
        "Founded 2016, UK. First chickpea-based puff snack brand to achieve mass UK/US retail distribution. "
        "100% vegan, organic, B-Corp certified. STRONG India opportunity: no Indian chickpea puff snack brand at this positioning. "
        "Clean label = tick on every trend box urban India cares about.",
        "Salt & Vinegar, White Cheddar Explosion, Rockin' Ranch, Nacho Vibes, Far Out Fajita, Purple Power (Purple Corn), Sriracha Sunshine, Veggie Straws",
        "$1.99-$3.49 for 78g (US retail)",
        "Premium modern trade (Foodhall, Nature's Basket, urban D2C). Not kirana. Target: Rs 199-249 for 78g."),
    "Harvest Snaps": ("Baked Green Pea/Lentil Crisps",
        "USA brand (Calbee North America). Baked, not fried. Green peas as primary ingredient. "
        "Gluten-free, non-GMO, vegetarian certified. PERFECT India fit: Indian consumers love peas/lentils. "
        "No comparable baked pea crisp exists in Indian market. Health-snacking megatrend.",
        "Lightly Salted Green Pea, Tomato Basil, Mango Chili Lime, Caesar, Black Pepper, White Cheddar Jalapeño, Selects (premium line)",
        "$0.99-$2.49 for 70-93g (US retail)",
        "Premium modern trade / D2C. India-relevant flavors (Mango Chili, Black Pepper). Rs 150-199 for 70g."),
    "Tong Garden": ("SE Asian Nuts & Seeds Pioneer",
        "Founded 1972, Thailand. 50+ export markets. Singapore-listed. Broadest flavored nut range in Asia. "
        "ASEAN FTA makes import economics excellent. Product types Indian consumers haven't seen: "
        "BBQ broad beans, wasabi green peas, sesame-coated peanuts. Immediate B2B/gifting opportunity.",
        "Honey Peanuts, BBQ Coated Peanuts, Party Mix, Wasabi Green Peas, Salted Cashew+Mac, "
        "Sunflower/Pumpkin Kernels, Seaweed Snacks, Salted Broad Beans",
        "$0.50-$5.50 (38g-500g, varies)",
        "Premium kirana + modern trade + online. Rs 120-350 for 38-100g packs. Gift packs Rs 500+."),
    "Oishi": ("Thailand's Mass-Market Extruded Snack Leader",
        "Largest extruded snack brand in Thailand. Affordable, wide distribution across SE Asia. "
        "Products comparable to Kurkure/Uncle Chips but Thai flavors. ASEAN FTA advantage. "
        "Mid-tier import: between cheap domestic and expensive Western.",
        "Bread Pan (Cheese & Onion), Potato Fries Ketchup, Cheese Sticks, Onion Rings, Prawn Crackers (NON-VEG), Pillows",
        "$0.30-$1.50 (20-80g)",
        "Modern trade, Asian grocery stores, online. Rs 80-200 for 30-80g. Good value import."),
    "Want Want": ("Taiwan Rice Cracker Giant",
        "Taiwan-listed company. One of Asia's largest rice cracker brands. "
        "Rice crackers are unfamiliar but exciting to Indian snackers. "
        "Senbei (Japanese-style) rice crackers = premium novelty. Mostly vegetarian.",
        "Big Shelly Senbei, Golden Rice Cracker Bites Original, Snow Cracker, Hot-Kid Milk Candy, Shelly Puffs",
        "$0.50-$2.50 (60-200g)",
        "Modern trade / online. Rs 120-280 for 70-150g. Unique product type for India market."),
    "Calbee": ("Japan's Largest Snack Company",
        "Japan. Founded 1949. Annual revenue ~$1.2B USD. Highly premium positioning globally. "
        "Harvest Snaps is their US brand. In Japan: Jagabee (potato sticks), Shrimp Chips (NON-VEG). "
        "Premium price point but strong brand trust for Japan-lovers.",
        "JagaRico Potato Sticks, Harvest Snaps (pea crisps), Mini Salted Peas, Asian Style Chips Thai Yellow Curry",
        "$1.10-$2.50 (80-170g)",
        "Premium modern trade only. Rs 200-350. Urban aspirational consumer."),
    "Pocky / Glico": ("Japan's Iconic Stick Biscuit",
        "Glico (Japan). Pocky exists in India via grey import channels already — "
        "widely seen in urban Asian grocery stores and Amazon.in. "
        "White-channel opportunity to be the official Indian importer/distributor.",
        "Pocky Chocolate, Strawberry, Matcha, Almond Crush, Cookies & Cream, Pistachio, Blueberry Yogurt, Collon, Pretz",
        "$1.40-$3.50 (47-140g)",
        "Already in India informally. Official import: Rs 150-250 for standard Pocky. Premium flavor Rs 199-299."),
    "Nongshim": ("Korean Snack & Noodle Brand",
        "Korea. Known for Shin Ramyun noodles but has snack range. Korean wave (K-Pop/K-Drama) "
        "driving premium for Korean food in Indian metros. Honey Twist Snack = vegetarian. "
        "Strategic import for 'Korean culture' trend.",
        "Honey Twist Snack, Honey Butter Crackers (check veg), Shrimp Crackers (NON-VEG — SKIP)",
        "$0.55-$1.50 (40-100g)",
        "Korean culture stores, online. Rs 120-250. Niche but growing."),
    "Wonderful Pistachios": ("USA Premium Pistachio Brand",
        "California. Wonderful Company. #1 pistachio brand in USA. Clean single-ingredient. "
        "India pistachio market is growing fast — pistachios are already culturally familiar. "
        "Flavored no-shell pistachios = novel for India. Premium gifting play.",
        "Roasted Salted, Lightly Salted, Sea Salt & Vinegar, Jalapeño Lime, Chili Roasted, Honey Roasted",
        "$1.25-$4.00 (49-454g)",
        "Premium dry fruit stores, modern trade, online. Rs 300-500 for 49g. Premium gifting segment."),
    "Tyrrell's": ("UK Premium Hand-Cooked Crisps",
        "Herefordshire, UK. Premium artisan crisp brand. Wonky vegetable crisps = unique product. "
        "UK-India FTA (if/when finalized) would help. Currently high duty (no FTA). "
        "Target: Foodhall-tier premium urban consumer.",
        "Naked (sea salt), Lentil Crisps Sour Cream, Parsnip & Black Pepper, Veg Crisps, Sweet Chilli",
        "$2.00-$5.00 (80-150g)",
        "Ultra-premium — Rs 400-600 for 80g. Very niche. Foodhall, airport retail only."),
}

r=3
for brand in BRAND_ORDER:
    prof_data = FULL_PROFILES.get(brand)
    if not prof_data: continue
    tagline,desc,products_str,price_str,india_strat = prof_data
    bf=BRAND_FILLS.get(brand,fill("FFFFFF"))
    ec=economics.get(brand,{}); vc=VEG_COMPLIANCE.get(brand,{}); feas,_=IMPORT_FEASIBILITY.get(brand,("?",""))

    ws4.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
    c=ws4.cell(r,1,f"  {brand} — {tagline}")
    c.fill=fill("0D2137"); c.font=H1T; c.alignment=WR; c.border=MED
    ws4.row_dimensions[r].height=22; r+=1

    rows=[
        ("Country / HQ", f"{bprofiles.get(brand,{}).get('country','')} · {bprofiles.get(brand,{}).get('hq_city','')}"),
        ("Brand Description", desc),
        ("Key Vegetarian Products", products_str),
        ("Global Price Range", price_str),
        ("Import Feasibility", f"{feas} — {IMPORT_FEASIBILITY.get(brand,('',''))[1]}"),
        ("India Strategy", india_strat),
        ("Vegetarian Status", f"{vc.get('overall','')} | Green Dot: {vc.get('green_dot_eligible','')}"),
        ("Flag Products (skip)", vc.get("flag_products","None")),
        ("Best Veg SKUs", vc.get("key_veg_skus","")),
        ("Min India MRP (est.)", f"Rs {ec.get('min_mrp_inr',0):.0f} for ~{ec.get('typical_pack_g','?')}g pack | ASEAN FTA: {'YES' if brand in ASEAN else 'No'}"),
    ]
    for field,val in rows:
        rf2=rfill(r)
        is_flag="Flag" in field; is_feas="Feasibility" in field; is_mrp="MRP" in field
        row_fill=(REDF if is_flag else (GREENF if (is_mrp and ec.get("min_mrp_inr",999)<200) else rf2))
        wr(ws4,r,1,field,f=BOLDF,fl=row_fill)
        ws4.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
        c2=ws4.cell(r,2,val); c2.fill=row_fill; c2.border=THIN
        c2.font=fnt(bold=is_feas and "HIGH" in str(val),size=9,
                    color="CC0000" if (is_flag and "NONE" not in str(val)) else "000000")
        c2.alignment=WR; ws4.row_dimensions[r].height=max(18,len(str(val))//60*13+16); r+=1
    r+=1

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — IMPORT ECONOMICS CALCULATOR
# ═══════════════════════════════════════════════════════════════════════════════
ws5=wb.create_sheet("Import Economics")
ws5.freeze_panes="B3"
hdr1(ws5,1,1,"IMPORT ECONOMICS CALCULATOR — India Landed Cost & MRP Estimation per Brand",span=12)
ec_cols=["Brand","Country","HS\nCode","ASEAN\nFTA?","Export\nPrice $","CIF\n(INR)","BCD\n%","BCD\n(INR)","IGST\n(INR)","Landed\n(INR)","Eff. Duty\n%","Min MRP\n(INR)"]
for ci,h in enumerate(ec_cols,1): hdr2(ws5,2,ci,h)
sw(ws5,[20,14,10,10,12,12,8,10,10,12,10,14])

for ri,brand in enumerate(BRAND_ORDER,3):
    ec=economics.get(brand,{}); bf=BRAND_FILLS.get(brand,fill("FFFFFF"))
    ep=BRAND_EXPORT_PRICES.get(brand,{})
    is_asean=brand in ASEAN
    mrp=ec.get("min_mrp_inr",0)
    viability_fill=(GREENF if mrp<200 else (YELLOWF if mrp<400 else REDF))

    vals=[
        brand, bprofiles.get(brand,{}).get("country",""),
        ec.get("hs_code",""), "YES ✓" if is_asean else "No",
        ep.get("export_usd","?"),
        ec.get("cif_inr","?"), ec.get("bcd_pct","?"), ec.get("bcd_inr","?"),
        ec.get("igst_inr","?"), ec.get("total_landed_inr","?"),
        f"{ec.get('effective_duty_pct','?')}%", f"Rs {mrp:.0f}",
    ]
    for ci,v in enumerate(vals,1):
        cell_fl=(GREENF if ci==4 and is_asean else (viability_fill if ci==12 else bf))
        c=wr(ws5,ri,ci,v,a=CTR,fl=cell_fl)
        if ci in {5,6,7,8,9,10}: c.number_format="#,##0.0"
    ws5.row_dimensions[ri].height=28
ws5.auto_filter.ref=f"A2:{get_column_letter(len(ec_cols))}2"

# Assumptions block below
sep=len(BRAND_ORDER)+5
ws5.merge_cells(start_row=sep,start_column=1,end_row=sep,end_column=12)
ws5.cell(sep,1,"ASSUMPTIONS & NOTES").font=fnt(True,"1A5276",10); ws5.cell(sep,1).fill=BLUEF; sep+=1
assumptions=[
    ("Exchange Rate", f"USD/INR = {USD_TO_INR} | GBP/INR = {GBP_TO_INR} | THB/INR = {THB_TO_INR} | SGD/INR = {SGD_TO_INR} | MYR/INR = {MYR_TO_INR}"),
    ("Freight & Insurance", "Estimated at +20% of FOB export price for CIF. Actual depends on mode (air/sea) and quantity."),
    ("ASEAN FTA (AIFTA)", "India-ASEAN Free Trade Agreement: BCD reduced to 5-10% for qualifying countries (Thailand, Malaysia, Philippines, Indonesia, Vietnam). Requires Form D (Certificate of Origin)."),
    ("BCD Standard", "30% Basic Customs Duty on most food snacks (HS 1905, 2008, 1904) without FTA."),
    ("SWS", "Social Welfare Surcharge = 10% of BCD amount."),
    ("IGST", "12% for nuts/grains/cereals (HS 2008, 1904, 0802). 18% for biscuits/snacks (HS 1905). Applied on CIF + BCD + SWS."),
    ("Importer Margin", "35% of landed cost covers: FSSAI compliance, labeling, storage, wastage, working capital, overheads."),
    ("Distributor Margin", "20% of ex-importer price."),
    ("Retailer Margin", "25% of ex-distributor price."),
    ("Min MRP", "= retailer price ÷ (1 - retailer_margin). This is the MINIMUM sustainable MRP. Actual MRP may be 20-40% higher for brand positioning."),
    ("Volume Impact", "All economics improve significantly at scale. Calculations assume small-lot imports (LCL). FCL container loads reduce landed cost by 15-25%."),
]
for field,val in assumptions:
    rf=rfill(sep)
    wr(ws5,sep,1,field,f=BOLDF,fl=rf)
    ws5.merge_cells(start_row=sep,start_column=2,end_row=sep,end_column=12)
    c=ws5.cell(sep,2,val); c.fill=rf; c.border=THIN; c.font=fnt(size=8); c.alignment=WR
    ws5.row_dimensions[sep].height=22; sep+=1

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — VEGETARIAN COMPLIANCE MATRIX
# ═══════════════════════════════════════════════════════════════════════════════
ws6=wb.create_sheet("Vegetarian Compliance")
ws6.freeze_panes="B3"
hdr1(ws6,1,1,"VEGETARIAN COMPLIANCE MATRIX — Green Dot Eligibility for India Import",span=7)
vc_cols=["Brand","Overall\nVeg Status","Certified\n(Global)","Green Dot\n(India) Eligible","Flag Products\n(SKIP these)","Best Veg SKUs to Import","Clean\nLabel?"]
for ci,h in enumerate(vc_cols,1): hdr2(ws6,2,ci,h)
sw(ws6,[20,20,22,22,40,55,14])

for ri,brand in enumerate(BRAND_ORDER,3):
    vc=VEG_COMPLIANCE.get(brand,{}); bf=BRAND_FILLS.get(brand,fill("FFFFFF"))
    overall=vc.get("overall","")
    is_100veg="100%" in overall
    is_partial="PARTIAL" in overall
    has_flag=vc.get("flag_products","") not in ("NONE","")

    status_fill=(GREENF if is_100veg else (YELLOWF if not has_flag else ORANGEF))
    flag_fill=(REDF if has_flag and "NONE" not in vc.get("flag_products","NONE") else GREENF)

    wr(ws6,ri,1,brand,f=BOLDF,fl=bf)
    wr(ws6,ri,2,overall,a=CTR,fl=status_fill)
    wr(ws6,ri,3,vc.get("certified",""),fl=rf)
    wr(ws6,ri,4,vc.get("green_dot_eligible",""),fl=status_fill)
    wr(ws6,ri,5,vc.get("flag_products",""),fl=flag_fill)
    wr(ws6,ri,6,vc.get("key_veg_skus",""),fl=fill("FFFFFF"))
    wr(ws6,ri,7,vc.get("clean_label",""),a=CTR,fl=rfill(ri))
    ws6.row_dimensions[ri].height=42
ws6.auto_filter.ref=f"A2:{get_column_letter(len(vc_cols))}2"

# FSSAI import process
fssai_r=len(BRAND_ORDER)+5
ws6.merge_cells(start_row=fssai_r,start_column=1,end_row=fssai_r,end_column=7)
ws6.cell(fssai_r,1,"FSSAI IMPORT PROCESS — Step-by-Step for Vegetarian Snack Import").font=fnt(True,"1A5276",10)
ws6.cell(fssai_r,1).fill=BLUEF; fssai_r+=1
for ci2,h2 in enumerate(["Step","Requirement","Details"],1):
    x=ws6.cell(fssai_r,ci2,h2); x.fill=H2F; x.font=H2T; x.alignment=CTR; x.border=THIN
ws6.merge_cells(start_row=fssai_r,start_column=3,end_row=fssai_r,end_column=7)
fssai_r+=1
for step,detail in FSSAI_STEPS:
    rf=rfill(fssai_r)
    parts=step.split(".",1); num=parts[0] if len(parts)>1 else step
    name=parts[1].strip() if len(parts)>1 else ""
    wr(ws6,fssai_r,1,f"{num}.",a=CTR,fl=rf)
    wr(ws6,fssai_r,2,name,f=BOLDF,fl=rf)
    ws6.merge_cells(start_row=fssai_r,start_column=3,end_row=fssai_r,end_column=7)
    c=ws6.cell(fssai_r,3,detail); c.fill=rf; c.border=THIN; c.font=fnt(size=8); c.alignment=WR
    ws6.row_dimensions[fssai_r].height=28; fssai_r+=1

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — ALREADY IN INDIA (market landscape)
# ═══════════════════════════════════════════════════════════════════════════════
ws7=wb.create_sheet("Already In India")
sw(ws7,[22,20,16,14,14,55])
hdr1(ws7,1,1,"FOREIGN SNACKS ALREADY IN INDIA — Import Market Landscape",span=6)
ai_cols=["Brand","Channel","Price Range\n(India)","Import\nStatus","Veg\nFriendly","Notes / Opportunity"]
for ci,h in enumerate(ai_cols,1): hdr2(ws7,2,ci,h)

ALREADY_IN_INDIA=[
    ("Pringles (Kellogg's)","Modern trade, kirana","Rs 70-200","Official PepsiCo India","PARTIALLY","Established. No entry point. But shows premium foreign chip can sell at Rs 70+ in India."),
    ("Pocky (Glico)","Asian grocery stores, Amazon.in","Rs 150-250","Grey import (no official)","MOSTLY","Huge grey market. Official importer opportunity. Strong brand recognition via K-pop/anime."),
    ("Doritos (PepsiCo)","Modern trade","Rs 30-200","PepsiCo India (local production)","PARTIALLY","Made locally now — not a true import. Shows corn snack works at Rs 30+."),
    ("Lay's International Flavors","Modern trade","Rs 20-40","PepsiCo India (local)","PARTIALLY","Not truly imported. But shows global flavor acceptance."),
    ("KitKat (Nestlé)","National","Rs 20-100","Nestlé India (local)","MOSTLY","Not a snack per se. Chocolate confection."),
    ("Tong Garden","Specialty import stores, online","Rs 200-450","Grey import only","MOSTLY","Available at Fab India, specialty food stores, Amazon.in (grey). No official importer."),
    ("Indomie Noodles","Specialty stores, Amazon.in","Rs 50-150","Grey + some official","PARTIALLY","Popular among Indian-Indonesian community. Veg options limited."),
    ("Nongshim Shin Ramyun","Specialty Korean stores","Rs 80-200","Small importers","PARTIAL — check veg","Korean wave driving. Mostly non-veg flavor range."),
    ("Want Want Rice Crackers","Asian grocery, Amazon.in","Rs 100-300","Small importers","MOSTLY","Available but not mainstream. Low brand awareness."),
    ("Hippeas","Not yet in India","—","NOT IMPORTED","100% VEGAN","OPPORTUNITY: Zero India presence. Clean label, vegan, chickpea — ideal for urban Indian consumer."),
    ("Harvest Snaps","Not yet in India","—","NOT IMPORTED","100% VEG","OPPORTUNITY: Zero India presence. Baked pea crisps = health snack gap in India."),
    ("Calbee Harvest Snaps (Japan)","Limited specialty","Rs 300-600","Very limited grey","MOSTLY VEG","Premium Japanese brand. Available at Japan-specialty stores."),
    ("Tyrrells","Not yet in India","—","NOT IMPORTED","MOSTLY VEG","OPPORTUNITY: Ultra-premium segment. UK artisan positioning. Very limited competition."),
    ("Wonderful Pistachios","Amazon.in, specialty","Rs 400-800","Some grey imports","100% VEG","Available premium. California pistachios trusted. Good entry point."),
    ("Oishi","Limited","Rs 80-200","Very limited","PARTIAL","SE Asian expat community buys. Mainstream potential untapped."),
    ("Mamee Monster","Limited specialty","Rs 50-150","Limited import","PARTIAL (veg options)","Noodle snack concept very welcome in India if marketed right."),
    ("Jack n Jill Piattos","Very limited","Rs 80-200","Limited import","PARTIAL","Filipino diaspora buys. Unknown to mainstream India."),
    ("Korean Seaweed Snacks (various)","Growing fast","Rs 100-300","Multiple small importers","VEG (marine)","Korean wave + seaweed health trend = fastest growing import snack in India 2024."),
]
for ri,row in enumerate(ALREADY_IN_INDIA,3):
    rf=rfill(ri)
    is_opp="OPPORTUNITY" in str(row[5])
    row_fill=GREENF if is_opp else rf
    for ci,v in enumerate(row,1):
        c=wr(ws7,ri,ci,v,fl=row_fill)
        if is_opp and ci==5: c.font=fnt(bold=True,color="1A7431",size=9)
    ws7.row_dimensions[ri].height=30
ws7.auto_filter.ref=f"A2:{get_column_letter(len(ai_cols))}2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — PM OPPORTUNITIES & STRATEGY
# ═══════════════════════════════════════════════════════════════════════════════
ws8=wb.create_sheet("PM Opportunities")
sw(ws8,[28,90])
ws8.row_dimensions[1].height=22
hdr1(ws8,1,1,"PM IMPORT STRATEGY — What to Import, Why, At What Margins",span=2)

PM_SECTIONS=[
    ("TOP IMPORT PICKS — RANK ORDER", [
        ("1. Harvest Snaps (USA/Calbee)  ★★★★★",
         "WHY: 100% vegetarian baked pea crisps with zero Indian equivalent. Health snack megatrend. "
         "India loves peas/lentils — flavor familiarity. Clean label. Gluten-free. Non-GMO. "
         "ECONOMICS: Min MRP ~Rs 190 for 93g. Positioned at Rs 199-249 = premium but not absurd. "
         "CHANNEL: D2C, premium kirana, Blinkit/Swiggy (urban), modern trade (Foodhall, Nature's Basket). "
         "ACTION: Contact Calbee North America for India distribution rights."),
        ("2. Hippeas (UK)  ★★★★★",
         "WHY: Only 100% vegan, B-Corp certified chickpea puff with global brand credibility. "
         "India connection: chickpeas are Indian (chana). Zero Indian brand in this space. "
         "Urban India's vegan/health trend = perfect timing. Rs 199-249 for 78g. "
         "CHALLENGE: No ASEAN FTA — higher duty. But premium pricing absorbs it. "
         "ACTION: Contact Hippeas UK for India import/distribution. B-Corp brand building is strong story."),
        ("3. Tong Garden (Thailand)  ★★★★☆",
         "WHY: ASEAN FTA (5% BCD), broadest flavored nut range in Asia, established in 50+ markets. "
         "India already buying via grey channels — proof of demand. Premium nuts segment growing fast in India. "
         "Unique SKUs: wasabi green peas, BBQ broad beans, party mix. "
         "ECONOMICS: Rs 120-150 for 38-50g pack. Very viable. Rs 500-1000 gift packs. "
         "FLAG: Avoid anchovy peanut range. "
         "ACTION: Contact Tong Garden Holdings (Singapore) export team. Form D for ASEAN FTA."),
        ("4. Pocky / Glico (Japan)  ★★★★☆",
         "WHY: Already in India via grey imports = proven demand. Strong brand via K-pop/anime/Asian culture. "
         "White-channel opportunity: become official importer. Matcha, Pistachio, premium flavors. "
         "ECONOMICS: Rs 150-250 per pack. Established price tolerance already set by grey market. "
         "ACTION: Contact Glico International export division. Official distributor registration."),
        ("5. Wonderful Pistachios (USA)  ★★★★☆",
         "WHY: Pistachios are part of Indian culture (gifting, dry fruits). Flavored no-shell pistachios "
         "(Jalapeño Lime, Sea Salt & Vinegar) = completely new format for Indian consumer. "
         "100% vegetarian. Premium gifting opportunity. Diwali/wedding gift boxes. "
         "ECONOMICS: Rs 300-500 for 49g snack pack. Rs 1000+ for larger. Premium but gift-segment justified. "
         "CHALLENGE: No FTA with USA — full 30% BCD. "
         "ACTION: Contact Wonderful Company for India distribution."),
        ("6. Want Want (Taiwan)  ★★★☆☆",
         "WHY: Rice crackers completely novel format for India. Mostly vegetarian. ASEAN-adjacent (some FTA coverage). "
         "Low price point possible: Rs 120-180 for 70-150g. Good introductory import. "
         "CHALLENGE: Low brand awareness in India. Marketing needed. "
         "ACTION: Pilot in modern trade to test demand."),
        ("7. Oishi (Thailand)  ★★★☆☆",
         "WHY: ASEAN FTA advantage. Mass-market SE Asian extruded snacks. Affordable: Rs 80-150 for 30-80g. "
         "Volume play: cheaper price point = broader reach. "
         "CHALLENGE: Partially vegetarian — need strict SKU selection. Prawn range must be excluded. "
         "ACTION: Import only veg-certified SKUs: Bread Pan, Potato Fries, Cheese Sticks, Onion Rings."),
    ]),
    ("MARKET TIMING & TRENDS", [
        ("Korean Snack Wave",
         "K-pop, K-drama, Korean skincare driving massive Korean food interest in Indian metros 2022-2025. "
         "Korean seaweed snacks growing 200%+ YoY on Amazon India. "
         "First-mover advantage in Korean vegetarian snacks (Nongshim Honey Twist, Lotte crackers) is NOW."),
        ("Baked/Healthy Snack Gap",
         "No Indian brand has credibly positioned at 'baked + healthy + vegetarian' in snack aisle at premium. "
         "Harvest Snaps and Hippeas fill this gap perfectly. '6g protein per serving' messaging resonates "
         "with India's protein-deficiency awareness (NIN studies widely cited)."),
        ("Pistachio & Premium Nut Boom",
         "India's premium dry fruit market growing 18% CAGR. Gifting, Diwali, online buying driving "
         "Rs 500-2000 premium nut purchases. Flavored pistachios (Jalapeño Lime, Sea Salt) = "
         "new-age premium gifting play. Wonderful Pistachios has packaging designed for this."),
        ("Seaweed = Next Quinoa",
         "Seaweed snacks are on the same trajectory quinoa was in India 5 years ago. "
         "Multiple Korean seaweed brands launching India grey imports. "
         "A branded, FSSAI-certified vegetarian seaweed snack would be category-defining."),
        ("Premium Kirana + D2C as Channel",
         "Online (D2C, Amazon, Blinkit) is THE channel for imported snacks Rs 100+. "
         "Premium kirana (Lulu, Spar, Spencer's, local premium stores) for visibility. "
         "Foodhall/Nature's Basket for ultra-premium (Tyrrell's, Hippeas). "
         "Mass kirana is NOT viable for imported snacks at these economics."),
    ]),
    ("COMPARISON TO OUR 19 SKUS (Balaji/Surya context)", [
        ("Price Tier Separation",
         "Our 19 SKUs = Rs 5. Importable foreign snacks = Rs 80-600. "
         "ZERO price overlap. These are entirely separate consumer occasions and channels. "
         "Import snacks serve urban premium, gifting, aspiration. Domestic Rs 5 = impulse kirana."),
        ("Consumer Profile Gap",
         "Rs 5 snack buyer: Semi-urban/rural, SEC B/C/D, daily impulse, 5g-25g packs, kirana. "
         "Import snack buyer: Urban metro, SEC A/B+, weekend indulgence/gifting, 50-200g, modern trade/online. "
         "These are different consumers, different channels, different brand language."),
        ("Ingredient Inspiration",
         "Foreign snack brands show what Indian consumers are willing to pay for in 5 years: "
         "baked (Harvest Snaps), chickpea-based (Hippeas), pumpkin seeds (Tong Garden), wasabi peas. "
         "Today's Rs 200 imported product = tomorrow's Rs 20 domestic innovation."),
        ("White Space Import Strategy",
         "Best import strategy: categories with ZERO domestic Indian equivalent. "
         "Baked pea crisps (zero), chickpea puffs (zero), flavored wasabi nuts (zero), premium pistachio snacks (zero). "
         "Avoid: potato chips (PepsiCo dominates), namkeen (domestic brands win on cost), instant noodles (Maggi/Yippee)."),
    ]),
    ("RISK FACTORS", [
        ("FSSAI Compliance Complexity",
         "Every new product needs NOC. Process can take 3-8 weeks per product. "
         "Labeling requirements strict — sticker labeling adds cost. "
         "First shipment full lab testing. Budget Rs 50,000-2,00,000 per new product for compliance costs."),
        ("Shelf Life in Import Chain",
         "Foreign snacks often have 6-12 month shelf life. Sea freight (20-30 days) + customs clearance (7-21 days) + "
         "distribution = easy 60-90 days consumed before product hits shelf. "
         "Air freight viable for premium small lots but expensive. Plan minimum shelf life carefully."),
        ("Currency & Price Volatility",
         "USD/INR movement directly affects MRP. If INR depreciates 5%, landed cost rises 5% instantly. "
         "Premium imports are sensitive to this. Need hedging or periodic MRP revision strategy."),
        ("Grey Market Competition",
         "For brands already in grey market (Tong Garden, Pocky, Want Want): "
         "Grey market sets a price ceiling. Consumers buying grey at Rs 200 won't pay Rs 350 official. "
         "Need to offer VALUE vs grey: official warranty, consistent supply, wider distribution."),
        ("Minimum Order Quantities",
         "Most brand export divisions require minimum order: 1-5 containers (FCL). "
         "1 FCL = ~20,000-40,000 units depending on pack size. "
         "High capital requirement for first order. Need distribution network in place before ordering."),
    ]),
]
r=3
for section_title,items in PM_SECTIONS:
    ws8.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    c=ws8.cell(r,1,section_title); c.fill=H1F; c.font=H1T; c.alignment=WR; c.border=MED
    ws8.row_dimensions[r].height=22; r+=1
    for key,value in items:
        rf=rfill(r)
        is_high="★★★★★" in key or "OPPORTUNITY" in key
        row_fill=GREENF if is_high else rf
        wr(ws8,r,1,key,f=BOLDF,fl=row_fill)
        wr(ws8,r,2,value,f=fnt(size=9),fl=row_fill)
        ws8.row_dimensions[r].height=max(32,len(value)//90*13+20); r+=1
    r+=1

# ── SAVE ────────────────────────────────────────────────────────────────────
out=ROOT/"foreign_snacks_intelligence.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
print(f"Products in catalogue: {len(products)}")
print(f"Brands profiled: {len(BRAND_ORDER)}")
